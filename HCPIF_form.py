# -*- coding: utf-8 -*-
"""
HCPIF extractor R2.7.4
Changes vs R2.7.3:

  FIX 1 – SLE OID incorrect matches
    add_sle_oid_from_legal_name now passes the HCPIF Country ISO2 into the Oracle
    hotel-name lookup so that deduplication is done on (HOTEL_NAME_MATCH, COUNTRY_ABBR)
    instead of name alone. This prevents a hotel whose legal name exists in multiple
    countries from being matched to the wrong record.

  FIX 2 – Currency "euro" (and similar words) resolved to None
    A CURRENCY_WORD_MAP is applied before the KNOWN_CURRENCIES validity check, so
    free-text values such as "euro", "US dollar", "pound" etc. are converted to their
    3-letter ISO codes. The DEFAULT_CURRENCY_BY_COUNTRY fallback was already correct
    but is now a genuine last resort rather than a silent workaround.

  FIX 3 – New "NO OID - EC" review tab
    After Oracle enrichment, rows where BOTH Found Hotel OID and SLE OID are blank AND
    the business model (TAI_BM) is NOT "GROUP" (i.e. Expedia Collect / Direct) are
    written to a separate sheet called "NO OID - EC". They are KEPT in the main
    Extraction sheet and WEBADI as before.

  FIX 4 – Remove Receipt Method from OC WEBADI
    col_current_rm write block removed from inject_into_webadi.
    "Current RM" is no longer pushed into column AB of the WEBADI sheet.
"""

# -------------------- DEFAULT PATHS --------------------
from pathlib import Path

DEFAULT_INPUT_DIR = r"C:\Users\josenjr\Downloads\HCPIFs"
DEFAULT_OUTPUT_FILE = r"C:\Users\josenjr\Downloads\HCPIFs\Output\HCPIF_extraction.xlsx"

DEFAULT_WEBADI_UPDATE_TEMPLATE = r"C:\Users\josenjr\Downloads\HCPIFs\OC WEBADI Update template.xlsm"
DEFAULT_WEBADI_ATTACH_SLE_TEMPLATE = r"C:\Users\josenjr\Downloads\HCPIFs\OC WEBADI Attach SLE template.xlsm"
DEFAULT_OID_CREATION_TEMPLATE = r"C:\Users\josenjr\Downloads\HCPIFs\OID Creation WEBADI.xlsm"

# -------------------- POPPLER / TESSERACT CONFIG --------------------

POPPLER_PATH = r"C:\Users\josenjr\OneDrive - Expedia Group\Desktop\Poppler\poppler-25.12.0\Library\bin"

TESSERACT_EXE = r"C:\Users\josenjr\OneDrive - Expedia Group\Desktop\Tessdata2\tesseract.exe"
TESSDATA_PREFIX = r"C:\Users\josenjr\OneDrive - Expedia Group\Desktop\Tessdata2\tessdata"

# -------------------- ORACLE CONFIG --------------------
ORACLE_CLIENT_DIR = r"C:\Users\josenjr\OneDrive - Expedia Group\Desktop\Oracle Instant Client\instantclient_23_0"
ORACLE_USERNAME = "xxxxxxxxxx"
ORACLE_PASSWORD = "XXXXXXXXXXX"  # WARNING: Hardcoded credential - should be moved to environment variable
ORACLE_DSN = "ashworaebsdb02-vip.datawarehouse.expecn.com:1526/ORAPRD_UI"

_ORACLE_CLIENT_INITIALIZED = False

def ensure_oracle_client():
    global _ORACLE_CLIENT_INITIALIZED
    if not _ORACLE_CLIENT_INITIALIZED:
        oracledb.init_oracle_client(lib_dir=ORACLE_CLIENT_DIR)
        _ORACLE_CLIENT_INITIALIZED = True

ORACLE_TCA_SQL_TEMPLATE = """
SELECT
    HCA.ACCOUNT_NUMBER AS ORACLE_ID,
    COUNT(HCA.ACCOUNT_NUMBER) OVER (PARTITION BY HCA.ACCOUNT_NUMBER) AS DUPLICATE_COUNT,
    HZP.PARTY_NUMBER,
    HCA.ATTRIBUTE1 AS EXPEDIA_ID,
    HZP.PARTY_NAME AS HOTEL_NAME,
    TAI.Attribute1 AS TAI_PC,
    TAI.Business_Model AS TAI_BM,
    SITEOU.NAME AS SITE_OU,
    HCA.CUSTOMER_CLASS_CODE,
    HCA.ATTRIBUTE3 AS SOURCE_SYSTEM,
    HCA.CREATION_DATE AS ACCOUNT_CREATION_DATE,
    SLE.PARTY_NAME AS SLE_NAME,
    SLE.SLE_OID,
    REL.START_DATE AS RELATIONSHIP_DATE,
    HZP.ADDRESS1,
    HZP.CITY,
    HZP.POSTAL_CODE,
    HZP.STATE,
    HZP.PROVINCE,
    HZP.COUNTRY AS COUNTRY_ABBR,
    TER.TERRITORY_SHORT_NAME AS COUNTRY_FULL_NAME,
    CONTACT.USER_NAME AS LAST_UPDATED_BY,
    CONTACT.FIRST_NAME,
    CONTACT.LAST_NAME,
    CONTACT.EMAIL_ADDRESS,
    CONTACT.PRIMARY_FLAG AS CONTACT_PRIMARY_FLAG,
    TAX.REGISTRATION_NUMBER,
    TAX.TAX,
    TAX.TAX_REGIME_CODE,
    TAX.REGISTRATION_STATUS_CODE,
    TAX.EFFECTIVE_FROM AS TAX_START_DATE,
    CURR.CURRENCY_CODE AS ORACLE_CURRENCY,
    RM.NAME AS RECEIPT_METHOD_NAME,
    RM.CURRENT_RM,
    RM.RM_START_DATE
FROM AR.HZ_CUST_ACCOUNTS HCA
LEFT JOIN AR.HZ_PARTIES HZP
    ON HZP.PARTY_ID = HCA.PARTY_ID
LEFT JOIN 
    (SELECT HR.OBJECT_ID,HR.SUBJECT_ID,HR.RELATIONSHIP_CODE,HR.START_DATE
        FROM AR.HZ_RELATIONSHIPS HR
            WHERE HR.RELATIONSHIP_CODE = 'LEGAL ENTITY/OWNER OF'
                AND HR.END_DATE LIKE '31-DEC-12') REL
                    ON REL.OBJECT_ID = HCA.PARTY_ID
LEFT JOIN
    (SELECT HZP2.PARTY_NAME,HZP2.PARTY_ID,HCA2.ACCOUNT_NUMBER AS SLE_OID
        FROM AR.HZ_PARTIES HZP2
            LEFT JOIN AR.HZ_CUST_ACCOUNTS HCA2
                ON HCA2.PARTY_ID = HZP2.PARTY_ID) SLE
                    ON SLE.PARTY_ID = REL.SUBJECT_ID
LEFT JOIN 
    (SELECT PTP.PARTY_ID,ZR.REGISTRATION_NUMBER,ZR.REGISTRATION_STATUS_CODE,ZR.EFFECTIVE_FROM,ZR.EFFECTIVE_TO,ZR.TAX_REGIME_CODE,ZR.TAX
        FROM APPS.ZX_PARTY_TAX_PROFILE PTP
            LEFT JOIN APPS.ZX_REGISTRATIONS ZR
                ON PTP.PARTY_TAX_PROFILE_ID = ZR.PARTY_TAX_PROFILE_ID
                    WHERE ZR.REGISTRATION_NUMBER IS NOT NULL
                        AND ZR.EFFECTIVE_TO IS NULL) TAX
                            ON TAX.PARTY_ID = HCA.PARTY_ID
LEFT JOIN
    (SELECT HCPA.CURRENCY_CODE,HCPA.CUST_ACCOUNT_ID
        FROM AR.HZ_CUST_PROFILE_AMTS HCPA
            WHERE HCPA.ATTRIBUTE1 IS NULL) CURR
                ON CURR.CUST_ACCOUNT_ID = HCA.CUST_ACCOUNT_ID
LEFT JOIN
    (SELECT CRM.CUSTOMER_ID,CRM.ATTRIBUTE1 AS CURRENT_RM,CRM.START_DATE AS RM_START_DATE,ARM.NAME
        FROM AR.RA_CUST_RECEIPT_METHODS CRM
            LEFT JOIN AR.AR_RECEIPT_METHODS ARM
                ON CRM.RECEIPT_METHOD_ID = ARM.RECEIPT_METHOD_ID
                    WHERE CRM.PRIMARY_FLAG = 'Y'
                        AND CRM.END_DATE IS NULL) RM
                            ON RM.CUSTOMER_ID = HCA.CUST_ACCOUNT_ID
LEFT JOIN
    (SELECT ACV.CUSTOMER_ID,ACV.FIRST_NAME,ACV.LAST_NAME,ACV.EMAIL_ADDRESS,CRV.PRIMARY_FLAG,FNDU.USER_NAME
        FROM APPS.AR_CONTACTS_V ACV
            LEFT JOIN APPS.AR_CONTACT_ROLES_V CRV
                ON ACV.CONTACT_ID = CRV.CONTACT_ID
                    LEFT JOIN APPS.FND_USER FNDU
                        ON FNDU.USER_ID = ACV.LAST_UPDATED_BY
                    WHERE STATUS = 'A'
                        AND CRV.PRIMARY_FLAG = 'Y') CONTACT
                            ON CONTACT.CUSTOMER_ID = HCA.CUST_ACCOUNT_ID
LEFT JOIN
    (SELECT TERRITORY_CODE,TERRITORY_SHORT_NAME
        FROM APPLSYS.FND_TERRITORIES_TL) TER
            ON TER.TERRITORY_CODE = HZP.COUNTRY
LEFT JOIN XXEXPD.XXEXPD_TCA_ADDITIONAL_INFO TAI
    ON HCA.Cust_Account_ID = TAI.partner_id
LEFT JOIN HZ_PARTY_SITES PS
    ON HZP.party_id = PS.party_id
LEFT JOIN HR_OPERATING_UNITS SITEOU  
    ON HCA.ORG_ID = SITEOU.ORGANIZATION_ID
WHERE HCA.CUSTOMER_CLASS_CODE = 'HOTEL'
  AND HCA.Attribute1 IN ({expedia_ids})
"""

# NOTE:
# This query below uses name normalization with REGEXP_REPLACE to improve match tolerance
# (special characters, '&' vs 'AND', repeated spaces, punctuation).
# It is slower for larger volumes.
# If performance becomes an issue, replace this logic with a simpler UPPER(TRIM(HZP.PARTY_NAME))
# comparison and accept stricter name matching.
ORACLE_HOTEL_BY_NAME_SQL_TEMPLATE = """
SELECT
    HCA.ACCOUNT_NUMBER AS ORACLE_ID,
    HCA.CUSTOMER_CLASS_CODE,
    HCA.ATTRIBUTE1 AS EXPEDIA_ID,
    HZP.PARTY_NAME AS HOTEL_NAME,
    HZP.COUNTRY AS COUNTRY_ABBR
FROM AR.HZ_CUST_ACCOUNTS HCA
JOIN AR.HZ_PARTIES HZP
    ON HZP.PARTY_ID = HCA.PARTY_ID
WHERE TRIM(
        REGEXP_REPLACE(
            REGEXP_REPLACE(
                REPLACE(UPPER(HZP.PARTY_NAME), '&', ' AND '),
                '[^A-Z0-9 ]',
                ' '
            ),
            ' +',
            ' '
        )
      ) IN ({hotel_names})
  AND HCA.CUSTOMER_CLASS_CODE = 'SLE'
"""

# -------------------- ENSURE / INSTALL DEPS --------------------
import sys
import subprocess
import importlib
import argparse
import copy
import re
import os
import time
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import oracledb

# -------------------- SHAREPOINT CONFIG (Graph + Azure App) --------------------
from msal import ConfidentialClientApplication  # msal already ensured by ensure_package

SP_TENANT_ID = "79efa2e2-5409-4b35-9714-ada0138ee76c"
SP_CLIENT_ID = "b516125e-8383-4640-a60a-f7e1524b871d"
SP_CLIENT_SECRET = os.environ["SP_CLIENT_SECRET"]  # defined via environment variable
SP_SCOPE = ["https://graph.microsoft.com/.default"]

SP_SITE_ID = "expediacorp.sharepoint.com,5e885651-f726-43f3-abad-1707182bd7be,3c6e8dfc-ae97-4e0a-a6e8-a58687688935"
SP_DRIVE_ID = "b!UVaIXib380OrrRcHGCvXvvyNbjyXrgpOpuilhodoiTWe-gXpkwfVSpMXueseaGTv"

SP_INPUT_FOLDER = "[Projects]/Bulk HCPIF Collab - CMD & Invoicing/PDFs Ready to Load"
SP_OUTPUT_BASE = "[Projects]/Bulk HCPIF Collab - CMD & Invoicing/Processed PDFs"
SP_ARCHIVE_FOLDER = "[Projects]/Bulk HCPIF Collab - CMD & Invoicing/Archived PDFs - Already Bulk Uploaded"
SP_SF_REPORT_FOLDER = "[Projects]/Bulk HCPIF Collab - CMD & Invoicing/SF Report"

# -------------------- COUNTRY / TAX MAPPING CONFIG --------------------
# Adjust the file name below to match the actual mapping spreadsheet name
COUNTRY_TAX_MAPPING_FILE = r"C:\Users\josenjr\Downloads\HCPIFs\Tax codes by country 4-30-26.xlsb"
COUNTRY_TAX_MAPPING_SHEET = "Sheet1"  # sheet tab name


def _sp_get_token() -> str:
    app = ConfidentialClientApplication(
        SP_CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{SP_TENANT_ID}",
        client_credential=SP_CLIENT_SECRET,
    )
    result = app.acquire_token_for_client(scopes=SP_SCOPE)
    if "access_token" not in result:
        raise RuntimeError(f"Error obtaining Graph token: {result}")
    return result["access_token"]


def _sp_graph_get(url: str) -> Dict:
    token = _sp_get_token()
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    resp.raise_for_status()
    return resp.json()


def _encode_path(path: str) -> str:
    """
    Encodes the folder/file path for use in Graph URL,
    preserving slashes (/) and encoding spaces, &, [, ], etc.
    """
    clean = path.strip("/")
    return quote(clean, safe="/")


def sp_list_children(folder_path: str) -> List[Dict]:
    """
    Lists items (files/folders) within a library folder.
    folder_path: relative path within the drive (e.g., "A/B/C").
    """
    encoded = _encode_path(folder_path)
    url = (
        f"https://graph.microsoft.com/v1.0/"
        f"sites/{SP_SITE_ID}/drives/{SP_DRIVE_ID}"
        f"/root:/{encoded}:/children"
    )
    print(f"[DEBUG] Graph URL children = {url}")
    data = _sp_graph_get(url)
    return data.get("value", [])


def sp_download_pdfs_from_folder(folder_path: str, dest_dir: Path) -> List[Path]:
    """
    Downloads all PDFs from a SharePoint folder to dest_dir.
    Returns the list of local files.
    """
    dest_dir.mkdir(parents=True, exist_ok=True)
    token = _sp_get_token()
    items = sp_list_children(folder_path)
    downloaded: List[Path] = []

    for item in items:
        name = item.get("name", "")
        if not name.lower().endswith(".pdf"):
            continue
        file_id = item["id"]
        url = (
            f"https://graph.microsoft.com/v1.0/"
            f"sites/{SP_SITE_ID}/drives/{SP_DRIVE_ID}/items/{file_id}/content"
        )
        resp = requests.get(url, headers={"Authorization": f"Bearer {token}"})
        resp.raise_for_status()
        local_path = dest_dir / name
        with open(local_path, "wb") as f:
            f.write(resp.content)
        downloaded.append(local_path)
    return downloaded


def sp_upload_file(local_path: Path, folder_path: str):
    """
    Uploads (or overwrites) a local file to a SharePoint folder.
    folder_path: relative path within the drive, e.g., "A/B/C".
    """
    token = _sp_get_token()
    file_name = local_path.name

    encoded_folder = _encode_path(folder_path)
    url = (
        f"https://graph.microsoft.com/v1.0/"
        f"sites/{SP_SITE_ID}/drives/{SP_DRIVE_ID}"
        f"/root:/{encoded_folder}/{file_name}:/content"
    )

    with open(local_path, "rb") as f:
        data = f.read()

    resp = requests.put(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/octet-stream",
        },
        data=data,
    )
    resp.raise_for_status()


def sp_get_most_recent_file(folder_path: str, extension: str = ".xlsm") -> Optional[Dict]:
    """
    Returns the metadata dict of the most recently modified file
    with the given extension inside the SharePoint folder.
    Returns None if no matching file is found.
    """
    items = sp_list_children(folder_path)
    candidates = [
        item for item in items
        if item.get("name", "").lower().endswith(extension.lower())
        and "file" in item  # only files, not sub-folders
    ]
    if not candidates:
        # Also try .xlsx if .xlsm not found
        candidates = [
            item for item in items
            if item.get("name", "").lower().endswith(".xlsx")
            and "file" in item
        ]
    if not candidates:
        return None

    # Sort by lastModifiedDateTime descending
    candidates.sort(
        key=lambda x: x.get("lastModifiedDateTime", ""),
        reverse=True,
    )
    return candidates[0]


def check_sf_report_rebill(
        df_extraction: "pd.DataFrame",
        dest_dir: Path,
) -> Tuple["pd.DataFrame", "pd.DataFrame"]:
    """
    Downloads the most recent SF Report from SharePoint, then cross-checks
    the 'Expedia ID' column in df_extraction against rows in the SF Report
    where Request Category (col I) == 'Rebill Request'.

    Returns:
        df_clean    – rows that did NOT match (safe to go to WEBADI)
        df_rebill   – rows that matched (should go to 'Review - Existing Rebill' tab)
                      with an extra 'SF_Case_Number' column and 'Comments' updated.
    """
    print("\n[INFO] Fetching most recent SF Report from SharePoint...")

    sf_item = sp_get_most_recent_file(SP_SF_REPORT_FOLDER, extension=".xlsm")
    if sf_item is None:
        sf_item = sp_get_most_recent_file(SP_SF_REPORT_FOLDER, extension=".xlsx")

    if sf_item is None:
        print("[WARN] No SF Report file found in SharePoint. Skipping rebill check.")
        return df_extraction, pd.DataFrame()

    file_id = sf_item["id"]
    file_name = sf_item.get("name", "SF_Report.xlsm")
    print(f"[INFO] Most recent SF Report: {file_name} "
          f"(modified: {sf_item.get('lastModifiedDateTime', 'unknown')})")

    token = _sp_get_token()
    url = (
        f"https://graph.microsoft.com/v1.0/"
        f"sites/{SP_SITE_ID}/drives/{SP_DRIVE_ID}/items/{file_id}/content"
    )
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    resp.raise_for_status()

    local_sf_path = dest_dir / file_name
    with open(local_sf_path, "wb") as f:
        f.write(resp.content)
    print(f"[INFO] SF Report downloaded to: {local_sf_path}")

    # Read the SF Report
    # Col B = Expedia Hotel ID (index 1), Col E = Case Number (index 4),
    # Col I = Request Category (index 8)
    # The file may have a header row; we read with header=0.
    try:
        # Try xlsm as xlsx (openpyxl handles both)
        df_sf = pd.read_excel(
            local_sf_path,
            sheet_name=0,          # first sheet
            dtype=str,             # everything as string to avoid float IDs
            engine="openpyxl",
        )
    except Exception as e:
        print(f"[WARN] Could not read SF Report with openpyxl ({e}). Skipping rebill check.")
        return df_extraction, pd.DataFrame()

    # Identify the relevant columns by position (B=index 1, E=index 4, I=index 8)
    # but also try by name for robustness
    def _col_by_name_or_pos(df_sf: "pd.DataFrame", names: list, pos: int) -> Optional[str]:
        for n in names:
            for c in df_sf.columns:
                if str(c).strip().lower() == n.lower():
                    return c
        # fall back to positional
        cols = list(df_sf.columns)
        if pos < len(cols):
            return cols[pos]
        return None

    col_eid      = _col_by_name_or_pos(df_sf, ["expedia hotel id", "expedia_hotel_id", "hotel id"], 1)
    col_case     = _col_by_name_or_pos(df_sf, ["case number", "case_number", "casenumber"], 4)
    col_category = _col_by_name_or_pos(df_sf, ["request category", "request_category", "type"], 8)

    if not col_eid or not col_category:
        print(f"[WARN] Could not identify required columns in SF Report "
              f"(eid_col={col_eid}, category_col={col_category}). "
              "Skipping rebill check.")
        return df_extraction, pd.DataFrame()

    print(f"[INFO] SF Report columns used → EID: '{col_eid}' | "
          f"Case: '{col_case}' | Category: '{col_category}'")

    # Filter SF Report: only "Rebill Request" rows
    mask_coo = df_sf[col_category].astype(str).str.strip().str.lower() == "rebill request"
    df_sf_coo = df_sf[mask_coo].copy()

    # Normalize EIDs in SF Report: strip whitespace, remove .0 suffix from floats, cast to str
    def _norm_eid(v) -> str:
        s = str(v).strip()
        # Remove trailing ".0" that comes from Excel reading integers as floats
        if s.endswith(".0"):
            s = s[:-2]
        return s

    df_sf_coo["_EID_NORM"] = df_sf_coo[col_eid].apply(_norm_eid)
    # Build a lookup: eid → case_number (take first match if duplicates)
    sf_coo_lookup: Dict[str, str] = {}
    for _, row in df_sf_coo.iterrows():
        eid = row["_EID_NORM"]
        if eid and eid not in ("", "nan", "None"):
            case = str(row[col_case]).strip() if col_case and pd.notna(row.get(col_case)) else ""
            if eid not in sf_coo_lookup:
                sf_coo_lookup[eid] = case

    if not sf_coo_lookup:
        print("[INFO] No 'Rebill Request' entries found in SF Report.")
        return df_extraction, pd.DataFrame()

    print(f"[INFO] {len(sf_coo_lookup)} EID(s) with 'Rebill Request' status in SF Report.")

    # Cross-check with extraction
    df_extraction = df_extraction.copy()
    df_extraction["_EID_NORM"] = (
        df_extraction["Expedia ID"]
        .astype(str)
        .str.strip()
        .apply(_norm_eid)
    )

    mask_rebill = df_extraction["_EID_NORM"].isin(sf_coo_lookup.keys())

    if not mask_rebill.any():
        print("[INFO] No extraction EIDs matched the SF Report 'Rebill Request' list.")
        df_extraction = df_extraction.drop(columns=["_EID_NORM"])
        return df_extraction, pd.DataFrame()

    matched_count = mask_rebill.sum()
    print(f"[INFO] {matched_count} extraction row(s) matched SF Report 'Rebill Request'. "
          "Moving to 'Review - Existing Rebill' tab.")

    # Build the rebill dataframe
    df_rebill = df_extraction[mask_rebill].copy()
    df_clean  = df_extraction[~mask_rebill].copy()

    # Add SF case number column
    df_rebill["SF_Case_Number"] = df_rebill["_EID_NORM"].map(sf_coo_lookup)

    # Overwrite / set Comments
    hold_comment = "Hold HCPIF - EID under Rebill process"
    if "Comments" in df_rebill.columns:
        df_rebill["Comments"] = hold_comment
    else:
        df_rebill.insert(len(df_rebill.columns), "Comments", hold_comment)

    # Clean up helper column
    df_rebill = df_rebill.drop(columns=["_EID_NORM"])
    df_clean  = df_clean.drop(columns=["_EID_NORM"])

    return df_clean, df_rebill


REQUIRED = [
    ("pdfplumber", "pdfplumber"),
    ("pandas", "pandas"),
    ("openpyxl", "openpyxl"),
    ("tqdm", "tqdm"),
    ("Unidecode", "unidecode"),
    ("pycountry", "pycountry"),
    ("pytesseract", "pytesseract"),
    ("pdf2image", "pdf2image"),
    ("msal", "msal"),
    ("requests", "requests"),
    ("pyxlsb", "pyxlsb")
]


def ensure_package(pip_name: str, import_name: str):
    try:
        return importlib.import_module(import_name)
    except ImportError:
        print(f" Installing {pip_name} ...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name])
        return importlib.import_module(import_name)


pdfplumber = ensure_package("pdfplumber", "pdfplumber")
pd = ensure_package("pandas", "pandas")
openpyxl = ensure_package("openpyxl", "openpyxl")
tqdm = ensure_package("tqdm", "tqdm")
unidecode_mod = ensure_package("Unidecode", "unidecode")
pytesseract = ensure_package("pytesseract", "pytesseract")
pdf2image = ensure_package("pdf2image", "pdf2image")
msal = ensure_package("msal", "msal")
requests = ensure_package("requests", "requests")
pyxlsb = ensure_package("pyxlsb", "pyxlsb")

from unidecode import unidecode
from pdf2image import convert_from_path

pytesseract.pytesseract.tesseract_cmd = TESSERACT_EXE
os.environ["TESSDATA_PREFIX"] = TESSDATA_PREFIX

try:
    pycountry = ensure_package("pycountry", "pycountry")
except Exception:
    pycountry = None

from openpyxl.packaging import manifest as _ox_manifest

for _ext in (".jpg", ".jpeg", ".JPG", ".JPEG"):
    try:
        _ox_manifest.mimetypes.add_type("image/jpeg", _ext)
    except Exception:
        pass

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from urllib.parse import quote

# -------------------- FIELD DEFINITIONS --------------------
FIELDS: List[Dict] = [
    {"col": "Today's date", "labels": [r"today[\\']s\s*date"]},
    {"col": "Effective Date of Change", "labels": [r"effective\s*date\s*of\s*change"]},
    {"col": "Expedia ID", "labels": [r"expedia\s*id"]},
    {"col": "Country", "labels": [r"country"]},
    {"col": "Currency", "labels": [r"currency", r"currency\s*code", r"curr\.?"]},
    {"col": "Legal Name", "labels": [r"legal\s*name"]},
    {"col": "Hotel Name", "labels": [r"hotel\s*name"]},
    {"col": "Address Line 1", "labels": [r"address\s*line\s*1"]},
    {"col": "City", "labels": [r"city"]},
    {
        "col": "State/Province",
        "labels": [
            r"state\s*/\s*province",
            r"state\s*province",
            r"state",
            r"province",
        ],
    },
    {"col": "Postal Code", "labels": [r"postal\s*code", r"zip\s*code"]},
    {"col": "Tax Registration Number", "labels": [r"tax\s*registration\s*number"]},
    {"col": "Tax Registration Status", "labels": [r"tax\s*registration\s*status"]},
    {"col": "First Name", "labels": [r"first\s*name"]},
    {"col": "Last Name", "labels": [r"last\s*name"]},
    {"col": "Email Address", "labels": [r"email\s*address", r"email"]},
    {"col": "Preferred Language", "labels": [r"preferred\s*language", r"language\s*preference"]},
]

SAME_LINE_PATTERN_TMPL = r"(?im)^\s*{label}\s*:?\s*(?P<val>[^\n\r]+?)\s*$"
LABEL_ONLY_PATTERN_TMPL = r"(?im)^\s*{label}\s*:?\s*$"

EMAIL_FALLBACK = re.compile(r"(?i)[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}")

DATE_PATTERNS = [
    "%d/%b/%Y",
    "%d/%B/%Y",
    "%d/%m/%Y",
    "%Y-%m-%d",
    "%d-%b-%Y",
    "%d-%m-%Y",
]

DATE_TOKEN_REGEX = re.compile(
    r"\b(\d{1,2}/[A-Za-z]{3}/\d{4}|\d{1,2}/\d{1,2}/\d{4}|"
    r"\d{4}-\d{2}-\d{2}|\d{1,2}-[A-Za-z]{3}-\d{4}|\d{1,2}-\d{1,2}-\d{4})\b"
)

MONTH3 = r"(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)"

UNICODE_LETTERS = r"A-Za-z-\u0370-\u03FF\u1F00-\u1FFF\u3040-\u30FF\u4E00-\u9FFF"

SIGNATURE_NOISE = re.compile(r"(?i)^\s*(signature|signed|signee|signer)\b")

HEADER_NOISE = re.compile(
    r"(?i)\b(city|state|province|billing|contact|information|signature|signed|signee|signer)\b"
)

POSTAL_LIKE = re.compile(r"(?i)^([A-Z]?\d[\dA-Z\- ]{2,}|[A-Z]\d[A-Z]\s*\d[A-Z]\d)$")

ALL_LABELS = [lab for f in FIELDS for lab in f["labels"]]
BARRIER_LABELS = ALL_LABELS + [r"signature", r"signed", r"signee", r"signer"]

NEXT_LABEL_BARRIER = re.compile(r"(?i)\b(?:" + "|".join(BARRIER_LABELS) + r")\b\s*:?")
LABEL_LINE_RE = re.compile(r"(?im)^\s*(?:" + "|".join(BARRIER_LABELS) + r")\s*:?\b")

KNOWN_CURRENCIES = {"AED", "AFN", "ALL", "AMD", "ANG", "AOA", "ARS", "AUD", "AWG", "AZN",
                    "BAM", "BBD", "BDT", "BGN", "BHD", "BIF", "BMD", "BND", "BOB", "BRL",
                    "BSD", "BTN", "BWP", "BYN", "BZD",
                    "CAD", "CDF", "CHF", "CLP", "CNY", "COP", "CRC", "CUP", "CVE", "CZK",
                    "DJF", "DKK", "DOP", "DZD",
                    "EGP", "ERN", "ETB", "EUR",
                    "FJD", "FKP",
                    "GBP", "GEL", "GHS", "GIP", "GMD", "GNF", "GTQ", "GYD",
                    "HKD", "HNL", "HTG", "HUF",
                    "IDR", "ILS", "INR", "IQD", "IRR", "ISK",
                    "JMD", "JOD", "JPY",
                    "KES", "KGS", "KHR", "KMF", "KPW", "KRW", "KWD", "KYD", "KZT",
                    "LAK", "LBP", "LKR", "LRD", "LSL", "LYD",
                    "MAD", "MDL", "MGA", "MKD", "MMK", "MNT", "MOP", "MRU", "MUR", "MVR",
                    "MWK", "MXN", "MYR", "MZN",
                    "NAD", "NGN", "NIO", "NOK", "NPR", "NZD",
                    "OMR",
                    "PAB", "PEN", "PGK", "PHP", "PKR", "PLN", "PYG",
                    "QAR",
                    "RON", "RSD", "RUB", "RWF",
                    "SAR", "SBD", "SCR", "SDG", "SEK", "SGD", "SHP", "SLE", "SLL", "SOS",
                    "SRD", "SSP", "STN", "SVC", "SYP", "SZL",
                    "THB", "TJS", "TMT", "TND", "TOP", "TRY", "TTD", "TWD", "TZS",
                    "UAH", "UGX", "USD", "UYU", "UZS",
                    "VED", "VES", "VND", "VUV",
                    "WST",
                    "XAF", "XCD", "XDR", "XOF", "XPF",
                    "YER",
                    "ZAR", "ZMW", "ZWL", }

DEFAULT_CURRENCY_BY_COUNTRY = {
    # Eurozone countries
    "GERMANY": "EUR", "DE": "EUR",
    "SPAIN": "EUR", "ES": "EUR",
    "ITALY": "EUR", "IT": "EUR",
    "AUSTRIA": "EUR", "AT": "EUR",
    "MONTENEGRO": "EUR", "ME": "EUR",
    "GREECE": "EUR", "GR": "EUR",
    "FRANCE": "EUR", "FR": "EUR",
    "PORTUGAL": "EUR", "PT": "EUR",
    "BELGIUM": "EUR", "BE": "EUR",
    "NETHERLANDS": "EUR", "NL": "EUR",
    "IRELAND": "EUR", "IE": "EUR",
    "LUXEMBOURG": "EUR", "LU": "EUR",
    "FINLAND": "EUR", "FI": "EUR",
    "SLOVAKIA": "EUR", "SK": "EUR",
    "SLOVENIA": "EUR", "SI": "EUR",
    "ESTONIA": "EUR", "EE": "EUR",
    "LATVIA": "EUR", "LV": "EUR",
    "LITHUANIA": "EUR", "LT": "EUR",
    "MALTA": "EUR", "MT": "EUR",
    "CYPRUS": "EUR", "CY": "EUR",
    "CROATIA": "EUR", "HR": "EUR",

    # Americas
    "MEXICO": "MXN", "MX": "MXN",
    "UNITED STATES": "USD", "USA": "USD", "US": "USD",
    "CANADA": "CAD", "CA": "CAD",
    "BRAZIL": "BRL", "BR": "BRL",
    "ARGENTINA": "ARS", "AR": "ARS",
    "CHILE": "CLP", "CL": "CLP",
    "COLOMBIA": "COP", "CO": "COP",
    "PERU": "PEN", "PE": "PEN",
    "COSTA RICA": "CRC", "CR": "CRC",
    "PANAMA": "USD", "PA": "USD",
    "DOMINICAN REPUBLIC": "DOP", "DO": "DOP",
    "JAMAICA": "JMD", "JM": "JMD",

    # Asia-Pacific
    "SWITZERLAND": "CHF", "CH": "CHF",
    "THAILAND": "THB", "TH": "THB",
    "AUSTRALIA": "AUD", "AU": "AUD",
    "HONG KONG": "HKD", "HK": "HKD",
    "SINGAPORE": "SGD", "SG": "SGD",
    "JAPAN": "JPY", "JP": "JPY",
    "CHINA": "CNY", "CN": "CNY",
    "INDIA": "INR", "IN": "INR",
    "MALAYSIA": "MYR", "MY": "MYR",
    "INDONESIA": "IDR", "ID": "IDR",
    "PHILIPPINES": "PHP", "PH": "PHP",
    "VIETNAM": "VND", "VN": "VND",
    "SOUTH KOREA": "KRW", "KR": "KRW",
    "KOREA": "KRW",
    "TAIWAN": "TWD", "TW": "TWD",
    "NEW ZEALAND": "NZD", "NZ": "NZD",
    "SRI LANKA": "LKR", "LK": "LKR",
    "MALDIVES": "MVR", "MV": "MVR",

    # Middle East
    "UNITED ARAB EMIRATES": "AED", "UAE": "AED", "AE": "AED",
    "SAUDI ARABIA": "SAR", "SA": "SAR",
    "ISRAEL": "ILS", "IL": "ILS",
    "TURKEY": "TRY", "TR": "TRY",
    "TURKIYE": "TRY",
    "EGYPT": "EGP", "EG": "EGP",
    "JORDAN": "JOD", "JO": "JOD",
    "QATAR": "QAR", "QA": "QAR",
    "BAHRAIN": "BHD", "BH": "BHD",
    "KUWAIT": "KWD", "KW": "KWD",
    "OMAN": "OMR", "OM": "OMR",

    # Europe (non-Euro)
    "UNITED KINGDOM": "GBP", "UK": "GBP", "GB": "GBP",
    "ENGLAND": "GBP", "SCOTLAND": "GBP", "WALES": "GBP",
    "NORWAY": "NOK", "NO": "NOK",
    "SWEDEN": "SEK", "SE": "SEK",
    "DENMARK": "DKK", "DK": "DKK",
    "POLAND": "PLN", "PL": "PLN",
    "CZECH REPUBLIC": "CZK", "CZ": "CZK",
    "CZECHIA": "CZK",
    "HUNGARY": "HUF", "HU": "HUF",
    "ROMANIA": "RON", "RO": "RON",
    "BULGARIA": "BGN", "BG": "BGN",
    "ICELAND": "ISK", "IS": "ISK",
    "SERBIA": "RSD", "RS": "RSD",
    "UKRAINE": "UAH", "UA": "UAH",
    "RUSSIA": "RUB", "RU": "RUB",

    # Africa
    "SOUTH AFRICA": "ZAR", "ZA": "ZAR",
    "MOROCCO": "MAD", "MA": "MAD",
    "KENYA": "KES", "KE": "KES",
    "NIGERIA": "NGN", "NG": "NGN",
    "GHANA": "GHS", "GH": "GHS",
    "TANZANIA": "TZS", "TZ": "TZS",
    "UGANDA": "UGX", "UG": "UGX",
    "MAURITIUS": "MUR", "MU": "MUR",
    "SEYCHELLES": "SCR", "SC": "SCR",
}

# FIX 2 – map free-text currency words to ISO codes BEFORE the KNOWN_CURRENCIES check
CURRENCY_WORD_MAP: Dict[str, str] = {
    "EURO": "EUR", "EUROS": "EUR", "EUR": "EUR",
    "US DOLLAR": "USD", "US DOLLARS": "USD", "DOLLAR": "USD", "DOLLARS": "USD",
    "POUND": "GBP", "POUNDS": "GBP", "POUND STERLING": "GBP", "STERLING": "GBP",
    "YEN": "JPY", "JAPANESE YEN": "JPY",
    "YUAN": "CNY", "RENMINBI": "CNY", "RMB": "CNY",
    "SWISS FRANC": "CHF", "FRANC": "CHF", "FRANCS": "CHF",
    "CANADIAN DOLLAR": "CAD", "CANADIAN DOLLARS": "CAD",
    "AUSTRALIAN DOLLAR": "AUD", "AUSTRALIAN DOLLARS": "AUD",
    "HONG KONG DOLLAR": "HKD",
    "MEXICAN PESO": "MXN", "PESO": "MXN",
    "REAL": "BRL", "REAIS": "BRL", "BRL": "BRL",
    "RUPEE": "INR", "RUPEES": "INR",
    "KRONA": "SEK", "KRONE": "NOK",
    "ZLOTY": "PLN",
    "FORINT": "HUF",
    "KORUNA": "CZK",
    "LEU": "RON",
    "DINAR": "RSD",
    "DIRHAM": "AED",
    "RIYAL": "SAR",
    "SHEKEL": "ILS",
    "BAHT": "THB",
    "RINGGIT": "MYR",
    "PESO ARGENTINO": "ARS",
    "CORONA": "SEK",
}

def normalize_currency_word(raw: str) -> Optional[str]:
    """
    Convert free-text currency names to ISO-4217 codes.
    Returns None if not recognised.
    """
    if not raw:
        return None
    s = raw.strip().upper()
    # Direct ISO match
    if re.fullmatch(r"[A-Z]{3}", s) and s in KNOWN_CURRENCIES:
        return s
    # Word-map lookup
    mapped = CURRENCY_WORD_MAP.get(s)
    if mapped:
        return mapped
    # Partial word match (e.g. "Euros" contained in key)
    for word, code in CURRENCY_WORD_MAP.items():
        if word in s or s in word:
            return code
    return None


def get_currency_from_country(country_name: Optional[str]) -> Optional[str]:
    """
    Gets the currency code for a country using multiple fallback strategies:
    1. Direct lookup in DEFAULT_CURRENCY_BY_COUNTRY (fast, pre-configured)
    2. ISO2 code lookup in DEFAULT_CURRENCY_BY_COUNTRY
    3. pycountry lookup (comprehensive, works for all countries)

    Returns the 3-letter ISO-4217 currency code or None if not found.
    """
    if not country_name:
        return None

    try:
        if pd.isna(country_name):
            return None
    except Exception:
        pass

    country_raw = str(country_name).strip()
    if not country_raw or country_raw.upper() == "<NA>":
        return None

    country_norm = country_raw.upper()

    # Strategy 1: Direct lookup by country name
    if country_norm in DEFAULT_CURRENCY_BY_COUNTRY:
        return DEFAULT_CURRENCY_BY_COUNTRY[country_norm]

    # Strategy 2: Lookup by ISO2 code
    try:
        iso2_country = to_iso2(country_raw)
        if iso2_country and iso2_country.upper() in DEFAULT_CURRENCY_BY_COUNTRY:
            return DEFAULT_CURRENCY_BY_COUNTRY[iso2_country.upper()]
    except Exception:
        pass

    # Strategy 3: Use pycountry to get currency (comprehensive fallback)
    if pycountry:
        try:
            # Try to find the country
            country_obj = None

            # Try direct lookup by name or code
            try:
                country_obj = pycountry.countries.lookup(country_raw)
            except Exception:
                pass

            # If we found a country, try to get its currency
            if country_obj:
                alpha_2 = getattr(country_obj, "alpha_2", None)
                if alpha_2:
                    # Try to find currency by country code
                    # Note: pycountry doesn't have direct country->currency mapping
                    # so we use a common ISO mapping
                    try:
                        # Import pycountry.currencies if available
                        import pycountry

                        # Common country code to currency mappings not in our dict
                        iso_currency_map = {
                            "AL": "ALL", "DZ": "DZD", "AO": "AOA", "AM": "AMD",
                            "AZ": "AZN", "BD": "BDT", "BY": "BYN", "BZ": "BZD",
                            "BO": "BOB", "BA": "BAM", "BW": "BWP", "BN": "BND",
                            "KH": "KHR", "CM": "XAF", "CV": "CVE", "TD": "XAF",
                            "CU": "CUP", "CD": "CDF", "DJ": "DJF", "EC": "USD",
                            "SV": "USD", "ER": "ERN", "ET": "ETB", "FJ": "FJD",
                            "GA": "XAF", "GM": "GMD", "GE": "GEL", "GN": "GNF",
                            "GY": "GYD", "HT": "HTG", "HN": "HNL", "IQ": "IQD",
                            "IR": "IRR", "KZ": "KZT", "KP": "KPW", "KG": "KGS",
                            "LA": "LAK", "LB": "LBP", "LR": "LRD", "LY": "LYD",
                            "MK": "MKD", "MG": "MGA", "MW": "MWK", "ML": "XOF",
                            "MR": "MRU", "MN": "MNT", "MZ": "MZN", "MM": "MMK",
                            "NA": "NAD", "NP": "NPR", "NI": "NIO", "NE": "XOF",
                            "PK": "PKR", "PG": "PGK", "PY": "PYG", "RW": "RWF",
                            "SN": "XOF", "SL": "SLL", "SO": "SOS", "SS": "SSP",
                            "SD": "SDG", "SR": "SRD", "SZ": "SZL", "SY": "SYP",
                            "TJ": "TJS", "TG": "XOF", "TT": "TTD", "TN": "TND",
                            "TM": "TMT", "UY": "UYU", "UZ": "UZS", "VU": "VUV",
                            "VE": "VES", "YE": "YER", "ZM": "ZMW", "ZW": "ZWL",
                        }

                        if alpha_2 in iso_currency_map:
                            return iso_currency_map[alpha_2]

                    except Exception:
                        pass
        except Exception:
            pass

    # If all strategies fail, return None
    return None


# -------------------- DATE / VALUE HELPERS --------------------
def is_valid_date_token(token: str) -> bool:
    token = token.strip()
    for fmt in DATE_PATTERNS:
        try:
            datetime.strptime(token, fmt)
            return True
        except ValueError:
            continue
    return False


def extract_first_date_token(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    for m in DATE_TOKEN_REGEX.finditer(s):
        token = m.group(0)
        if is_valid_date_token(token):
            return token
    return None


def extract_two_dates_from_lines(full_text: str):
    tokens = DATE_TOKEN_REGEX.findall(full_text or "")
    valids = [t for t in tokens if is_valid_date_token(t)]
    if len(valids) >= 2:
        return valids[0], valids[1]
    if len(valids) == 1:
        return valids[0], None
    return None, None


def contains_date_like(s: str) -> bool:
    if not s:
        return False
    if DATE_TOKEN_REGEX.search(s):
        return True
    if re.search(rf"\b{MONTH3}\b\s+\d{{1,2}},\s*\d{{4}}", s, re.IGNORECASE):
        return True
    if re.search(r"GMT\s*[+-]\d+", s, re.IGNORECASE):
        return True
    return False


def normalize_block_text(s: str) -> str:
    return (s or "").replace("\u00A0", " ").replace("\t", " ")


def cut_at_next_label(s: str) -> str:
    if not s:
        return s
    m = NEXT_LABEL_BARRIER.search(s)
    return s[: m.start()].strip() if m else s.strip()


def strip_signature_prefix(s: Optional[str]) -> Optional[str]:
    if not s:
        return s
    return (
            re.sub(r"(?i)^\s*(signature|signed|signee|signer)\s*:\s*", "", s).strip() or None
    )


def strip_parenthesized_dates(s: Optional[str]) -> Optional[str]:
    if not s:
        return s
    s = re.sub(
        rf"\(\s*{MONTH3}\s+\d{{1,2}},\s*\d{{4}}[^()]*\)", "", s, flags=re.IGNORECASE
    )
    s = re.sub(r"\([^0-9()]*\d[^()]*\)", "", s)
    return s.strip() or None


LEADING_JUNK_RE = re.compile(r"^[^0-9A-Za-z]+")


def strip_leading_junk(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    s = str(s).lstrip()
    s = LEADING_JUNK_RE.sub("", s)
    return s


def clean_extracted_value(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    v = normalize_block_text(s)
    v = cut_at_next_label(v)
    v = strip_signature_prefix(v) or v
    v = strip_parenthesized_dates(v) or v
    v = re.sub(r"\s{2,}", " ", v).strip()
    v = strip_leading_junk(v) or ""
    if not re.search(rf"[{UNICODE_LETTERS}0-9]", v):
        return None
    if contains_date_like(v):
        return None
    return v or None


def looks_like_postal(s: Optional[str]) -> bool:
    if not s:
        return False
    return bool(POSTAL_LIKE.match(s.strip()))


def sanitize_country(val: Optional[str], currency: Optional[str]) -> Optional[str]:
    if not val:
        return None
    v = clean_extracted_value(val)
    if not v:
        return None
    v_clean = v.strip()
    v_upper = v_clean.upper()
    if currency:
        v_clean = re.sub(
            rf"(?i)^\s*{re.escape(str(currency).strip())}\s+",
            "",
            v_clean,
        ).strip()
        v_upper = v_clean.upper()
    if re.fullmatch(r"[A-Za-z]{2,4}", v_clean):
        if currency and v_upper == (currency or "").strip().upper():
            return None
        return None
    if (
            HEADER_NOISE.search(v_clean)
            or re.search(r"\d", v_clean)
            or SIGNATURE_NOISE.match(v_clean)
            or contains_date_like(v_clean)
    ):
        return None
    return v_clean


def sanitize_state_and_postal(
        state_val: Optional[str], postal_val: Optional[str]
) -> Tuple[Optional[str], Optional[str]]:
    s = clean_extracted_value(state_val) if state_val else None
    p = clean_extracted_value(postal_val) if postal_val else None

    if s and looks_like_postal(s):
        if not p:
            p = s
        s = None

    if s and not re.search(rf"[{UNICODE_LETTERS}]", s) and len(s) <= 2:
        s = None

    if s and (contains_date_like(s) or SIGNATURE_NOISE.match(s)):
        s = None

    return s, p


def is_alpha_line(s: str) -> bool:
    s = s.strip()
    if not s or "@" in s:
        return False
    if contains_date_like(s) or SIGNATURE_NOISE.match(s):
        return False
    return re.fullmatch(rf"[{UNICODE_LETTERS} .'\-()]{2, 120}", s) is not None


def try_same_line_block(text: str, label_regex: str) -> Optional[str]:
    pattern = re.compile(
        SAME_LINE_PATTERN_TMPL.format(label=label_regex),
        flags=re.IGNORECASE | re.MULTILINE,
    )
    m = pattern.search(text)
    if m:
        cand = clean_extracted_value(m.group("val"))
        if cand and not contains_date_like(cand) and not SIGNATURE_NOISE.match(cand):
            return cand
    return None


def try_next_line_block(
        text: str, label_regex: str, all_label_variants: List[str]
) -> Optional[str]:
    label_only = re.compile(
        LABEL_ONLY_PATTERN_TMPL.format(label=label_regex),
        flags=re.IGNORECASE | re.MULTILINE,
    )
    m = label_only.search(text)
    if not m:
        return None
    tail = text[m.end():]
    start_of_label = re.compile(
        r"(?im)^\s*(?:" + "|".join(all_label_variants) + r")\s*:\b"
    )
    for ln in tail.splitlines():
        cand = clean_extracted_value(ln)
        if not cand:
            continue
        if start_of_label.match(cand):
            return None
        if HEADER_NOISE.search(cand) or SIGNATURE_NOISE.match(cand) or contains_date_like(
                cand
        ):
            continue
        return cand
    return None


def try_next_line_block_state(text: str, label_regex: str) -> Optional[str]:
    """
    Fallback específico para State/Province:
    - Se logo após o label vier outra linha que já é um label (Postal Code, Tax Registration Number,
      Signature, etc.), consideramos o campo vazio e NÃO descemos até linhas de assinatura.
    """
    label_only = re.compile(
        LABEL_ONLY_PATTERN_TMPL.format(label=label_regex),
        flags=re.IGNORECASE | re.MULTILINE,
    )
    m = label_only.search(text)
    if not m:
        return None

    tail = text[m.end():]

    for ln in tail.splitlines():
        # Se a linha já começa com outro label conhecido, paramos imediatamente
        if LABEL_LINE_RE.match(ln):
            return None

        cand = clean_extracted_value(ln)
        if not cand:
            continue
        if (
                HEADER_NOISE.search(cand)
                or SIGNATURE_NOISE.match(cand)
                or contains_date_like(cand)
        ):
            continue
        return cand

    return None


def words_from_page(page) -> List[Dict]:
    return page.extract_words(
        x_tolerance=3,
        y_tolerance=3,
        keep_blank_chars=False,
        use_text_flow=True,
    ) or []


def find_label_anchors(words: List[Dict], label_tokens: List[str]) -> List[Dict]:
    def _norm(s: str) -> str:
        s = (s or "").lower().replace("’", "'").strip()
        return s[:-1] if s.endswith(":") else s

    anchors = []
    n, m = len(words), len(label_tokens)
    norm_labels = [_norm(x) for x in label_tokens]
    for i in range(n - m + 1):
        if all(_norm(words[i + j]["text"]) == norm_labels[j] for j in range(m)):
            seq = words[i: i + m]
            anchors.append(
                {
                    "x0": min(w["x0"] for w in seq),
                    "x1": max(w["x1"] for w in seq),
                    "top": min(w["top"] for w in seq),
                    "bottom": max(w["bottom"] for w in seq),
                }
            )
    return anchors


def collect_tokens_on_same_line_right(
        words: List[Dict], anchor: Dict, y_tol: float = 6.0
):
    a_top, a_bot, a_x1 = anchor["top"], anchor["bottom"], anchor["x1"]
    line_tokens = []
    for w in words:
        if w["x0"] <= a_x1:
            continue
        if abs(w["top"] - a_top) <= y_tol or abs(w["bottom"] - a_bot) <= y_tol:
            line_tokens.append(w)
    line_tokens.sort(key=lambda k: k["x0"])
    return line_tokens


def collect_tokens_next_line_down(
        words, anchor, x_window=360.0, v_range=(0.5, 80.0)
):
    a_bot, a_x1 = anchor["bottom"], anchor["x1"]
    v_min, v_max = a_bot + v_range[0], a_bot + v_range[1]
    h_min, h_max = a_x1, a_x1 + x_window
    cand = [
        w
        for w in words
        if (v_min <= w["top"] <= v_max and h_min <= w["x0"] <= h_max)
    ]
    cand.sort(key=lambda k: (k["top"], k["x0"]))
    return cand


def join_tokens_until_next_label(tokens: List[Dict]) -> str:
    kept = []
    for t in tokens:
        txt = (t.get("text") or "").strip()
        if not txt:
            continue
        if txt.endswith(":"):
            break
        kept.append(t)
    return " ".join(t["text"] for t in kept).strip()


def join_tokens(tokens: List[Dict]) -> str:
    return " ".join(t["text"] for t in tokens).strip()


CURRENCY3 = re.compile(r"^[A-Za-z]{3}$")


def extract_currency_positional(words: List[Dict], anchors: List[Dict]) -> Optional[str]:
    for a in anchors:
        same = collect_tokens_on_same_line_right(words, a)
        for t in same:
            cur = (t.get("text") or "").strip().upper()
            if CURRENCY3.match(cur) and cur in KNOWN_CURRENCIES:
                return cur

        down = collect_tokens_next_line_down(words, a, x_window=360, v_range=(0.5, 100))
        for t in down:
            cur = (t.get("text") or "").strip().upper()
            if CURRENCY3.match(cur) and cur in KNOWN_CURRENCIES:
                return cur
    return None


def extract_date_positional(words: List[Dict], anchors: List[Dict]) -> Optional[str]:
    for a in anchors:
        same = collect_tokens_on_same_line_right(words, a)
        if same:
            txt = join_tokens(same)
            m = DATE_TOKEN_REGEX.search(txt)
            if m and is_valid_date_token(m.group(1)):
                return m.group(1)
        down = collect_tokens_next_line_down(words, a, x_window=420, v_range=(0.5, 120))
        if down:
            txt = join_tokens(down)
            m = DATE_TOKEN_REGEX.search(txt)
            if m and is_valid_date_token(m.group(1)):
                return m.group(1)
    return None


def extract_country_positional(words: List[Dict], anchors: List[Dict]) -> Optional[str]:
    for a in anchors:
        same = collect_tokens_on_same_line_right(words, a)
        raw = join_tokens_until_next_label(same)

        if raw:
            val = clean_extracted_value(raw)
            if (
                    val
                    and not HEADER_NOISE.search(val)
                    and not re.search(r"\d", val)
                    and not re.fullmatch(r"[A-Za-z]{2,4}", val)
            ):
                return val

        down = collect_tokens_next_line_down(words, a, x_window=480, v_range=(0.5, 120))
        raw = join_tokens(down)
        if raw:
            val = clean_extracted_value(raw)
            if (
                    val
                    and not HEADER_NOISE.search(val)
                    and not re.search(r"\d", val)
                    and not re.fullmatch(r"[A-Za-z]{2,4}", val)
            ):
                return val
    return None


def extract_preferred_language_positional(
        words: List[Dict], anchors: List[Dict]
) -> Optional[str]:
    for a in anchors:
        same = collect_tokens_on_same_line_right(words, a)
        raw = join_tokens_until_next_label(same)
        if raw:
            val = clean_extracted_value(raw)
            if (
                    val
                    and not HEADER_NOISE.search(val)
                    and not SIGNATURE_NOISE.match(val)
                    and not contains_date_like(val)
            ):
                return val

        down = collect_tokens_next_line_down(words, a, x_window=300, v_range=(0.5, 80))
        raw = join_tokens(down)
        if raw:
            val = clean_extracted_value(raw)
            if (
                    val
                    and not HEADER_NOISE.search(val)
                    and not SIGNATURE_NOISE.match(val)
                    and not contains_date_like(val)
            ):
                return val
    return None


TRN_BAD = re.compile(r"(?i)^(registered|select a status)$")


def sanitize_trn(val: Optional[str]) -> Optional[str]:
    v = clean_extracted_value(val)
    if not v:
        return None
    if TRN_BAD.match(v) or SIGNATURE_NOISE.match(v) or contains_date_like(v):
        return None
    if not re.search(r"\d", v):
        return None
    return v


def extract_trn_same_line(words: List[Dict], anchors: List[Dict]) -> Optional[str]:
    for a in anchors:
        same = collect_tokens_on_same_line_right(words, a, y_tol=6.0)
        if not same:
            continue
        raw = join_tokens(same)
        val = sanitize_trn(raw)
        if val:
            return val
    return None


def is_garbage_text(text: str) -> bool:
    """
    Simple heuristic to detect scrambled text.
    If we don't find any typical HCPIF keywords, treat as "garbage" and use OCR.
    """
    if not text:
        return True
    s = text.strip()
    if len(s) < 40:
        return True
    lowered = s.lower()
    keywords = [
        "hotel collect payment information form",
        "legal name",
        "expedia id",
        "country",
        "currency",
        "first name",
        "last name",
        "email address",
        "preferred language",
    ]
    if any(kw in lowered for kw in keywords):
        return False
    return True


def ocr_extract_full_text(pdf_path: Path) -> str:
    """
    Convert PDF to images and use OCR (pytesseract) to extract readable text.
    Uses DPI 350 and LSTM engine (oem 3, psm 6) for better quality.
    """
    texts = []
    images = convert_from_path(str(pdf_path), dpi=350, poppler_path=POPPLER_PATH)
    for img in images:
        try:
            t = pytesseract.image_to_string(
                img,
                lang="eng",
                config="--oem 3 --psm 6",
            )
            if t:
                texts.append(t)
        except Exception:
            continue
    return "\n".join(texts)


def strip_state_if_person_name(results: Dict[str, Optional[str]]) -> None:
    """
    If State/Province equals First Name, Last Name, or "First Last", clears the field.
    Avoids cases where empty state pulls contact name.
    """
    s = results.get("State/Province")
    if not s:
        return
    fn = results.get("First Name")
    ln = results.get("Last Name")
    if not fn and not ln:
        return

    def norm(x):
        return unidecode(str(x)).strip().lower()

    s_n = norm(s)
    candidates = set()
    if fn:
        candidates.add(norm(fn))
    if ln:
        candidates.add(norm(ln))
    if fn and ln:
        candidates.add((norm(fn) + " " + norm(ln)).strip())

    if s_n in candidates:
        results["State/Province"] = None


def extract_fields_positional(pdf_path: Path) -> Dict[str, Optional[str]]:
    results = {f["col"]: None for f in FIELDS}
    with pdfplumber.open(str(pdf_path)) as pdf:
        full_text_pages = []
        pages_bundle = []
        for idx, page in enumerate(pdf.pages, start=1):
            words = words_from_page(page)
            text = page.extract_text() or ""
            full_text_pages.append(text)
            pages_bundle.append((idx, page, words, text))

        full_text_raw = normalize_block_text("\n".join(full_text_pages))
        use_ocr = is_garbage_text(full_text_raw)
        if use_ocr:
            ocr_text = ocr_extract_full_text(pdf_path)
            full_text = normalize_block_text(ocr_text)
            pages_bundle = []
        else:
            full_text = full_text_raw

        for idx, page, words, text in pages_bundle:
            if not words:
                continue

            anchors_currency = find_label_anchors(words, ["currency"])
            anchors_country = find_label_anchors(words, ["country"])
            anchors_edc = find_label_anchors(words, ["effective", "date", "of", "change"])
            anchors_today = find_label_anchors(words, ["today's", "date"])
            anchors_trn = find_label_anchors(words, ["tax", "registration", "number"])
            anchors_state = find_label_anchors(words, ["state", "province"]) + \
                            find_label_anchors(words, ["state"]) + \
                            find_label_anchors(words, ["province"])
            anchors_language = find_label_anchors(words, ["preferred", "language"])

            if not results["Preferred Language"]:
                results["Preferred Language"] = extract_preferred_language_positional(
                    words, anchors_language
                )

            if not results["Currency"]:
                results["Currency"] = extract_currency_positional(words, anchors_currency)

            if not results["Effective Date of Change"]:
                results["Effective Date of Change"] = extract_date_positional(
                    words, anchors_edc
                )

            if not results["Today's date"]:
                results["Today's date"] = extract_date_positional(words, anchors_today)

            if not results["Country"]:
                results["Country"] = extract_country_positional(words, anchors_country)

            if not results["Tax Registration Number"]:
                results["Tax Registration Number"] = extract_trn_same_line(
                    words, anchors_trn
                )

            # State/Province ONLY from labeled field (positional)
            if not results["State/Province"] and anchors_state:
                for a in anchors_state:
                    same = collect_tokens_on_same_line_right(words, a, y_tol=6.0)
                    if same:
                        raw = join_tokens(same)
                        val = clean_extracted_value(raw)
                        if val:
                            results["State/Province"] = val
                            break

        d1, d2 = extract_two_dates_from_lines(full_text)
        if d1 and not results["Today's date"]:
            results["Today's date"] = d1
        if d2 and not results["Effective Date of Change"]:
            results["Effective Date of Change"] = d2

        # Regex fallback for Currency field
        if not results["Currency"]:
            # Try multiple patterns to catch currency
            patterns = [
                r"(?is)\bcurrency\s*:?\s*([A-Za-z]{3})\b",  # "Currency: USD" or "Currency USD"
                r"(?is)\bcurrency\s*:?\s*([A-Za-z ]{3,20})\b",  # "Currency: US Dollar"
                r"(?is)\bcurrency\s*code\s*:?\s*([A-Za-z]{3})\b",  # "Currency Code: USD"
            ]
            for pattern in patterns:
                m = re.search(pattern, full_text)
                if m:
                    raw_cur = m.group(1).strip()
                    normalized = normalize_currency_word(raw_cur)
                    if normalized:
                        results["Currency"] = normalized
                        break

        if not results["Country"]:
            m_same = re.search(r"(?im)^\s*country\s*:\s*(.+)$", full_text)
            if m_same:
                val = clean_extracted_value(m_same.group(1))
                if val:
                    results["Country"] = val

        # ------- Regex fallback for remaining fields -------
        for f in FIELDS:
            col = f["col"]
            if results.get(col):
                continue

            # Controlled fallback for State/Province:
            if col == "State/Province":
                found = None
                specific_labels = [r"state\s*/\s*province", r"state\s*province"]
                for lab in specific_labels:
                    # 1) try same line as label
                    found = try_same_line_block(full_text, lab)
                    # 2) if not found, use SPECIFIC State/Province fallback
                    if not found:
                        found = try_next_line_block_state(full_text, lab)
                    if found:
                        break
                results[col] = found
                continue

            found = None
            if col == "Tax Registration Number":
                for lab in f["labels"]:
                    found = try_same_line_block(full_text, lab)
                    if found:
                        found = sanitize_trn(found)
                    if found:
                        break
            else:
                for lab in f["labels"]:
                    found = try_same_line_block(full_text, lab)
                    if found:
                        break
            if not found:
                for lab in f["labels"]:
                    found = try_next_line_block(full_text, lab, BARRIER_LABELS)
                    if found:
                        break
            if col.lower().startswith("email"):
                if found and "@" not in str(found):
                    found = None
                if not found:
                    m = EMAIL_FALLBACK.search(full_text)
                    if m:
                        found = m.group(0)
            results[col] = found
        # ------- end FIELDS fallback -------

        # --- FIX 2: normalize currency word before KNOWN_CURRENCIES check ---
        if results["Currency"]:
            raw_cur = str(results["Currency"]).strip()
            # Try direct normalize first
            resolved = normalize_currency_word(raw_cur)
            results["Currency"] = resolved  # may be None if unrecognised – fallback below

        # Clear state if it equals contact name
        strip_state_if_person_name(results)

        # Only use sanitize_state_and_postal to handle obvious swaps;
        results["State/Province"], results["Postal Code"] = sanitize_state_and_postal(
            results.get("State/Province"), results.get("Postal Code")
        )

        for name_col in ("First Name", "Last Name"):
            if results.get(name_col):
                v = clean_extracted_value(results[name_col])
                v = strip_parenthesized_dates(strip_signature_prefix(v) or v) or v
                results[name_col] = None if (v and contains_date_like(v)) else v

        trn = results.get("Tax Registration Number")
        results["Tax Registration Number"] = sanitize_trn(trn) if trn else None

        # --- Extra sanity for Postal Code (only digits, OCR fix O->0, G->8) ---
        p_val = results.get("Postal Code")
        if p_val:
            s = str(p_val)
            trans = str.maketrans({"O": "0", "o": "0", "G": "8", "g": "8"})
            s2 = s.translate(trans)
            digits = "".join(ch for ch in s2 if ch.isdigit())
            results["Postal Code"] = digits or None
            p_val = results["Postal Code"]

        if p_val and not re.search(r"\d", str(p_val)):
            results["Postal Code"] = None

        # Special case: Hong Kong should not have short numeric "CEP" like 852
        country_raw = (results.get("Country") or "").strip()
        iso2_country = None
        try:
            iso2_country = to_iso2(country_raw) or ""
        except Exception:
            iso2_country = ""
        if iso2_country.upper() == "HK":
            p_val = results.get("Postal Code")
            if p_val and re.fullmatch(r"\d{1,4}", str(p_val).strip()):
                results["Postal Code"] = None

        cur = results.get("Currency")
        c_val = sanitize_country(results.get("Country"), cur)
        if c_val and (
                re.fullmatch(r"[A-Za-z]{2,4}", c_val)
                or (cur and c_val.upper() == cur.upper())
                or contains_date_like(c_val)
        ):
            c_val = None
        results["Country"] = c_val or results.get("Country")

        for date_col in ("Today's date", "Effective Date of Change"):
            raw = results.get(date_col)
            token = extract_first_date_token(raw)
            results[date_col] = token

        if results.get("Today's date") and not results.get("Effective Date of Change"):
            d1, d2 = extract_two_dates_from_lines(full_text)
            if not d2 and d1 == results["Today's date"]:
                results["Effective Date of Change"] = d1

        cur = results.get("Currency")
        if isinstance(cur, str):
            c = cur.strip().upper()
            if not re.fullmatch(r"[A-Z]{3}", c):
                m = re.search(r"\b([A-Z]{3})\b", c)
                c = m.group(1).upper() if m else None
            if c and c not in KNOWN_CURRENCIES:
                c = None
            results["Currency"] = c

        # Final fallback: if currency is still None, derive from country
        if not results.get("Currency"):
            country_raw = (results.get("Country") or "").strip()
            if country_raw:
                cur_final = get_currency_from_country(country_raw)
                results["Currency"] = cur_final
                if debug := False:  # Enable for debugging currency issues
                    if cur_final:
                        print(f"[DEBUG] Currency derived from country '{country_raw}': {cur_final}")
                    else:
                        print(f"[WARN] Could not derive currency for country: '{country_raw}'")

        translit_fields = {
            "First Name",
            "Last Name",
            "City",
            "State/Province",
            "Country",
            "Legal Name",
            "Hotel Name",
            "Address Line 1",
            "Preferred Language",
            "Tax Registration Status",
        }
        for col in translit_fields:
            if results.get(col):
                results[col] = unidecode(results[col]).strip()

        return results


def find_pdfs(input_dir: Path) -> List[Path]:
    return sorted([p for p in input_dir.rglob("*.pdf") if p.is_file()])


WEBADI_SHEET_DEFAULT = "WebADI"


def to_iso2(country_name: Optional[str]) -> Optional[str]:
    if country_name is None:
        return None
    try:
        if pd.isna(country_name):
            return None
    except Exception:
        pass

    v = str(country_name).strip()
    if not v or v.upper() == "<NA>":
        return None

    v_norm = unidecode(v).strip().upper()

    MANUAL_ISO2 = {
        "TURKEY": "TR",
        "TURKIYE": "TR",
    }
    if v_norm in MANUAL_ISO2:
        return MANUAL_ISO2[v_norm]

    if re.fullmatch(r"[A-Za-z]{2}", v):
        return v.upper()

    if pycountry:
        for cand in (v_norm, v):
            try:
                c = pycountry.countries.lookup(cand)
                return getattr(c, "alpha_2", None)
            except Exception:
                continue
    return None


def normalize_name_for_match(val: Optional[str]) -> Optional[str]:
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except Exception:
        pass

    s = unidecode(str(val)).upper().strip()
    if not s or s == "<NA>":
        return None

    s = s.replace("&", " AND ")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()

    return s or None

def normalize_hotel_name_for_sql(val: Optional[str]) -> Optional[str]:
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except Exception:
        pass

    s = unidecode(str(val)).upper().strip()
    if not s or s == "<NA>":
        return None

    s = s.replace("&", " AND ")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()

    # Safety check: reject names that are too short (likely to cause false matches)
    # Must have at least 3 characters after normalization
    if len(s) < 3:
        return None

    # Safety check: reject common generic terms that would cause false matches
    generic_terms = {"INC", "LLC", "LTD", "CORP", "CO", "THE", "AND", "SA", "SPA",
                     "GMBH", "SRL", "BV", "NV", "AG", "PLC", "PTY"}
    if s in generic_terms:
        return None

    return s or None

def _norm_header_key(s: str) -> str:
    if s is None:
        return ""
    s = str(s).replace("\u00A0", " ")
    s = s.replace("\n", " ").replace("\r", " ")
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    s = s.upper()
    s = s.replace("/", "_").replace("-", "_").replace(" ", "_")
    s = re.sub(r"[^A-Z0-9_]", "", s)
    return s


HEADER_SYNONYMS = {
    "BATCH_NAME": {"BATCH", "BATCHNAME"},
    "HOTEL_ID": {"PROPERTY_ID", "PROPERTYID", "EXPEDIA_ID", "HOTEL", "HOTELID", "PROPERTY"},
    "CUSTOMER_NAME": {"CUSTOMER", "CUSTOMERNAME", "LEGAL_NAME"},
    "ADDRESS_LINE_1": {"ADDRESS1", "ADDR1", "ADDRESS_LINE1", "ADDRESS"},
    "CITY": {"TOWN"},
    "STATE": {"STATE_PROVINCE", "STATEORPROVINCE", "STATE_", "STATE__PROVINCE"},
    "PROVINCE": {"STATE_PROVINCE", "STATEORPROVINCE"},
    "POSTAL_CODE": {"ZIP", "ZIP_CODE", "POSTCODE", "POSTALCODE"},
    "COUNTRY": {"COUNTRY_CODE", "COUNTRYNAME"},
    "BILLING_CURRENCY": {"CURRENCY", "CURRENCY_CODE"},
    "FIRST_NAME": {"FIRSTNAME", "CONTACT_FIRST_NAME"},
    "LAST_NAME": {"LASTNAME", "CONTACT_LAST_NAME"},
    "EMAIL_ADDRESS": {"EMAIL", "CONTACT_EMAIL", "E_MAIL"},
    "PREFERRED_LANGUAGE": {"LANGUAGE", "LANG"},
    "TAX_REG_NUMBER": {
        "TAX_REGISTRATION_NUMBER",
        "TRN",
        "VAT",
        "VAT_NUMBER",
        "TAXID",
        "TAX_ID",
        "TAX_REG_NUM",
    },
    "SITE_PURPOSE": {"SITEPURPOSE", "SITE_PURP", "SITE_PURPOSE_"},
    "BILL_TO": {"BILLTO", "SITE_USE_CODE", "SITE_USE", "SITEUSECODE"},
}


def _apply_synonym(key: str) -> str:
    for target, aliases in HEADER_SYNONYMS.items():
        if key == target or key in aliases:
            return target
    return key


def _debug_dump_first_rows(ws, n=6):
    print("\n[DEBUG] Sample of first rows (normalized):")
    for r in range(1, min(ws.max_row, n) + 1):
        vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            vals.append(_apply_synonym(_norm_header_key(v)) if isinstance(v, str) else "")
        print(f"  R{r:02d}: {vals}")


def _debug_dump_row(ws, r):
    vals = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        vals.append((_apply_synonym(_norm_header_key(v)) if isinstance(v, str) else "", v))
    print(f"[DEBUG] Row {r} -> {[x for x in vals if x[1]]}")


def find_header_row(
        ws,
        must_have_cols=("HOTEL_ID", "CUSTOMER_NAME"),
        start_at_row: int = 5,
        debug: bool = False,
) -> Tuple[int, Dict[str, int]]:
    must_have = {_apply_synonym(_norm_header_key(x)) for x in must_have_cols}
    if debug:
        _debug_dump_first_rows(ws, n=max(6, start_at_row + 2))
        print(
            f"[DEBUG] Searching for header starting at row {start_at_row} "
            f"(must_have={sorted(list(must_have))})"
        )
    for r in range(max(1, start_at_row), ws.max_row + 1):
        header_map_norm: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if not isinstance(val, str):
                continue
            raw = (
                val.replace("\u00A0", " ")
                .replace("\n", " ")
                .replace("\r", " ")
                .strip()
            )
            if not raw:
                continue
            key = _apply_synonym(_norm_header_key(raw))
            if key and key not in header_map_norm:
                header_map_norm[key] = c
        if must_have.issubset(header_map_norm.keys()):
            if debug:
                print(
                    f"[DEBUG] Header found at row {r}. "
                    f"Keys: {sorted(header_map_norm.keys())}"
                )
                _debug_dump_row(ws, r)
            return r, header_map_norm
    raise RuntimeError(
        "Could not find header row in WebADI sheet. "
        "Check the template."
    )


def last_data_row(ws, header_row: int, key_col_index: int) -> int:
    r = ws.max_row
    while r > header_row:
        if ws.cell(r, key_col_index).value not in (None, "", " "):
            return r
        r -= 1
    return header_row


def clone_row(ws, src_row: int, dest_row: int):
    for c in range(1, ws.max_column + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dest_row, c)
        dst.value = src.value
        dst.font = copy.copy(src.font)
        dst.border = copy.copy(src.border)
        dst.fill = copy.copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy.copy(src.protection)
        dst.alignment = copy.copy(src.alignment)


def unprotect_workbook_and_sheets(wb, debug: bool = False):
    try:
        if hasattr(wb, "security"):
            wb.security.lockStructure = False
            wb.security.workbookPassword = None
            if debug:
                print("[DEBUG] Workbook security: lockStructure=False, password cleared.")
    except Exception as e:
        if debug:
            print(f"[DEBUG] Failed to modify wb.security: {e}")

    for ws in wb.worksheets:
        try:
            ws.protection.sheet = False
            try:
                ws.protection.password = None
            except Exception:
                pass
            if debug:
                print(f"[DEBUG] Unprotected sheet: {ws.title}")
        except Exception as e:
            if debug:
                print(f"[DEBUG] Failed to unprotect sheet {ws.title}: {e}")


def _is_blank(val) -> bool:
    """True for None, empty string, or NaN."""
    if val is None:
        return True
    if isinstance(val, str) and val.strip() == "":
        return True
    try:
        import pandas as _pd
        return bool(_pd.isna(val))
    except Exception:
        return False


def format_comment_from_filename(fn: str) -> str:
    if not fn:
        return ""
    base = os.path.splitext(str(fn))[0]

    m = re.search(r"(case[^-_]*)[-_ ]*(eid[^-_]*)", base, flags=re.IGNORECASE)
    if m:
        case_part = m.group(1).strip()
        eid_part = m.group(2).strip()
        return f"{case_part} - {eid_part}"

    m = re.search(r"(eid[^-_]*)[-_ ]*(case[^-_]*)", base, flags=re.IGNORECASE)
    if m:
        eid_part = m.group(1).strip()
        case_part = m.group(2).strip()
        return f"{case_part} - {eid_part}"

    return base.strip()


def fetch_oracle_tca(expedia_ids: List[str]) -> pd.DataFrame:
    expedia_ids = sorted({str(x).strip() for x in expedia_ids if pd.notna(x) and str(x).strip()})
    if not expedia_ids:
        return pd.DataFrame()

    ids_literal = ", ".join(f"'{eid}'" for eid in expedia_ids)
    sql = ORACLE_TCA_SQL_TEMPLATE.format(expedia_ids=ids_literal)

    print(f"\n[INFO] Running Oracle query for {len(expedia_ids)} Expedia ID(s)...")

    ensure_oracle_client()

    with oracledb.connect(user=ORACLE_USERNAME, password=ORACLE_PASSWORD, dsn=ORACLE_DSN) as con:
        with con.cursor() as cur:
            cur.execute("ALTER SESSION SET CURRENT_SCHEMA = APPS")

        t0 = time.time()
        with con.cursor() as cur:
            cur.execute(sql)
            cols = [d[0] for d in cur.description]
            rows = cur.fetchall()
        elapsed = time.time() - t0

    print(f"[INFO] Oracle query completed. Rows: {len(rows)}. Time: {elapsed:.2f}s")

    df_oracle = pd.DataFrame(rows, columns=cols)

    if not df_oracle.empty:
        df_oracle["EXPEDIA_ID"] = df_oracle["EXPEDIA_ID"].astype(str).str.strip()
        df_oracle["COUNTRY_ABBR"] = df_oracle["COUNTRY_ABBR"].astype(str).str.strip()
        df_oracle = df_oracle.sort_values(
            ["EXPEDIA_ID", "COUNTRY_ABBR", "DUPLICATE_COUNT"],
            ascending=[True, True, True],
        )
        df_oracle = df_oracle.drop_duplicates(subset=["EXPEDIA_ID", "COUNTRY_ABBR"], keep="first")

    return df_oracle

# FIX 1 – pass country_iso2_list so we match (name, country) in Oracle
def fetch_oracle_hotels_by_name(hotel_names: List[str], country_iso2_list: Optional[List[str]] = None) -> pd.DataFrame:
    """
    Queries Oracle for SLE records matching the given hotel/legal names.

    FIX 1: The result is now deduplicated on (HOTEL_NAME_MATCH, COUNTRY_ABBR)
    when country info is available, preventing the wrong SLE OID from being
    returned when the same legal name appears in more than one country.

    The SQL already returns COUNTRY_ABBR, so no schema change is needed.
    """
    hotel_names_norm = sorted({
        normalize_hotel_name_for_sql(x)
        for x in hotel_names
        if normalize_hotel_name_for_sql(x)
    })
    if not hotel_names_norm:
        print("[INFO] No valid hotel names for SLE OID lookup (all filtered out).")
        return pd.DataFrame()

    # Safety check: warn if we're searching for very short names
    short_names = [n for n in hotel_names_norm if len(n) < 5]
    if short_names:
        print(f"[WARN] Some normalized legal names are very short and may cause false matches: {short_names[:5]}")

    names_literal = ", ".join("'" + n.replace("'", "''") + "'" for n in hotel_names_norm)
    sql = ORACLE_HOTEL_BY_NAME_SQL_TEMPLATE.format(hotel_names=names_literal)
    print(f"\n[INFO] Running Oracle hotel-name query for {len(hotel_names_norm)} hotel name(s)...")
    ensure_oracle_client()
    with oracledb.connect(user=ORACLE_USERNAME, password=ORACLE_PASSWORD, dsn=ORACLE_DSN) as con:
        with con.cursor() as cur:
            cur.execute("ALTER SESSION SET CURRENT_SCHEMA = APPS")
        t0 = time.time()
        with con.cursor() as cur:
            cur.execute(sql)
            cols = [d[0] for d in cur.description]
            rows = cur.fetchall()
        elapsed = time.time() - t0
    print(f"[INFO] Oracle hotel-name query completed. Rows: {len(rows)}. Time: {elapsed:.2f}s")

    # Warning if we got way too many results (possible false match scenario)
    if len(rows) > 100:
        print(f"[WARN] Oracle returned {len(rows)} SLE records. This may indicate overly broad name matching.")

    df_oracle = pd.DataFrame(rows, columns=cols)
    if df_oracle.empty:
        return df_oracle

    df_oracle["HOTEL_NAME_MATCH"] = df_oracle["HOTEL_NAME"].apply(normalize_name_for_match)
    df_oracle["COUNTRY_ABBR"] = df_oracle["COUNTRY_ABBR"].astype(str).str.strip().str.upper()
    df_oracle = df_oracle.sort_values(["HOTEL_NAME_MATCH", "COUNTRY_ABBR", "ORACLE_ID"])

    if country_iso2_list:
        # Build a set of (name_norm, iso2) pairs from the HCPIF for guided deduplication
        hcpif_pairs = set()
        for name, iso2 in zip(
            [normalize_hotel_name_for_sql(x) for x in hotel_names],
            [str(c).strip().upper() if c else "" for c in country_iso2_list]
        ):
            if name:
                hcpif_pairs.add((name, iso2))

        # Prefer the row whose COUNTRY_ABBR matches the HCPIF country; fall back to first row
        def _pick_best(group):
            # First pass: exact country match
            for idx in group.index:
                name_match = group.loc[idx, "HOTEL_NAME_MATCH"]
                country = group.loc[idx, "COUNTRY_ABBR"]
                if (name_match, country) in hcpif_pairs:
                    return group.loc[[idx]]

            # Second pass: if no exact match and group has multiple countries,
            # prefer the one with an Expedia ID (more likely to be a real hotel account)
            if len(group) > 1:
                with_eid = group[group["EXPEDIA_ID"].notna() & (group["EXPEDIA_ID"].astype(str).str.strip() != "")]
                if not with_eid.empty:
                    return with_eid.iloc[[0]]

            # Final fallback: first row (alphabetical by country)
            return group.iloc[[0]]

        df_oracle = (
            df_oracle.groupby("HOTEL_NAME_MATCH", group_keys=False)
            .apply(_pick_best, include_groups=True)
            .reset_index(drop=True)
        )
    else:
        # Legacy behaviour: deduplicate on name only
        df_oracle = df_oracle.drop_duplicates(subset=["HOTEL_NAME_MATCH"], keep="first")

    return df_oracle

def add_sle_oid_from_legal_name(df: pd.DataFrame) -> pd.DataFrame:
    """
    FIX 1: pass both Legal Name AND Country ISO2 to fetch_oracle_hotels_by_name
    so that when the same SLE name exists in multiple countries the correct OID
    (matching the HCPIF country) is selected.

    Also adds a validation column to flag suspicious matches.
    """
    if "Legal Name" not in df.columns:
        return df

    df = df.copy()

    # Gather country ISO2 list aligned with Legal Name list
    country_iso2_list: Optional[List[str]] = None
    if "Country ISO2" in df.columns:
        country_iso2_list = df["Country ISO2"].tolist()
    elif "Country" in df.columns:
        country_iso2_list = [to_iso2(c) or "" for c in df["Country"].tolist()]

    oracle_base = fetch_oracle_hotels_by_name(
        df["Legal Name"].tolist(),
        country_iso2_list=country_iso2_list,
    )

    insert_pos = list(df.columns).index("Legal Name") + 1 if "Legal Name" in df.columns else len(df.columns)

    if oracle_base.empty:
        if "SLE OID" not in df.columns:
            df.insert(loc=insert_pos, column="SLE OID", value=pd.NA)
            df.insert(loc=insert_pos + 1, column="SLE Match Status", value=pd.NA)
        return df

    df["LEGAL_NAME_MATCH"] = df["Legal Name"].apply(normalize_name_for_match)

    # Add HCPIF country for secondary match
    if country_iso2_list:
        df["_HCPIF_ISO2"] = [str(c).strip().upper() if c else "" for c in country_iso2_list]
    else:
        df["_HCPIF_ISO2"] = ""

    # Keep the Oracle country in the merge for validation
    match_df = df.merge(
        oracle_base[["HOTEL_NAME_MATCH", "COUNTRY_ABBR", "ORACLE_ID", "EXPEDIA_ID"]],
        how="left",
        left_on=["LEGAL_NAME_MATCH", "_HCPIF_ISO2"],
        right_on=["HOTEL_NAME_MATCH", "COUNTRY_ABBR"],
        suffixes=("", "_ORACLE")
    )

    # Track which rows got country-matched vs name-only matched
    match_df["_COUNTRY_MATCHED"] = match_df["ORACLE_ID"].notna()

    # If country-keyed merge left some rows unmatched, try name-only fallback for those rows
    if "ORACLE_ID" in match_df.columns:
        no_match_mask = match_df["ORACLE_ID"].isna()
        if no_match_mask.any():
            name_only = oracle_base.drop_duplicates(subset=["HOTEL_NAME_MATCH"], keep="first")
            fallback = (
                df.loc[no_match_mask, ["LEGAL_NAME_MATCH"]]
                .merge(name_only[["HOTEL_NAME_MATCH", "ORACLE_ID", "COUNTRY_ABBR", "EXPEDIA_ID"]],
                       how="left",
                       left_on="LEGAL_NAME_MATCH",
                       right_on="HOTEL_NAME_MATCH")
            )
            match_df.loc[no_match_mask, "ORACLE_ID"] = fallback["ORACLE_ID"].values
            match_df.loc[no_match_mask, "COUNTRY_ABBR"] = fallback["COUNTRY_ABBR"].values
            match_df.loc[no_match_mask, "EXPEDIA_ID_ORACLE"] = fallback["EXPEDIA_ID"].values

    # Create validation status column
    def _validate_sle_match(row):
        if pd.isna(row.get("ORACLE_ID")) or str(row.get("ORACLE_ID")).strip() == "":
            return "No SLE Found"

        # Check if country matched
        hcpif_country = str(row.get("_HCPIF_ISO2", "")).strip().upper()
        oracle_country = str(row.get("COUNTRY_ABBR", "")).strip().upper()

        if hcpif_country and oracle_country and hcpif_country != oracle_country:
            return f"Country Mismatch (HCPIF: {hcpif_country}, Oracle: {oracle_country})"

        # Check if the normalized name is very short (higher risk of false match)
        norm_name = str(row.get("LEGAL_NAME_MATCH", "")).strip()
        if len(norm_name) < 5:
            return f"Short Name Match (length {len(norm_name)}) - Review"

        # Good match
        if row.get("_COUNTRY_MATCHED"):
            return "Match (Name + Country)"
        else:
            return "Match (Name Only) - Review if country important"

    match_df["SLE Match Status"] = match_df.apply(_validate_sle_match, axis=1)

    # Drop existing columns before inserting at specific position
    if "SLE OID" in match_df.columns:
        match_df = match_df.drop(columns=["SLE OID"])

    # Store SLE Match Status values before dropping
    sle_match_status_values = match_df["SLE Match Status"]
    if "SLE Match Status" in match_df.columns:
        match_df = match_df.drop(columns=["SLE Match Status"])

    # Insert columns at desired positions
    match_df.insert(loc=insert_pos, column="SLE OID", value=match_df.get("ORACLE_ID", pd.NA))
    match_df.insert(loc=insert_pos + 1, column="SLE Match Status", value=sle_match_status_values)

    match_df = match_df.drop(
        columns=["LEGAL_NAME_MATCH", "HOTEL_NAME_MATCH", "ORACLE_ID", "COUNTRY_ABBR",
                 "_HCPIF_ISO2", "_COUNTRY_MATCHED", "EXPEDIA_ID_ORACLE"],
        errors="ignore",
    )
    return match_df

def create_sle_match_review_tab(df: pd.DataFrame, df_oracle: pd.DataFrame) -> pd.DataFrame:
    """
    Creates a review tab for rows where the HCPIF data matches an existing SLE in Oracle.

    Flags rows where:
    - SLE OID exists
    - Legal Name matches Oracle SLE_NAME (normalized)
    - Country matches Oracle COUNTRY_ABBR

    Returns DataFrame with flagged rows for coordinator review.
    """
    if df_oracle.empty or "SLE OID" not in df.columns:
        return pd.DataFrame()

    # Only check rows that have an SLE OID
    df_with_sle = df[df["SLE OID"].notna() & (df["SLE OID"].astype(str).str.strip() != "")].copy()

    if df_with_sle.empty:
        return pd.DataFrame()

    # Merge with Oracle data to get SLE_NAME and COUNTRY_ABBR
    df_with_sle["SLE OID"] = df_with_sle["SLE OID"].astype(str).str.strip()

    # Create a lookup from Oracle data (SLE OID -> SLE_NAME, COUNTRY_ABBR)
    oracle_sle_lookup = {}
    if "SLE_OID" in df_oracle.columns and "SLE_NAME" in df_oracle.columns and "COUNTRY_ABBR" in df_oracle.columns:
        for _, row in df_oracle.iterrows():
            sle_oid = str(row.get("SLE_OID", "")).strip()
            sle_name = str(row.get("SLE_NAME", "")).strip()
            country = str(row.get("COUNTRY_ABBR", "")).strip().upper()
            if sle_oid and sle_oid not in ("", "nan", "None"):
                oracle_sle_lookup[sle_oid] = {
                    "SLE_NAME": sle_name,
                    "COUNTRY_ABBR": country
                }

    if not oracle_sle_lookup:
        print("[INFO] No Oracle SLE data available for match review.")
        return pd.DataFrame()

    # Check for matches
    matches = []
    for idx, row in df_with_sle.iterrows():
        sle_oid = str(row.get("SLE OID", "")).strip()
        hcpif_legal_name = str(row.get("Legal Name", "")).strip()
        hcpif_country_iso2 = str(row.get("Country ISO2", "")).strip().upper()

        if sle_oid in oracle_sle_lookup:
            oracle_data = oracle_sle_lookup[sle_oid]
            oracle_sle_name = oracle_data["SLE_NAME"]
            oracle_country = oracle_data["COUNTRY_ABBR"]

            # Normalize for comparison
            hcpif_name_norm = normalize_name_for_match(hcpif_legal_name)
            oracle_name_norm = normalize_name_for_match(oracle_sle_name)

            # Check if name and country match
            name_match = hcpif_name_norm == oracle_name_norm if hcpif_name_norm and oracle_name_norm else False
            country_match = hcpif_country_iso2 == oracle_country if hcpif_country_iso2 and oracle_country else False

            if name_match and country_match:
                # This is a potential duplicate/already exists scenario
                row_copy = row.copy()
                row_copy["Oracle SLE Name"] = oracle_sle_name
                row_copy["Oracle Country"] = oracle_country
                row_copy["Match Reason"] = "Legal Name + Country match existing SLE - Review before processing"
                matches.append(row_copy)

    if not matches:
        print("[INFO] No SLE matches found requiring review.")
        return pd.DataFrame()

    df_review = pd.DataFrame(matches)
    print(f"[INFO] {len(df_review)} row(s) flagged for SLE match review.")
    return df_review


def enrich_hcpif_with_oracle(
        df: pd.DataFrame, base_columns: List[str]
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if "Expedia ID" not in df.columns:
        print("[WARN] Column 'Expedia ID' not found; skipping Oracle enrichment.")
        return df, pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    expedia_ids = df["Expedia ID"].tolist()
    df_oracle = fetch_oracle_tca(expedia_ids)
    if df_oracle.empty:
        print("[INFO] No Oracle data returned; HCPIF remains without enrichment.")
        return df, pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    df["Expedia ID"] = df["Expedia ID"].astype(str).str.strip()
    if "Country ISO2" not in df.columns:
        df["Country ISO2"] = df["Country"].map(lambda x: to_iso2(x) or "")

    df_oracle["EXPEDIA_ID"] = df_oracle["EXPEDIA_ID"].astype(str).str.strip()
    df_oracle["COUNTRY_ABBR"] = df_oracle["COUNTRY_ABBR"].astype(str).str.strip()

    merged = df.merge(
        df_oracle,
        how="left",
        left_on=["Expedia ID", "Country ISO2"],
        right_on=["EXPEDIA_ID", "COUNTRY_ABBR"],
        suffixes=("", "_ORACLE"),
    )

    # ── FALLBACK: if no match by country, try by Expedia ID only ──
    mask_no_match = merged["ORACLE_ID"].isna() if "ORACLE_ID" in merged.columns else pd.Series([False] * len(merged))
    if mask_no_match.any():
        df_oracle_by_eid = df_oracle.drop_duplicates(subset=["EXPEDIA_ID"], keep="first")
        fallback = (
            df.loc[mask_no_match, ["Expedia ID"]]
            .merge(
                df_oracle_by_eid,
                how="left",
                left_on="Expedia ID",
                right_on="EXPEDIA_ID",
            )
        )
        oracle_fill_cols = [
            "ORACLE_ID", "HOTEL_NAME", "ORACLE_CURRENCY",
            "CURRENT_RM", "RECEIPT_METHOD_NAME", "SLE_NAME",
            "SLE_OID", "TAI_BM", "DUPLICATE_COUNT",
        ]
        for col in oracle_fill_cols:
            if col in fallback.columns and col in merged.columns:
                merged.loc[mask_no_match, col] = fallback[col].values
    # ── end fallback ──

    merged["Found Hotel OID"] = merged["ORACLE_ID"]
    merged["Found Hotel Name"] = merged["HOTEL_NAME"]

    if "Hotel Name" in merged.columns:
        merged["Hotel Name"] = merged["Hotel Name"].where(
            merged["Hotel Name"].notna()
            & (merged["Hotel Name"].astype(str).str.strip() != ""),
            merged["HOTEL_NAME"],
        )
    else:
        merged["Hotel Name"] = merged["HOTEL_NAME"]

    merged["Oracle Currency"] = merged["ORACLE_CURRENCY"]
    merged["Current RM"] = merged["CURRENT_RM"]

    if "Currency" in merged.columns:
        merged["Currency"] = merged["Currency"].astype(str).str.upper().str.strip()
        merged["Oracle Currency"] = merged["Oracle Currency"].astype(str).str.upper().str.strip()

        def _match(cur, oc):
            if not cur or not oc:
                return "NO"
            return "YES" if cur == oc else "NO"

        merged["Currency Matches Oracle"] = [
            _match(c, o) for c, o in zip(merged["Currency"], merged["Oracle Currency"])
        ]
    else:
        merged["Currency Matches Oracle"] = pd.NA

    comment_col = "Comments"
    new_comment = merged.get("file_name", pd.Series([""] * len(merged))).apply(
        lambda fn: f"Updated Ownership per HCPIF - {format_comment_from_filename(fn)}"
    )
    if comment_col in merged.columns:
        merged[comment_col] = merged[comment_col].fillna("").astype(str)
        merged[comment_col] = merged[comment_col].where(
            merged[comment_col].str.strip() != "",
            new_comment,
        )
    else:
        merged[comment_col] = new_comment

    oracle_key_cols = [c for c in ["Expedia ID", "Country ISO2", "file_name", "file_path"] if c in merged.columns]
    exclude_for_details = set(base_columns) | {
        "Found Hotel OID",
        "Found Hotel Name",
        "Oracle Currency",
        "Currency Matches Oracle",
        "Comments",
    }
    oracle_value_cols = [c for c in merged.columns if c not in exclude_for_details]
    mask_has_oracle = merged["ORACLE_ID"].notna() if "ORACLE_ID" in merged.columns else merged[
        "Found Hotel OID"].notna()
    df_oracle_details = merged.loc[mask_has_oracle, oracle_key_cols + oracle_value_cols].copy()

    series_rm = merged.get("RECEIPT_METHOD_NAME", pd.Series([""] * len(merged))).astype(str)
    mask_direct_debit = series_rm.str.contains("DIRECT DEBIT", case=False, na=False)

    series_tai = merged.get("TAI_BM", pd.Series([""] * len(merged))).astype(str)
    mask_tai_group = series_tai.str.contains("GROUP", case=False, na=False)

    mask_excluded = mask_direct_debit | mask_tai_group

    df_main_raw = merged.copy()
    df_excluded_raw = merged[mask_excluded].copy()

    if not df_excluded_raw.empty:
        print(
            f"[INFO] {len(df_excluded_raw)} row(s) flagged for DIRECT DEBIT/TAI_BM GROUP review."
        )

    # --- FIX 3: NO OID - EC tab ---
    # Rows where BOTH Found Hotel OID (Hotel OID from TCA query) AND SLE OID are blank,
    # AND the business model is NOT GROUP (i.e. Expedia Collect / Direct).
    has_hotel_oid = merged["Found Hotel OID"].notna() & (merged["Found Hotel OID"].astype(str).str.strip() != "")
    has_sle_oid = False  # default
    if "SLE OID" in merged.columns:
        has_sle_oid = merged["SLE OID"].notna() & (merged["SLE OID"].astype(str).str.strip().isin(["", "nan", "None", "<NA>"]) == False)
    elif "SLE_OID" in merged.columns:
        has_sle_oid = merged["SLE_OID"].notna() & (merged["SLE_OID"].astype(str).str.strip().isin(["", "nan", "None", "<NA>"]) == False)

    mask_no_oid_at_all = (~has_hotel_oid) & (~has_sle_oid if isinstance(has_sle_oid, pd.Series) else True)
    mask_not_group = ~series_tai.str.contains("GROUP", case=False, na=False)
    mask_no_oid_ec = mask_no_oid_at_all & mask_not_group

    df_no_oid_ec_raw = merged[mask_no_oid_ec].copy()
    if not df_no_oid_ec_raw.empty:
        print(f"[INFO] {len(df_no_oid_ec_raw)} row(s) flagged for 'NO OID - EC' review tab.")

    extra_cols = ["Found Hotel OID", "Found Hotel Name", "Oracle Currency", "Currency Matches Oracle", "Comments",
                  "Current RM"]
    final_main_cols = list(base_columns)
    for col in extra_cols:
        if col in df_main_raw.columns and col not in final_main_cols:
            final_main_cols.append(col)

    df_main = df_main_raw.loc[:, final_main_cols].copy()
    df_excluded = df_excluded_raw.loc[:, final_main_cols].copy()
    df_no_oid_ec = df_no_oid_ec_raw.loc[:, final_main_cols].copy() if not df_no_oid_ec_raw.empty else pd.DataFrame()

    # Create SLE match review tab
    df_sle_review = create_sle_match_review_tab(df_main, df_oracle)

    return df_main, df_excluded, df_oracle_details, df_no_oid_ec, df_sle_review


def normalize_language_to_code(val: Optional[str]) -> Optional[str]:
    """
    Converts language (string) to ISO 639-1 2-letter code (EN, PT, ES, ...).
    - Uses pycountry.languages as primary source (international standard).
    - Accepts names ('English', 'Portuguese (Brazil)'), codes with region ('en-US', 'pt_BR').
    - Returns uppercase codes for consistency with Oracle EBS.
    - If not recognized with confidence, returns None (leaves blank in WEBADI).
    """
    if val is None:
        return None

    s_raw = str(val).strip()
    if not s_raw:
        return None

    # Normalize accents and case
    s = unidecode(s_raw).strip()

    # 1) Already a code like "en", "EN", "pt", etc.
    if len(s) == 2 and s.isalpha():
        code = s.upper()
        # Validate it's a real language code
        if pycountry is not None:
            try:
                pycountry.languages.get(alpha_2=code.lower())
                return code
            except:
                pass
        return code

    # 2) Get 2-letter prefix before "-" or "_" (e.g., "en-US", "pt_BR", "en (US)")
    m = re.match(r"\s*([A-Za-z]{2})\b", s)
    if m:
        code = m.group(1).upper()
        # Validate it's a real language code
        if pycountry is not None:
            try:
                pycountry.languages.get(alpha_2=code.lower())
                return code
            except:
                pass
        return code

    # 3) Try lookup via pycountry (name, alpha_2, alpha_3, bibliographic, etc.)
    if pycountry is not None:
        try:
            lang = pycountry.languages.lookup(s.lower())
            alpha2 = getattr(lang, "alpha_2", None)
            if alpha2 and len(alpha2) == 2:
                return alpha2.upper()
        except Exception:
            pass

    # 4) Manual fallback for common names (in English/Portuguese/Spanish)
    # Only use these as last resort if pycountry lookup failed
    s_low = s.lower()
    if "english" in s_low:
        return "EN"
    if "portugu" in s_low:
        return "PT"
    if "spanish" in s_low or "espanol" in s_low or "espanhol" in s_low:
        return "ES"
    if "french" in s_low or "frances" in s_low:
        return "FR"
    if "german" in s_low or "alemao" in s_low or "aleman" in s_low:
        return "DE"
    if "italian" in s_low or "italiano" in s_low:
        return "IT"
    if "dutch" in s_low or "holandes" in s_low or "neerlandes" in s_low:
        return "NL"  # Dutch = NL (ISO 639-1 standard)
    if "japanese" in s_low or "japones" in s_low:
        return "JA"
    if "chinese" in s_low or "mandarin" in s_low:
        return "ZH"
    if "korean" in s_low:
        return "KO"
    if "russian" in s_low:
        return "RU"

    # If could not map with confidence, leave blank in WEBADI
    return None


def inject_into_webadi_update(
        template_path: Path,
        out_path: Optional[Path],
        df: pd.DataFrame,
        mode: str = "replace",
        sheet_name: str = WEBADI_SHEET_DEFAULT,
        header_start_row: int = 5,
        debug: bool = False,
        unprotect: bool = True,
):
    """
    Fills the OC WEBADI UPDATE template.

    Includes: Tax info, payment method, currency
    ACTION = "Update"
    GROUP_PARENT_SLE_ACCOUNT_NUMBER = SLE OID
    """
    if not template_path.exists():
        raise FileNotFoundError(f"WEBADI UPDATE template not found: {template_path}")

    wb = load_workbook(
        str(template_path),
        keep_vba=True,
        data_only=False,
        keep_links=True,
    )

    if unprotect:
        if debug:
            print("[DEBUG] webadi_unprotect=True: unprotecting workbook/sheets.")
        unprotect_workbook_and_sheets(wb, debug=debug)
    else:
        if debug:
            print("[DEBUG] webadi_unprotect=False: keeping protections.")

    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' not found in template.")
    ws = wb[sheet_name]

    # Update BATCH_NAME in E3 with today's date
    try:
        today_str = datetime.now().strftime("%m/%d/%Y")
        ws["E3"].value = f"OC WEBADI UPDATE {today_str}"
        if debug:
            print(f"[DEBUG] E3 (BATCH_NAME) set to 'OC WEBADI UPDATE {today_str}'")
    except Exception as e:
        if debug:
            print(f"[DEBUG] Failed to update BATCH_NAME in E3: {e}")

    header_row, header_map = find_header_row(
        ws, start_at_row=header_start_row, debug=debug
    )

    if debug:
        print(f"[DEBUG] UPDATE WebADI: Found {len(header_map)} headers in template:")
        for norm_name, col_idx in sorted(header_map.items()):
            print(f"  {norm_name} → Column {col_idx}")

    # Mapping: normalized WEBADI header -> df column (UPDATE template includes tax info)
    wanted = {
        "GROUP_PARENT_SLE_ACCOUNT_NUMBER": "SLE OID",
        "HOTEL_ID": "Expedia ID",
        "EFFECTIVE_DATE": "Effective Date of Change",
        "CUSTOMER_NAME": "Hotel Name",
        "TAX_REG_NUM": "Tax Registration Number",
        "TAX_REGIME_CODE": "WEBADI_TAX_REGIME_CODE",
        "TAX": "WEBADI_TAX",
        "TAX_REG_STATUS": "Tax Registration Status",
        "ADDRESS_LINE_1": "Address Line 1",  # Changed from ADDRESS1 to match OID Creation
        "CITY": "City",
        "POSTAL_CODE": "Postal Code",
        "COUNTRY": "WEBADI_COUNTRY",
        "PAYMENT_METHOD": "Current RM",
        "BILLING_CURRENCY": "Currency",
        "FIRST_NAME": "First Name",
        "LAST_NAME": "Last Name",
        "EMAIL_ADDRESS": "Email Address",
        "PREFERRED_LANGUAGE": "Preferred Language",
        "COMMENTS": "file_name",  # Will be converted to comment format
    }

    col_state = header_map.get("STATE")
    col_province = header_map.get("PROVINCE")
    col_site_purpose = header_map.get("SITE_PURPOSE")
    col_action = header_map.get("ACTION")
    col_upl = column_index_from_string("B")
    col_tax_reg_type = header_map.get("TAX_REG_TYPE")

    key_col = header_map.get("HOTEL_ID") or header_map.get("CUSTOMER_NAME")
    if not key_col:
        raise RuntimeError(
            "Could not find key columns (HOTEL_ID/CUSTOMER_NAME) in UPDATE WebADI header."
        )

    # Clear or clone rows according to mode
    if mode.lower() == "replace":
        last = ws.max_row
        for _ in range(header_row + 1, last + 1):
            ws.delete_rows(header_row + 1)
        base_row = header_row
        if debug:
            print(
                f"[DEBUG] Replace mode: cleared rows after header row {header_row}."
            )
    else:
        base_row = last_data_row(ws, header_row, key_col)
        if debug:
            print(
                f"[DEBUG] Append mode: last data row detected = {base_row}."
            )

    current = base_row

    for _, rec in df.iterrows():
        iso2 = to_iso2(rec.get("Country"))
        stateprov = rec.get("State/Province")

        current += 1
        src_row = base_row if base_row > header_row else header_row + 1
        clone_row(ws, src_row, current)

        # Set fixed columns for UPDATE template
        ws.cell(current, col_upl).value = "O"
        if col_action:
            ws.cell(current, col_action).value = "Update"
        if col_site_purpose:
            ws.cell(current, col_site_purpose).value = "BILL_TO"
        if col_tax_reg_type:
            ws.cell(current, col_tax_reg_type).value = "Others"

        # Always clear STATE/PROVINCE in new rows to avoid inheriting junk
        if col_state:
            ws.cell(current, col_state).value = None
        if col_province:
            ws.cell(current, col_province).value = None

        # Fill columns according to mapping
        for webadi_col, script_col in wanted.items():
            col_idx = header_map.get(webadi_col)
            if not col_idx:
                if debug:
                    print(f"[DEBUG] Column '{webadi_col}' not found in template header, skipping")
                continue
            val = rec.get(script_col)
            if debug and webadi_col in ["FIRST_NAME", "LAST_NAME", "EMAIL_ADDRESS", "ADDRESS_LINE_1"]:
                print(f"[DEBUG] Filling {webadi_col} (col {col_idx}) with value: {val}")

            # COMMENTS: Format as "filename Updated per HCPIF - update hotel name, address, tax, contact"
            # Note: file_name already contains Expedia IDs, so don't duplicate them
            if webadi_col == "COMMENTS":
                file_name = rec.get("file_name", "")
                if not _is_blank(file_name):
                    comment_text = f"{file_name} Updated per HCPIF - update hotel name, address, tax, contact"
                    ws.cell(current, col_idx).value = comment_text
                else:
                    ws.cell(current, col_idx).value = None

            # TAX_REG_STATUS: "REGISTERED" if contains "REGISTERED"
            elif webadi_col == "TAX_REG_STATUS":
                if _is_blank(val):
                    ws.cell(current, col_idx).value = None
                else:
                    s = str(val).strip().upper()
                    ws.cell(current, col_idx).value = "REGISTERED" if "REGISTERED" in s else None

            # Currency: 3 letters or blank
            elif webadi_col == "BILLING_CURRENCY":
                if _is_blank(val):
                    ws.cell(current, col_idx).value = None
                else:
                    if isinstance(val, str) and re.fullmatch(r"[A-Za-z]{3}", val.strip()):
                        ws.cell(current, col_idx).value = val.strip().upper()
                    else:
                        ws.cell(current, col_idx).value = val

            # EFFECTIVE_DATE: Format as '15/Feb/2026' (DD/MMM/YYYY)
            elif webadi_col == "EFFECTIVE_DATE":
                if _is_blank(val):
                    ws.cell(current, col_idx).value = None
                else:
                    # Try to parse and reformat the date
                    try:
                        if isinstance(val, datetime):
                            formatted_date = val.strftime("%d/%b/%Y")
                            ws.cell(current, col_idx).value = formatted_date
                        elif isinstance(val, str):
                            # Try to parse string date
                            from dateutil import parser
                            parsed_date = parser.parse(val)
                            formatted_date = parsed_date.strftime("%d/%b/%Y")
                            ws.cell(current, col_idx).value = formatted_date
                        else:
                            ws.cell(current, col_idx).value = val
                    except:
                        # If parsing fails, write as-is
                        ws.cell(current, col_idx).value = val

            # All other fields
            else:
                if _is_blank(val):
                    ws.cell(current, col_idx).value = None
                else:
                    ws.cell(current, col_idx).value = val

        # Fill State/Province in WebADI, after clearing the columns above
        if not _is_blank(stateprov) and isinstance(stateprov, str):
            if iso2 == "CA" and col_province:
                ws.cell(current, col_province).value = stateprov
            elif iso2 != "CA" and col_state:
                ws.cell(current, col_state).value = stateprov

    if not out_path:
        out_path = template_path

    wb.save(str(out_path))
    print(f"[OK] UPDATE WebADI filled: {out_path}")


LEGAL_NAME_CLEAN_RE = re.compile(r"[^A-Za-z0-9 ]+")

def inject_into_webadi_attach_sle(
        template_path: Path,
        out_path: Optional[Path],
        df: pd.DataFrame,
        mode: str = "replace",
        sheet_name: str = WEBADI_SHEET_DEFAULT,
        header_start_row: int = 5,
        debug: bool = False,
        unprotect: bool = True,
):
    """
    Fills the OC WEBADI ATTACH SLE template.

    Basic info only (no tax, no payment method, no currency)
    ACTION = "Independent to Independent"
    GROUP_PARENT_SLE_ACCOUNT_NUMBER = SLE OID
    """
    if not template_path.exists():
        raise FileNotFoundError(f"WEBADI ATTACH SLE template not found: {template_path}")

    wb = load_workbook(
        str(template_path),
        keep_vba=True,
        data_only=False,
        keep_links=True,
    )

    if unprotect:
        if debug:
            print("[DEBUG] webadi_unprotect=True: unprotecting workbook/sheets.")
        unprotect_workbook_and_sheets(wb, debug=debug)
    else:
        if debug:
            print("[DEBUG] webadi_unprotect=False: keeping protections.")

    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' not found in template.")
    ws = wb[sheet_name]

    # Update BATCH_NAME in E3 with today's date
    try:
        today_str = datetime.now().strftime("%m/%d/%Y")
        ws["E3"].value = f"OC WEBADI ATTACH SLE {today_str}"
        if debug:
            print(f"[DEBUG] E3 (BATCH_NAME) set to 'OC WEBADI ATTACH SLE {today_str}'")
    except Exception as e:
        if debug:
            print(f"[DEBUG] Failed to update BATCH_NAME in E3: {e}")

    header_row, header_map = find_header_row(
        ws, start_at_row=header_start_row, debug=debug
    )

    if debug:
        print(f"[DEBUG] ATTACH SLE WebADI: Found {len(header_map)} headers in template:")
        for norm_name, col_idx in sorted(header_map.items()):
            print(f"  {norm_name} → Column {col_idx}")

    # Mapping for ATTACH SLE template (basic info only, no tax)
    wanted = {
        "GROUP_PARENT_SLE_ACCOUNT_NUMBER": "SLE OID",
        "HOTEL_ID": "Expedia ID",
        "EFFECTIVE_DATE": "Effective Date of Change",
        "CUSTOMER_NAME": "Hotel Name",
        "ADDRESS_LINE_1": "Address Line 1",  # Changed from ADDRESS1 to match OID Creation
        "CITY": "City",
        "POSTAL_CODE": "Postal Code",
        "COUNTRY": "WEBADI_COUNTRY",
        "FIRST_NAME": "First Name",
        "LAST_NAME": "Last Name",
        "EMAIL_ADDRESS": "Email Address",  # Added per user request - from column U of extraction
        "COMMENTS": "file_name",  # Will be converted to comment format
    }

    col_state = header_map.get("STATE")
    col_province = header_map.get("PROVINCE")
    col_site_purpose = header_map.get("SITE_PURPOSE")
    col_action = header_map.get("ACTION")
    col_upl = column_index_from_string("B")

    key_col = header_map.get("HOTEL_ID") or header_map.get("CUSTOMER_NAME")
    if not key_col:
        raise RuntimeError(
            "Could not find key columns (HOTEL_ID/CUSTOMER_NAME) in ATTACH SLE WebADI header."
        )

    # Clear or clone rows according to mode
    if mode.lower() == "replace":
        last = ws.max_row
        for _ in range(header_row + 1, last + 1):
            ws.delete_rows(header_row + 1)
        base_row = header_row
        if debug:
            print(
                f"[DEBUG] Replace mode: cleared rows after header row {header_row}."
            )
    else:
        base_row = last_data_row(ws, header_row, key_col)
        if debug:
            print(
                f"[DEBUG] Append mode: last data row detected = {base_row}."
            )

    current = base_row

    for _, rec in df.iterrows():
        iso2 = to_iso2(rec.get("Country"))
        stateprov = rec.get("State/Province")

        current += 1
        src_row = base_row if base_row > header_row else header_row + 1
        clone_row(ws, src_row, current)

        # Set fixed columns for ATTACH SLE template
        ws.cell(current, col_upl).value = "O"
        if col_action:
            ws.cell(current, col_action).value = "Independent to Independent"
        if col_site_purpose:
            ws.cell(current, col_site_purpose).value = "BILL_TO"

        # Always clear STATE/PROVINCE in new rows to avoid inheriting junk
        if col_state:
            ws.cell(current, col_state).value = None
        if col_province:
            ws.cell(current, col_province).value = None

        # IMPORTANT: Clear tax-related columns (J-N) and payment/currency columns (AB, AC)
        # These should remain BLANK in ATTACH SLE template
        for col_letter in ["J", "K", "L", "M", "N", "AB", "AC"]:
            col_idx = column_index_from_string(col_letter)
            ws.cell(current, col_idx).value = None

        # Fill columns according to mapping (basic info only)
        for webadi_col, script_col in wanted.items():
            col_idx = header_map.get(webadi_col)
            if not col_idx:
                if debug:
                    print(f"[DEBUG] ATTACH SLE: Column '{webadi_col}' not found in template header, skipping")
                continue
            val = rec.get(script_col)
            if debug and webadi_col in ["FIRST_NAME", "LAST_NAME", "ADDRESS_LINE_1"]:
                print(f"[DEBUG] ATTACH SLE: Filling {webadi_col} (col {col_idx}) with value: {val}")

            # COMMENTS: Format as "filename Updated per HCPIF - Attach SLE"
            # Note: file_name already contains Expedia IDs, so don't duplicate them
            if webadi_col == "COMMENTS":
                file_name = rec.get("file_name", "")
                if not _is_blank(file_name):
                    comment_text = f"{file_name} Updated per HCPIF - Attach SLE"
                    ws.cell(current, col_idx).value = comment_text
                else:
                    ws.cell(current, col_idx).value = None

            # EFFECTIVE_DATE: Format as '15/Feb/2026' (DD/MMM/YYYY)
            elif webadi_col == "EFFECTIVE_DATE":
                if _is_blank(val):
                    ws.cell(current, col_idx).value = None
                else:
                    try:
                        if isinstance(val, datetime):
                            formatted_date = val.strftime("%d/%b/%Y")
                            ws.cell(current, col_idx).value = formatted_date
                        elif isinstance(val, str):
                            from dateutil import parser
                            parsed_date = parser.parse(val)
                            formatted_date = parsed_date.strftime("%d/%b/%Y")
                            ws.cell(current, col_idx).value = formatted_date
                        else:
                            ws.cell(current, col_idx).value = val
                    except:
                        ws.cell(current, col_idx).value = val

            # All other fields
            elif _is_blank(val):
                ws.cell(current, col_idx).value = None
            else:
                ws.cell(current, col_idx).value = val

        # Fill State/Province
        if not _is_blank(stateprov) and isinstance(stateprov, str):
            if iso2 == "CA" and col_province:
                ws.cell(current, col_province).value = stateprov
            elif iso2 != "CA" and col_state:
                ws.cell(current, col_state).value = stateprov

    if not out_path:
        out_path = template_path

    wb.save(str(out_path))
    print(f"[OK] ATTACH SLE WebADI filled: {out_path}")




def inject_into_oid_creation_webadi(
        template_path: Path,
        out_path: Optional[Path],
        df: pd.DataFrame,
        mode: str = "replace",
        sheet_name: str = WEBADI_SHEET_DEFAULT,
        header_start_row: int = 5,
        debug: bool = False,
        unprotect: bool = True,
):
    if not template_path.exists():
        raise FileNotFoundError(f"OID Creation template not found: {template_path}")

    wb = load_workbook(
        str(template_path),
        keep_vba=True,
        data_only=False,
        keep_links=True,
    )

    if unprotect:
        unprotect_workbook_and_sheets(wb, debug=debug)

    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' not found in OID Creation template.")
    ws = wb[sheet_name]

    # Header / batch
    try:
        ws["E3"] = f"SLE OID Creation {datetime.now().strftime('%m%d%Y')}"
    except Exception as e:
        if debug:
            print(f"[DEBUG] Failed to update E3 in OID Creation template: {e}")

    header_row, header_map = find_header_row(
        ws, start_at_row=header_start_row, debug=debug
    )

    wanted = {
        "HOTEL_ID": "Expedia ID",
        "CUSTOMER_NAME": "Legal Name",
        "ADDRESS_LINE_1": "Address Line 1",
        "CITY": "City",
        "POSTAL_CODE": "Postal Code",
        "COUNTRY": "Country ISO2",
        "FIRST_NAME": "First Name",
        "LAST_NAME": "Last Name",
        "EMAIL_ADDRESS": "Email Address",
    }

    fixed_values = {
        "SOURCE_SYSTEM": "ARLoader",
        "CUSTOMER_LEVEL": "1",
        "CUSTOMER_TYPE": "SLE",
        "BUSINESS_MODEL_CODE": "DIR",
        "ACCOUNT_STATUS": "ACTIVE",
        "INVOICE_TYPE": "INDIVIDUAL",
        "BILLTO_PARTY_FLAG": "Y",
        "BILLTO_ENTITY_TYPE": "SLE",
        "SITE_PURPOSE": "BILL_TO",
        "PRIMARY_SITE_FLAG": "Y",
        "PRIMARY_METHOD_FLAG": "Y",
        "CONTACT_ROLE": "BILL_TO",
        "PRIMARY_CONTACT_FLAG": "Y",
    }

    col_state = header_map.get("STATE")
    col_province = header_map.get("PROVINCE")

    col_billing_currency = header_map.get("BILLING_CURRENCY") or column_index_from_string("AS")
    col_payment_terms = column_index_from_string("AU")
    col_preferred_language = header_map.get("PREFERRED_LANGUAGE") or column_index_from_string("BI")

    key_col = header_map.get("HOTEL_ID") or header_map.get("CUSTOMER_NAME")
    if not key_col:
        raise RuntimeError(
            "Could not find key columns (HOTEL_ID/CUSTOMER_NAME) in OID Creation header."
        )

    if mode.lower() == "replace":
        last = ws.max_row
        for _ in range(header_row + 1, last + 1):
            ws.delete_rows(header_row + 1)
        base_row = header_row
    else:
        base_row = last_data_row(ws, header_row, key_col)

    current = base_row

    for _, rec in df.iterrows():
        iso2 = to_iso2(rec.get("Country"))
        stateprov = rec.get("State/Province")

        current += 1
        src_row = base_row if base_row > header_row else header_row + 1
        clone_row(ws, src_row, current)

        # clear state/province first
        if col_state:
            ws.cell(current, col_state).value = None
        if col_province:
            ws.cell(current, col_province).value = None

        # OID Creation fixed patterns
        if col_billing_currency:
            ws.cell(current, col_billing_currency).value = None

        ws.cell(current, col_payment_terms).value = "MONTHLY-14"
        ws.cell(current, col_preferred_language).value = "EN"

        # fixed/default values
        for webadi_col, fixed_val in fixed_values.items():
            col_idx = header_map.get(webadi_col)
            if col_idx:
                ws.cell(current, col_idx).value = fixed_val

        for webadi_col, script_col in wanted.items():
            col_idx = header_map.get(webadi_col)
            if not col_idx:
                continue

            val = rec.get(script_col)

            if _is_blank(val):
                ws.cell(current, col_idx).value = None
            else:
                if webadi_col == "BILLING_CURRENCY" and isinstance(val, str):
                    ws.cell(current, col_idx).value = val.strip().upper()
                else:
                    ws.cell(current, col_idx).value = val

        # State/Province
        if not _is_blank(stateprov) and isinstance(stateprov, str):
            if iso2 == "CA" and col_province:
                ws.cell(current, col_province).value = stateprov
            elif iso2 != "CA" and col_state:
                ws.cell(current, col_state).value = stateprov

    if not out_path:
        out_path = template_path

    wb.save(str(out_path))
    print(f"[OK] OID Creation WEBADI filled: {out_path}")


def add_country_iso2_column(df: pd.DataFrame) -> pd.DataFrame:
    if "Country" not in df.columns:
        return df

    iso2_values = []
    for v in df["Country"]:
        try:
            iso2_values.append(to_iso2(v))
        except Exception:
            iso2_values.append(None)

    insert_pos = list(df.columns).index("Country") + 1
    df.insert(loc=insert_pos, column="Country ISO2", value=iso2_values)
    return df


def add_language_abbreviation_column(df: pd.DataFrame) -> pd.DataFrame:
    if "Preferred Language" not in df.columns:
        return df

    lang_values = []
    for v in df["Preferred Language"]:
        try:
            lang_values.append(normalize_language_to_code(v))
        except Exception:
            lang_values.append(None)

    insert_pos = list(df.columns).index("Preferred Language") + 1
    df.insert(loc=insert_pos, column="Language Abbreviation", value=lang_values)
    return df


def extract_eid_from_filename(filename: Optional[str]) -> Optional[str]:
    """
    Extracts the Expedia ID from the filename pattern.
    Expected pattern: [case number (8-9 digits)][separator][Expedia ID]
    Ignores anything in parentheses.

    Examples:
      - "12345678-987654.pdf" -> "987654"
      - "153967337 23687.pdf" -> "23687"
      - "12345678_987654 (some notes).pdf" -> "987654"
      - "12345678 EID 987654.pdf" -> "987654"
    """
    if not filename or not isinstance(filename, str):
        return None

    # Remove file extension
    base = os.path.splitext(filename)[0]

    # Remove anything in parentheses
    base = re.sub(r'\([^)]*\)', '', base).strip()

    # Look for pattern: 8-9 digit case number followed by separator and then EID
    # The separator can be: -, _, space, or text like "EID", "eid", "-EID", etc.
    # Changed from \d{8} to \d{8,9} to handle 8 or 9 digit case numbers
    pattern = r'^\d{8,9}[\s\-_]*(?:EID|eid|Case)?[\s\-_]*(\d+)'
    match = re.search(pattern, base, re.IGNORECASE)

    if match:
        return match.group(1).strip()

    # Fallback: if no 8-9 digit pattern, look for any number after common separators
    # This handles cases where the case number might have different length
    pattern_fallback = r'^\d+[\s\-_]+(\d+)'
    match_fallback = re.search(pattern_fallback, base)

    if match_fallback:
        return match_fallback.group(1).strip()

    return None


def add_eid_validation_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Adds columns to validate that the Expedia ID in the filename matches
    the Expedia ID extracted from the PDF.

    Adds:
      - "File Name EID": Expedia ID extracted from filename
      - "EID Match Status": Validation result
    """
    if "file_name" not in df.columns or "Expedia ID" not in df.columns:
        return df

    df = df.copy()

    # Extract EID from filename
    df["File Name EID"] = df["file_name"].apply(extract_eid_from_filename)

    # Compare and create status
    def _validate_eid(row):
        pdf_eid = str(row.get("Expedia ID", "")).strip()
        file_eid = str(row.get("File Name EID", "")).strip()

        # Handle empty/missing values
        if not pdf_eid or pdf_eid in ("", "nan", "None", "<NA>"):
            return "PDF EID Missing"

        if not file_eid or file_eid in ("", "nan", "None", "<NA>"):
            return "File Name EID Missing"

        # Compare
        if pdf_eid == file_eid:
            return "Match"
        else:
            return "EID Provided in the PDF doesn't match"

    df["EID Match Status"] = df.apply(_validate_eid, axis=1)

    # FALLBACK LOGIC: Use File Name EID when PDF EID is wrong/missing
    # Store original PDF EID before overwriting
    df["Original PDF EID"] = df["Expedia ID"].copy()

    def _apply_fallback(row):
        """
        If PDF EID doesn't match filename EID, use filename EID as fallback.
        This allows processing to continue while keeping the issue flagged for review.
        """
        status = row.get("EID Match Status", "")
        file_eid = str(row.get("File Name EID", "")).strip()
        pdf_eid = str(row.get("Expedia ID", "")).strip()

        # Use filename EID as fallback in these cases:
        # 1. PDF EID is missing but filename has one
        # 2. PDF EID doesn't match filename EID
        if status == "PDF EID Missing" and file_eid and file_eid not in ("", "nan", "None"):
            return file_eid
        elif status == "EID Provided in the PDF doesn't match" and file_eid and file_eid not in ("", "nan", "None"):
            return file_eid
        else:
            # Keep original PDF EID
            return pdf_eid

    # Apply fallback: overwrite Expedia ID with filename EID when needed
    df["Expedia ID"] = df.apply(_apply_fallback, axis=1)

    # Insert these columns right after Expedia ID
    if "Expedia ID" in df.columns:
        eid_pos = list(df.columns).index("Expedia ID") + 1

        # Remove from current position and insert after Expedia ID
        original_eid_col = df.pop("Original PDF EID")
        file_eid_col = df.pop("File Name EID")
        status_col = df.pop("EID Match Status")

        df.insert(loc=eid_pos, column="Original PDF EID", value=original_eid_col)
        df.insert(loc=eid_pos + 1, column="File Name EID", value=file_eid_col)
        df.insert(loc=eid_pos + 2, column="EID Match Status", value=status_col)

    return df


def create_eid_mismatch_review_tab(df: pd.DataFrame) -> pd.DataFrame:
    """
    Creates a review tab for rows where the EID from the PDF doesn't match
    the EID from the filename.

    IMPORTANT: The script automatically uses the filename EID as fallback,
    so processing continues, but these rows are flagged for manual review.

    Returns DataFrame with rows that need manual review.
    """
    if "EID Match Status" not in df.columns:
        return pd.DataFrame()

    # Filter rows where EID doesn't match OR is missing from PDF
    mismatch_df = df[
        (df["EID Match Status"] == "EID Provided in the PDF doesn't match") |
        (df["EID Match Status"] == "PDF EID Missing")
    ].copy()

    if mismatch_df.empty:
        return pd.DataFrame()

    # Add explanation column with details about what happened
    def _explain_fallback(row):
        status = row.get("EID Match Status", "")
        original = row.get("Original PDF EID", "")
        file_eid = row.get("File Name EID", "")
        current = row.get("Expedia ID", "")

        if status == "PDF EID Missing":
            return (
                f"PDF had no Expedia ID. Using filename EID ({file_eid}) as fallback. "
                "Verify this is correct before uploading to Oracle."
            )
        elif status == "EID Provided in the PDF doesn't match":
            return (
                f"PDF EID ({original}) does NOT match filename EID ({file_eid}). "
                f"Using filename EID ({current}) as fallback. "
                "Please verify which EID is correct before uploading to Oracle."
            )
        else:
            return "Unknown status - manual review required"

    mismatch_df["Review Reason"] = mismatch_df.apply(_explain_fallback, axis=1)
    mismatch_df["Action Taken"] = "Used filename EID as fallback - row will process normally"

    return mismatch_df


def sanitize_legal_name_column(df: pd.DataFrame) -> pd.DataFrame:
    if "Legal Name" not in df.columns:
        return df

    def _clean_ln(x):
        if x in (None, "", " "):
            return x
        x = unidecode(str(x))
        x = LEGAL_NAME_CLEAN_RE.sub("", x)
        return x.strip()

    df["Legal Name"] = df["Legal Name"].astype(str).apply(_clean_ln)
    return df


def load_country_tax_mapping() -> pd.DataFrame:
    """
    Reads the mapping spreadsheet (COUNTRY, COUNTRY_NAME, Tax Regime Code, Tax)
    and returns a clean DataFrame.
    """
    path = Path(COUNTRY_TAX_MAPPING_FILE)
    if not path.exists():
        print(f"[WARN] COUNTRY_TAX_MAPPING_FILE not found: {COUNTRY_TAX_MAPPING_FILE}")
        return pd.DataFrame()

    df_map = pd.read_excel(
        path,
        sheet_name=COUNTRY_TAX_MAPPING_SHEET,
        engine="pyxlsb",
        dtype=str,
    )

    # Expected columns:
    # A: COUNTRY
    # B: COUNTRY_NAME
    # C: Tax Regime Code
    # D: Tax
    col_rename = {}
    for c in df_map.columns:
        c_norm = str(c).strip().upper().replace(" ", "_")
        if c_norm == "COUNTRY":
            col_rename[c] = "MAP_COUNTRY"
        elif c_norm == "COUNTRY_NAME":
            col_rename[c] = "COUNTRY_NAME"
        elif c_norm in ("TAX_REGIME_CODE", "TAX_REGIME_CODE"):
            col_rename[c] = "TAX_REGIME_CODE"
        elif c_norm == "TAX":
            col_rename[c] = "TAX"
    df_map = df_map.rename(columns=col_rename)

    needed = {"MAP_COUNTRY", "COUNTRY_NAME", "TAX_REGIME_CODE", "TAX"}
    if not needed.issubset(df_map.columns):
        print("[WARN] Country/Tax mapping does not have all expected columns "
              "(COUNTRY, COUNTRY_NAME, Tax Regime Code, Tax).")
        return pd.DataFrame()

    for c in ["MAP_COUNTRY", "COUNTRY_NAME", "TAX_REGIME_CODE", "TAX"]:
        df_map[c] = df_map[c].astype(str).str.strip()

    df_map = df_map[df_map["COUNTRY_NAME"] != ""].copy()
    return df_map


def enrich_df2_with_country_tax(df2: pd.DataFrame) -> pd.DataFrame:
    """
    Performs the 'VLOOKUP':
      df2['Country']  --> mapping['COUNTRY_NAME']

    Adds columns:
      - WEBADI_COUNTRY         (goes to COUNTRY column in WEBADI)
      - WEBADI_TAX_REGIME_CODE (goes to TAX_REGIME_CODE in WEBADI)
      - WEBADI_TAX             (goes to TAX in WEBADI)
    """
    df_map = load_country_tax_mapping()
    if df_map.empty or "Country" not in df2.columns:
        return df2

    df2 = df2.copy()

    df2["Country_norm"] = df2["Country"].astype(str).str.strip().str.upper()
    df_map["COUNTRY_NAME_norm"] = df_map["COUNTRY_NAME"].astype(str).str.strip().str.upper()

    df2 = df2.merge(
        df_map[["COUNTRY_NAME_norm", "MAP_COUNTRY", "TAX_REGIME_CODE", "TAX"]],
        how="left",
        left_on="Country_norm",
        right_on="COUNTRY_NAME_norm",
    )

    df2["WEBADI_COUNTRY"] = df2["MAP_COUNTRY"]
    df2["WEBADI_TAX_REGIME_CODE"] = df2["TAX_REGIME_CODE"]
    df2["WEBADI_TAX"] = df2["TAX"]

    df2 = df2.drop(columns=["Country_norm", "COUNTRY_NAME_norm", "MAP_COUNTRY", "TAX_REGIME_CODE", "TAX"])
    return df2


# -------------------- MAIN --------------------
def main():
    ap = argparse.ArgumentParser(
        description="Extract HCPIF fields into Excel (R2.6 + WEBADI injection + OCR fallback + Oracle enrich)."
    )
    ap.add_argument("--input_dir", type=str, default=str(DEFAULT_INPUT_DIR))
    ap.add_argument("--output", type=str, default=str(DEFAULT_OUTPUT_FILE))
    ap.add_argument(
        "--webadi_template",
        type=str,
        default=None,
        help="Path to WEBADI .xlsm/.xlsx template",
    )
    ap.add_argument(
        "--webadi_output",
        type=str,
        default=None,
        help="Path to save filled copy. "
             "If omitted, generates 'OC WEBADI MMDDYYYY.xlsm' in the same folder as the template.",
    )
    ap.add_argument(
        "--webadi_mode",
        type=str,
        choices=["append", "replace"],
        default="replace",
        help="append (clone last row); replace (clear data region before writing).",
    )
    ap.add_argument(
        "--webadi_sheet",
        type=str,
        default=WEBADI_SHEET_DEFAULT,
        help="WEBADI sheet name (default: WebADI).",
    )
    ap.add_argument(
        "--webadi_header_start_row",
        type=int,
        default=5,
        help="Row to start searching for header (default: 5).",
    )
    ap.add_argument(
        "--webadi_debug",
        action="store_true",
        help="Prints header/sheet diagnostics.",
    )
    ap.add_argument(
        "--webadi_unprotect",
        dest="webadi_unprotect",
        action="store_true",
        help="Unprotect workbook/sheets before writing (default: True).",
    )
    ap.add_argument(
        "--webadi_keep_protection",
        dest="webadi_unprotect",
        action="store_false",
        help="Do NOT unprotect workbook/sheets before writing.",
    )

    ap.set_defaults(webadi_unprotect=True)

    args = ap.parse_args()

    # ===== NEW: local folder by date (mm.dd.yyyy) =====
    today_str = datetime.now().strftime("%m.%d.%Y")

    base_local_dir = Path(args.input_dir or DEFAULT_INPUT_DIR)
    run_dir = base_local_dir / today_str
    run_dir.mkdir(parents=True, exist_ok=True)

    input_dir = run_dir
    output_path = run_dir / "HCPIF_extraction.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"[INFO] Local run folder: {run_dir}")

    # ===== Check for required templates BEFORE processing =====
    print("\n[INFO] Checking for required templates...")
    missing_templates = []

    update_template = Path(DEFAULT_WEBADI_UPDATE_TEMPLATE)
    if not update_template.exists():
        missing_templates.append(f"  - OC WEBADI UPDATE template: {update_template}")

    attach_sle_template = Path(DEFAULT_WEBADI_ATTACH_SLE_TEMPLATE)
    if not attach_sle_template.exists():
        missing_templates.append(f"  - OC WEBADI ATTACH SLE template: {attach_sle_template}")

    oid_template = Path(DEFAULT_OID_CREATION_TEMPLATE)
    if not oid_template.exists():
        missing_templates.append(f"  - OID Creation template: {oid_template}")

    tax_mapping = Path(COUNTRY_TAX_MAPPING_FILE)
    if not tax_mapping.exists():
        missing_templates.append(f"  - Tax mapping file: {tax_mapping}")

    if missing_templates:
        print("[WARN] The following required files are missing:")
        for item in missing_templates:
            print(item)
        print("\n[WARN] WebADI files will NOT be generated without templates.")
        print(f"[WARN] Please ensure all templates exist in: {base_local_dir}")
        user_input = input("\nContinue anyway? (y/n): ")
        if user_input.lower() != 'y':
            print("[INFO] Aborting script. Please restore templates and try again.")
            sys.exit(0)
    else:
        print("[OK] All templates found!")

    # ===== NEW: download PDFs from SharePoint (PDFs Ready to Load) =====
    print("\n[INFO] Downloading PDFs from SharePoint input folder...")
    downloaded = sp_download_pdfs_from_folder(SP_INPUT_FOLDER, input_dir)
    print(f"[INFO] {len(downloaded)} PDF(s) downloaded to {input_dir}")

    if not input_dir.exists():
        print(f"[ERROR] Input folder not found: {input_dir}")
        sys.exit(1)

    pdf_files = find_pdfs(input_dir)
    if not pdf_files:
        print(f"[ERROR] No PDFs found in: {input_dir}")
        sys.exit(2)

    rows = []

    print(f"\nReading {len(pdf_files)} PDF(s) from: {input_dir}\n")
    for pdf in tqdm.tqdm(pdf_files, desc="Processing PDFs"):
        try:
            data = extract_fields_positional(pdf)
            data["file_name"] = pdf.name
            data["file_path"] = str(pdf.resolve())
            rows.append(data)
        except Exception as e:
            print(f"[ERROR] Failed to process {pdf.name}: {e}")
            rows.append(
                {
                    "file_name": pdf.name,
                    "file_path": str(pdf.resolve()),
                    "status": f"error: {e}",
                }
            )

    df = pd.DataFrame(rows)

    ordered_cols = [f["col"] for f in FIELDS]
    for col in ordered_cols:
        if col not in df.columns:
            df[col] = pd.NA

    aux_cols = [c for c in df.columns if c not in ordered_cols]
    df = df[ordered_cols + aux_cols]

    def _strip_cell(v):
        if isinstance(v, str):
            return strip_leading_junk(v)
        return v

    df = df.apply(lambda col: col.map(_strip_cell))

    df = add_country_iso2_column(df)
    df = add_language_abbreviation_column(df)
    df = add_eid_validation_columns(df)  # Validate EID from filename vs PDF

    # Create EID mismatch review tab BEFORE enrichment
    df_eid_mismatch = create_eid_mismatch_review_tab(df)

    df = add_sle_oid_from_legal_name(df)

    base_columns = list(df.columns)

    df, df_excluded, df_oracle_details, df_no_oid_ec, df_sle_review = enrich_hcpif_with_oracle(df, base_columns)

    df = sanitize_legal_name_column(df)
    if not df_excluded.empty:
        df_excluded = sanitize_legal_name_column(df_excluded)

    # ── NEW: SF Report "Rebill Request" check ──────────────────────────
    df, df_rebill = check_sf_report_rebill(df, run_dir)
    # df_excluded also needs to have rebill rows removed
    if not df_excluded.empty and not df_rebill.empty:
        rebill_eids = set(
            df_rebill["Expedia ID"].astype(str).str.strip()
            if "Expedia ID" in df_rebill.columns else []
        )
        df_excluded = df_excluded[
            ~df_excluded["Expedia ID"].astype(str).str.strip().isin(rebill_eids)
        ].copy()
    # ── END NEW ─────────────────────────────────────────────────────────────

    oid_creation_template_path = None
    default_oid_tpl = Path(DEFAULT_OID_CREATION_TEMPLATE)
    if default_oid_tpl.exists():
        oid_creation_template_path = default_oid_tpl

    oid_creation_out_path = None
    df_oid_creation = pd.DataFrame()

    oid_need_cols = [
        c for c in [
            "Expedia ID",
            "Legal Name",
            "Address Line 1",
            "City",
            "State/Province",
            "Postal Code",
            "Country",
            "Country ISO2",
            "First Name",
            "Last Name",
            "Email Address",
            "SLE OID",
        ] if c in df.columns
    ]

    df_oid_creation = df.loc[
        (df["Expedia ID"].notna()) | (df["Legal Name"].notna()),
        oid_need_cols,
    ].copy()

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Extraction", index=False)
            if not df_excluded.empty:
                df_excluded.to_excel(
                    writer, sheet_name="Excluded_DIRECTDEBIT_TAIBM", index=False
                )
            if not df_oracle_details.empty:
                df_oracle_details.to_excel(
                    writer, sheet_name="Oracle_Details", index=False
                )
            # ── NEW: write rebill tab ───────────────────────────────────────
            if not df_rebill.empty:
                df_rebill.to_excel(
                    writer, sheet_name="Review - Existing Rebill", index=False
                )
                print(f"[INFO] {len(df_rebill)} row(s) written to 'Review - Existing Rebill' tab.")
            # ── END NEW ────────────────────────────────────────────────────

            # ── NEW: write NO OID - EC tab ──────────────────────────────────
            if not df_no_oid_ec.empty:
                df_no_oid_ec.to_excel(
                    writer, sheet_name="NO OID - EC", index=False
                )
                print(f"[INFO] {len(df_no_oid_ec)} row(s) written to 'NO OID - EC' tab.")
            # ── END NEW ────────────────────────────────────────────────────

            # ── NEW: write Review - SLE Match tab ──────────────────────────
            if not df_sle_review.empty:
                df_sle_review.to_excel(
                    writer, sheet_name="Review - SLE Match", index=False
                )
                print(f"[INFO] {len(df_sle_review)} row(s) written to 'Review - SLE Match' tab.")
            # ── END NEW ─────────────────────────────────────────────────────

            # ── NEW: write Review - EID Mismatch tab ────────────────────────
            if not df_eid_mismatch.empty:
                df_eid_mismatch.to_excel(
                    writer, sheet_name="Review - EID Mismatch", index=False
                )
                print(f"[INFO] {len(df_eid_mismatch)} row(s) written to 'Review - EID Mismatch' tab.")
            # ── END NEW ─────────────────────────────────────────────────────

        print("\n[OK] Done!")
        print(f"- PDFs read from: {input_dir}")
        print(f"- Excel saved to: {output_path}\n")
    except PermissionError:
        print("\n[ERROR] Could not write the Excel file.")
        print(f"Destination: {output_path}")
        print("Common causes:")
        print("  1) File is currently OPEN in Excel.")
        print("  2) You lack permission.")
        raise

    # ===== WEBADI GENERATION: 3 FILES =====
    # Prepare data for WebADI files
    need_cols = {
        "Expedia ID", "SLE OID", "Legal Name", "Hotel Name",
        "Address Line 1", "City", "State/Province", "Postal Code", "Country",
        "Currency", "Tax Registration Number", "Tax Registration Status",
        "Effective Date of Change", "Current RM", "file_name",
        "First Name", "Last Name", "Email Address", "Preferred Language",
    }

    present = [c for c in need_cols if c in df.columns]
    # Always include file_name if it exists
    if "file_name" in df.columns and "file_name" not in present:
        present.append("file_name")
    if present:
        df_webadi = df.loc[
            (df["Expedia ID"].notna()) | (df["Legal Name"].notna()),
            present,
        ].copy()

        if len(df_webadi) > 0:
            # Apply VLOOKUP Country/Tax enrichment
            df_webadi = enrich_df2_with_country_tax(df_webadi)
            today_str_for_webadi = datetime.now().strftime("%m%d%Y")

            # FILE 1: OC WEBADI UPDATE
            update_template = Path(DEFAULT_WEBADI_UPDATE_TEMPLATE)
            if update_template.exists():
                update_out_path = update_template.parent / f"OC WEBADI UPDATE {today_str_for_webadi}.xlsm"
                print(f"\n[INFO] Generating UPDATE WebADI...")
                print(f"  Template: {update_template}")
                print(f"  Output: {update_out_path}")
                print(f"  Rows: {len(df_webadi)}")
                try:
                    inject_into_webadi_update(
                        update_template,
                        update_out_path,
                        df_webadi,
                        mode="replace",
                        sheet_name="WebADI",
                        header_start_row=5,
                        debug=args.webadi_debug,
                        unprotect=args.webadi_unprotect,
                    )
                except Exception as ex:
                    print(f"[ERROR] Failed to fill UPDATE WebADI: {ex}")
            else:
                print(f"[WARN] UPDATE template not found: {update_template}")

            # FILE 2: OC WEBADI ATTACH SLE
            attach_sle_template = Path(DEFAULT_WEBADI_ATTACH_SLE_TEMPLATE)
            if attach_sle_template.exists():
                attach_sle_out_path = attach_sle_template.parent / f"OC WEBADI Attach SLE {today_str_for_webadi}.xlsm"
                print(f"\n[INFO] Generating ATTACH SLE WebADI...")
                print(f"  Template: {attach_sle_template}")
                print(f"  Output: {attach_sle_out_path}")
                print(f"  Rows: {len(df_webadi)}")
                try:
                    inject_into_webadi_attach_sle(
                        attach_sle_template,
                        attach_sle_out_path,
                        df_webadi,
                        mode="replace",
                        sheet_name="WebADI",
                        header_start_row=5,
                        debug=args.webadi_debug,
                        unprotect=args.webadi_unprotect,
                    )
                except Exception as ex:
                    print(f"[ERROR] Failed to fill ATTACH SLE WebADI: {ex}")
            else:
                print(f"[WARN] ATTACH SLE template not found: {attach_sle_template}")
        else:
            print("[INFO] No rows available for WebADI generation.")
    else:
        print("[WARN] Required columns for WebADI not found in extraction.")

    if oid_creation_template_path:
        if not df_oid_creation.empty:
            today_str_for_oid = datetime.now().strftime("%m%d%Y")
            oid_creation_out_path = oid_creation_template_path.parent / f"OID Creation {today_str_for_oid}.xlsm"

            print(f"[INFO] OID Creation template: {oid_creation_template_path}")
            print(f"[INFO] OID Creation output: {oid_creation_out_path}")
            print(f"[INFO] OID Creation rows: {len(df_oid_creation)}")

            try:
                inject_into_oid_creation_webadi(
                    oid_creation_template_path,
                    oid_creation_out_path,
                    df_oid_creation,
                    mode="replace",
                    sheet_name="WebADI",
                    header_start_row=5,
                    debug=args.webadi_debug,
                    unprotect=args.webadi_unprotect,
                )
            except Exception as ex:
                print(f"[ERROR] Failed to fill OID Creation WEBADI: {ex}")
        else:
            print("[INFO] No rows available for OID Creation WEBADI.")
    else:
        print("[INFO] OID Creation template not found. Skipping additional WEBADI.")

    # ===== NEW: Upload results to SharePoint =====
    try:
        sp_output_folder = f"{SP_OUTPUT_BASE}/{today_str}"

        print(f"[INFO] Uploading HCPIF_extraction.xlsx to SharePoint: {sp_output_folder}")
        sp_upload_file(output_path, sp_output_folder)

        # Upload UPDATE WebADI
        if 'update_out_path' in locals() and update_out_path is not None and update_out_path.exists():
            print(f"[INFO] Uploading UPDATE WebADI to SharePoint: {sp_output_folder}")
            sp_upload_file(update_out_path, sp_output_folder)

        # Upload ATTACH SLE WebADI
        if 'attach_sle_out_path' in locals() and attach_sle_out_path is not None and attach_sle_out_path.exists():
            print(f"[INFO] Uploading ATTACH SLE WebADI to SharePoint: {sp_output_folder}")
            sp_upload_file(attach_sle_out_path, sp_output_folder)

        # Upload OID Creation WebADI
        if 'oid_creation_out_path' in locals() and oid_creation_out_path is not None and oid_creation_out_path.exists():
            print(f"[INFO] Uploading OID Creation WEBADI to SharePoint: {sp_output_folder}")
            sp_upload_file(oid_creation_out_path, sp_output_folder)

        # Archive input PDFs
        for pdf in pdf_files:
            print(f"[INFO] Archiving PDF to SharePoint: {SP_ARCHIVE_FOLDER} -> {pdf.name}")
            sp_upload_file(pdf, SP_ARCHIVE_FOLDER)

        print("[INFO] SharePoint upload completed (HCPIF, 3 WebADIs, PDFs archived).")
    except Exception as e:
        print(f"[WARN] Failed to upload results to SharePoint: {e}")


if __name__ == "__main__":
    main()
