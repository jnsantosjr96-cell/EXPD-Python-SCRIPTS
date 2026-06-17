"""
COA vs WO Reconciliation Script (Refactored)

This script reconciles Customer On Account (COA) balances against Write-Off (WO) adjustments
in Oracle E-Business Suite, generating debit memos and adjustment files for finance operations.

Main workflow:
1. Extract customer list from COA report
2. Query Oracle for WO adjustments and RELO credit memos
3. Match COA receipts against WO transactions
4. Generate output files: CSV adjustments, RELO DM templates, HC DM templates
5. Send summary email via Outlook

Author: Jose Santos (josenjr@expediagroup.com)
"""

import os
import sys
import logging
import re
from pathlib import Path
from datetime import date
from typing import Generator, Optional, List, Tuple
from dataclasses import dataclass

import oracledb
import pandas as pd
from openpyxl import load_workbook

# Optional Outlook integration
try:
    import win32com.client as win32
    OUTLOOK_AVAILABLE = True
except ImportError:
    win32 = None
    OUTLOOK_AVAILABLE = False

# ==========================
# LOGGING CONFIGURATION
# ==========================

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# ==========================
# CONSTANTS
# ==========================

# File paths
DOWNLOADS_DIR = Path(os.path.expanduser("~")) / "Downloads"
BASE_OUTPUT_DIR = Path(
    r"C:\Users\josenjr\OneDrive - Expedia Group\Desktop\Daily activities\COA vs WO Process"
)
CONCAT_PATH = BASE_OUTPUT_DIR / "Concat.xlsx"
RELO_DM_TEMPLATE = BASE_OUTPUT_DIR / "4 Non-Lodging Relo CM.xlsm"
HC_DM_TEMPLATE = BASE_OUTPUT_DIR / "3 HC Relo Upload DM.xlsm"

# Oracle configuration
ORACLE_CLIENT_DIR = r"C:\Users\josenjr\OneDrive - Expedia Group\Desktop\Oracle Instant Client\instantclient_23_0"
USERNAME = os.getenv("ORACLE_USER", "josenjr")
PASSWORD = os.getenv("ORACLE_PASSWORD", "XXXXXXXX")  # TODO: Move to environment variable
DSN = os.getenv("ORACLE_DSN", "ashworaebsdb02-vip.datawarehouse.expecn.com:1526/ORAPRD_UI")

# Processing constants
BATCH_SIZE = 1000
CONCAT_START_ROW = 2
TEMPLATE_ROW_RELO = 10
WO_DATA_COLUMN_INDEX = 19

# Sheet names
SHEET_DATA = "Data"
SHEET_CUST_LIST = "CustList"
SHEET_WO = "WO"
SHEET_RELO_CMS = "Relo CMs"
SHEET_MATCHES = "Matches"
SHEET_NL_REVIEW = "NL review"
SHEET_MATCH_RELO = "Match RELO"
SHEET_CMS_REVERSED = "CMs Already Reversed"
SHEET_UNKNOWN_MATCHES = "Unknown OID Matches"
SHEET_NEGATIVE_DAYS = "Negative Days"
SHEET_EVC_RELO_CM = "EVC RELO CM"

# Entity mappings
ENTITY_US = "US"
ENTITY_CH = "CH"
ENTITY_TS = "TS"
ENTITY_BR = "BR"

# Email configuration
EMAIL_TO = "hotelcollectbilling@expedia.com"

# ==========================
# DATA CLASSES
# ==========================

@dataclass
class OutputFiles:
    """Container for generated output file paths."""
    coa_workbook: Path
    cash_xlsx: Optional[Path] = None
    csv_adjustments: Optional[Path] = None
    relo_dm: Optional[Path] = None
    hc_dm: Optional[Path] = None

    def get_attachments(self) -> List[Path]:
        """Return list of non-None file paths for email attachments."""
        return [f for f in [self.cash_xlsx, self.csv_adjustments, self.relo_dm, self.hc_dm] if f]

# ==========================
# SQL QUERIES
# ==========================

# WO adjustments query
BASE_QUERY = """
SELECT
CTA.TRX_NUMBER              AS "Transaction Number",
CTTA.NAME                   AS "Transaction Type",
CTA.TRX_DATE                AS "Transaction Date",
HCA.ACCOUNT_NUMBER          AS "Account Number",
CTA.INVOICE_CURRENCY_CODE   AS "Entered Currency",
AAA.AMOUNT                  AS "Entered Amount",
AAA.ADJUSTMENT_NUMBER       AS "Adjustment Number",
AAA.REASON_CODE             AS "Reason Code"
FROM AR.AR_ADJUSTMENTS_ALL      AAA
JOIN AR.RA_CUSTOMER_TRX_ALL     CTA  ON CTA.CUSTOMER_TRX_ID   = AAA.CUSTOMER_TRX_ID
JOIN AR.RA_CUST_TRX_TYPES_ALL   CTTA ON CTTA.CUST_TRX_TYPE_ID = CTA.CUST_TRX_TYPE_ID
JOIN AR.HZ_CUST_ACCOUNTS        HCA  ON HCA.CUST_ACCOUNT_ID   = CTA.BILL_TO_CUSTOMER_ID
JOIN APPS.HR_OPERATING_UNITS    HOU  ON HOU.ORGANIZATION_ID   = CTA.ORG_ID
JOIN APPS.GL_LEDGER_LE_V        GLL  ON GLL.LEGAL_ENTITY_ID   = HOU.DEFAULT_LEGAL_CONTEXT_ID
WHERE
GLL.LEDGER_CATEGORY_CODE = 'PRIMARY'
AND CTTA.NAME NOT LIKE '%EAC%'
AND UPPER(CTTA.NAME) NOT LIKE '%GROUP%'
AND (
INSTR(HOU.NAME, '11105') > 0
OR INSTR(HOU.NAME, '11115') > 0
OR INSTR(HOU.NAME, '12305') > 0
OR INSTR(HOU.NAME, '14101') > 0
OR INSTR(HOU.NAME, '23310') > 0
OR INSTR(HOU.NAME, '72110') > 0
)
AND HCA.ACCOUNT_NUMBER IN ({placeholders})
ORDER BY
HCA.ACCOUNT_NUMBER,
CTA.TRX_NUMBER,
AAA.ADJUSTMENT_NUMBER
"""

# RELO CMs query (with Reference)
RELO_QUERY = """
SELECT
CTA.TRX_NUMBER                  AS "Transaction Number",
CTTA.NAME                       AS "Transaction Type",
HCA.ACCOUNT_NUMBER              AS "Account Number",
CTA.INVOICE_CURRENCY_CODE       AS "Entered Currency",
PSA.AMOUNT_DUE_ORIGINAL         AS "Entered Amount",
PSA.AMOUNT_DUE_REMAINING        AS "Open Balance",
CTA.REASON_CODE                 AS "Reason Code",
CTA.TRX_DATE                    AS "Transaction Date",
CTA.INTERFACE_HEADER_ATTRIBUTE1 AS "Reference"
FROM AR.RA_CUSTOMER_TRX_ALL      CTA
JOIN AR.RA_CUST_TRX_TYPES_ALL    CTTA ON CTTA.CUST_TRX_TYPE_ID = CTA.CUST_TRX_TYPE_ID
JOIN AR.HZ_CUST_ACCOUNTS         HCA  ON HCA.CUST_ACCOUNT_ID   = CTA.BILL_TO_CUSTOMER_ID
JOIN AR.AR_PAYMENT_SCHEDULES_ALL PSA  ON PSA.CUSTOMER_TRX_ID   = CTA.CUSTOMER_TRX_ID
JOIN APPS.HR_OPERATING_UNITS     HOU  ON HOU.ORGANIZATION_ID   = CTA.ORG_ID
JOIN APPS.GL_LEDGER_LE_V         GLL  ON GLL.LEGAL_ENTITY_ID   = HOU.DEFAULT_LEGAL_CONTEXT_ID
WHERE
GLL.LEDGER_CATEGORY_CODE = 'PRIMARY'
AND PSA.CLASS = 'CM'
AND CTA.REASON_CODE LIKE '%OFFSET_ACCRUED_AR%'
AND CTTA.NAME IN (
'CH_DIR_CM_RELO_USD',
'US_DIR_CM_RELO_USD',
'TS_DIR_CM_RELO_USD',
'BR_DIR_CM_RELO_BRL',
'TS_CM_RELO'
)
AND HCA.ACCOUNT_NUMBER IN ({placeholders})
ORDER BY
CTA.TRX_DATE DESC,
CTA.TRX_NUMBER
"""

# Transaction Register for CMs in Match RELO tab with applied invoice
CM_APPLIED_QUERY = """
/*
* Oracle Transaction Register - UAT
* Author: Jose Santos (josenjr@expediagroup.com)
*/

SELECT

CASE
  WHEN HZP.COUNTRY in ('AS','AU','BD','BN','BT','CC','CK','CN','FJ','FM','GU','HK','ID','IN','JP','KH','KP','KR','LA','LK','MH','MM','MN','MO','MP','MV','MY','NC','NF','NP','NU','NZ','PF','PG','PH','PK','PW','SB','SG','TH','TK','TL','TO','TW','VN','VU','WF','WS')
  THEN 'APAC'
  WHEN HZP.COUNTRY in ('AG','AI','AN','AR','AW','BB','BL','BM','BO','BQ','BR','BS','BZ','CL','CO','CR','CU','CW','DM','DO','EC','FK','GD','GF','GP','GT','GY','HN','HT','JM','KN','KY','LC','MF','MQ','MS','MX','NI','PA','PE','PN','PR','PY','SR','SV','SX','TC','TT','UY','VC','VE','VG','VI')
  THEN 'LATAM'
  WHEN HZP.COUNTRY in ('AL','AM','AT','AZ','BA','BE','BG','BY','CH','CZ','DE','DK','EE','FI','FO','FR','GB','GE','GG','GL','HR','HU','IE','IS','JE','KG','KZ','LI','LT','LU','LV','MD','ME','MK','NL','NO','PL','PM','RO','RS','RU','SE','SH','SI','SJ','SK','TJ','TM','UA','UZ')
  THEN 'N-EMEA'
  WHEN HZP.COUNTRY in ('AD','AE','AF','AO','BF','BH','BI','BJ','BV','BW','CD','CF','CG','CI','CM','CV','CY','DJ','DZ','EG','ER','ES','ET','GA','GH','GI','GM','GN','GQ','GR','GW','IL','IQ','IR','IT','JO','KE','KM','KW','LB','LR','LS','LY','MA','MC','MG','ML','MR','MT','MU','MW','MZ','NA','NE','NG','OM','PS','PT','QA','RE','RW','SA','SC','SD','SL','SM','SN','SS','ST','SY','SZ','TD','TG','TN','TR','TZ','UG','VA','YE','YT','ZA','ZM','ZW')
  THEN 'S-EMEA'
  WHEN HZP.COUNTRY in ('CA','US')
  THEN 'NAMER'
  ELSE 'NOT_DEFINED'
  END AS "POAR Super Region"

,GLL.LEDGER_NAME "Ledger"
,HOU.NAME "OU NAME"
,SUBSTR(HOU.NAME,6,51) "Legal Entity"
,SUBSTR(HOU.NAME,1,5) "Company"
,PSA.INVOICE_CURRENCY_CODE "Currency"
,PSA.CLASS "CLASS"
,CTA.ATTRIBUTE4 "XLR ID"
,CTA.TRX_NUMBER "Transaction Number"
,CTTA.NAME "TRX Type Name"
,HZP.PARTY_NAME "Customer Name"
,HCA.ACCOUNT_NUMBER "Account Number"
,CTA.TRX_DATE "TRX DATE"
,PSA.GL_DATE "GL Date"
,TO_CHAR(CTA.TRX_DATE,'YYYY-MM')  "Year / Month"
,PSA.AMOUNT_DUE_ORIGINAL "Entered Amt"
,PSA.AMOUNT_DUE_REMAINING "Open Balance"
,CASE
    WHEN SUBSTR(CTTA.NAME,1,2) in ('CH','US')
    THEN ROUND(CTA.EXCHANGE_RATE * PSA.AMOUNT_DUE_ORIGINAL,2)
    ELSE PSA.AMOUNT_DUE_ORIGINAL
    END AS "Functional Amt"
,CASE
    WHEN PSA.INVOICE_CURRENCY_CODE = 'USD'
    THEN 1
    ELSE ROUND(GLDR.CONVERSION_RATE,15)
    END AS "Conversion Rate"
,CASE WHEN PSA.INVOICE_CURRENCY_CODE = 'USD'
    THEN PSA.AMOUNT_DUE_ORIGINAL
    ELSE ROUND(ROUND(GLDR.CONVERSION_RATE,15) * PSA.AMOUNT_DUE_ORIGINAL,11)
    END AS "USD Amount"
,TO_CHAR(CTA.PRINTING_LAST_PRINTED ,'MM/DD/YYYY HH:MI')  "Invoice Print Date"
,CTA.INTERFACE_HEADER_ATTRIBUTE1 "Reference"
,CTA.COMMENTS "Comments"
,CTA.INTERFACE_HEADER_CONTEXT "Batch Name"
,CTA.REASON_CODE "Reason Code"
,ARM.NAME "Receipt Method"
,HZP.CITY "City"
,HZP.COUNTRY "Country"
,ROUND(CTLA.TAX_AMOUNT,2) "Tax Amount"
,CASE  WHEN PSA.AMOUNT_IN_DISPUTE IS NULL
    THEN 0
    ELSE PSA.AMOUNT_IN_DISPUTE
    END AS "Amount In Dispute"
,CASE WHEN PSA.AMOUNT_IN_DISPUTE IS NULL or PSA.AMOUNT_IN_DISPUTE = '0.00'
    THEN 'NO'
    ELSE 'YES'
    END AS "TRX ON DISPUTE"
,CASE
    WHEN PSA.DISPUTE_DATE IS NULL
    THEN NULL
    ELSE PSA.DISPUTE_DATE
    END AS "Dispute Date"
,HCA.ATTRIBUTE1 "Hotel ID"
,TT.NAME "Term Name"
,FNU.USER_NAME "Created By User"

,CASE
WHEN HCA.CUSTOMER_CLASS_CODE = 'TA'
    THEN 'Travel Ad'
    WHEN HCA.CUSTOMER_CLASS_CODE = 'HOTEL'
            AND CTTA.NAME NOT LIKE '%GROUP%'
            THEN 'Hotel Collect Independent'
    WHEN HCA.CUSTOMER_CLASS_CODE = 'HOTEL'
            AND CTTA.NAME LIKE '%GROUP%'
            THEN 'Hotel Collect Corporate'
    WHEN HCA.CUSTOMER_CLASS_CODE = 'MESODB'
    THEN 'Meso Direct Bill'
    WHEN HCA.CUSTOMER_CLASS_CODE = 'SUPPLIER'
            OR HCA.CUSTOMER_CLASS_CODE = 'OTHER'
            OR CTTA.NAME LIKE '%SO%'
            THEN 'Supplier Other'
    WHEN HCA.CUSTOMER_CLASS_CODE = 'GROUP PARENT'
    THEN 'Hotel Collect Corporate'
    WHEN HCA.CUSTOMER_CLASS_CODE = 'MESOMF'
    THEN 'Meso Marketing Funds'
    WHEN HCA.CUSTOMER_CLASS_CODE = 'ELE'
    THEN 'Expedia Local Expert'
    ELSE 'UNLABLED ITEM - PLEASE CONTACT ICORT'
    END AS "Business Model"

,CASE
WHEN CTTA.NAME LIKE '%RELO%'
    THEN 'RELO'
    WHEN CTTA.NAME LIKE '%VOIDED%'
    THEN 'VOIDED'
    WHEN CTTA.NAME LIKE '%FRAUD%'
    THEN 'FRAUD'
    WHEN CTTA.NAME LIKE '%REBILL%'
    THEN 'REBILL'
    ELSE 'Standard TRX'
    END AS "TRX Type"

,CASE
WHEN ARM.RECEIPT_METHOD_ID IN ('4004','56065','4009','4005','4003','4008','56066','39002','4002','4006','4000','4011','56064','32000','37000','39003','4010','72021','72020','72022')
    THEN 'AP'
    ELSE 'Non AP'
    END AS "AP Status"

,INV.TRX_NUMBER  AS "Applied Invoice Number"
,INV.TRX_DATE    AS "Applied Invoice Date"

FROM AR.RA_CUSTOMER_TRX_ALL CTA

LEFT JOIN AR.HZ_CUST_ACCOUNTS HCA
  ON CTA.BILL_TO_CUSTOMER_ID = HCA.CUST_ACCOUNT_ID

LEFT JOIN AR.HZ_PARTIES HZP
  ON HCA.PARTY_ID = HZP.PARTY_ID

LEFT JOIN APPS.HR_OPERATING_UNITS HOU
  ON HOU.ORGANIZATION_ID = CTA.ORG_ID

LEFT JOIN APPS.GL_LEDGER_LE_V GLL
  ON GLL.LEGAL_ENTITY_ID = HOU.DEFAULT_LEGAL_CONTEXT_ID

LEFT JOIN AR.AR_PAYMENT_SCHEDULES_ALL PSA
  ON PSA.CUSTOMER_TRX_ID = CTA.CUSTOMER_TRX_ID

LEFT JOIN AR.AR_RECEIPT_METHODS ARM
  ON CTA.RECEIPT_METHOD_ID = ARM.RECEIPT_METHOD_ID

LEFT JOIN AR.RA_CUST_TRX_TYPES_ALL CTTA
  ON CTTA.CUST_TRX_TYPE_ID = CTA.CUST_TRX_TYPE_ID

LEFT JOIN APPS.AR_CUSTOMER_PROFILES_V CPV
  ON HCA.CUST_ACCOUNT_ID = CPV.CUSTOMER_ID

LEFT JOIN APPS.FND_USER FNU
  ON FNU.USER_ID = CTA.CREATED_BY

LEFT JOIN (
  SELECT SUM(AR.RA_CUSTOMER_TRX_LINES_ALL.EXTENDED_AMOUNT) "TAX_AMOUNT",
         AR.RA_CUSTOMER_TRX_LINES_ALL.CUSTOMER_TRX_ID,
         AR.RA_CUSTOMER_TRX_LINES_ALL.LINE_TYPE
  FROM AR.RA_CUSTOMER_TRX_LINES_ALL
  WHERE AR.RA_CUSTOMER_TRX_LINES_ALL.LINE_TYPE = 'TAX'
  GROUP BY AR.RA_CUSTOMER_TRX_LINES_ALL.CUSTOMER_TRX_ID,
           AR.RA_CUSTOMER_TRX_LINES_ALL.LINE_TYPE
) CTLA
  ON CTA.CUSTOMER_TRX_ID = CTLA.CUSTOMER_TRX_ID

LEFT JOIN (
  SELECT GL.GL_DAILY_RATES.CONVERSION_RATE,
         GL.GL_DAILY_RATES.FROM_CURRENCY,
         GL.GL_DAILY_RATES.TO_CURRENCY,
         GL.GL_DAILY_RATES.CONVERSION_DATE
  FROM GL.GL_DAILY_RATES
  WHERE GL.GL_DAILY_RATES.TO_CURRENCY = 'USD'
    AND GL.GL_DAILY_RATES.CONVERSION_TYPE = 'Corporate'
) GLDR
  ON (GLDR.FROM_CURRENCY = PSA.INVOICE_CURRENCY_CODE
      AND GLDR.CONVERSION_DATE = PSA.GL_DATE)

LEFT JOIN AR.RA_TERMS_TL TT
  ON PSA.TERM_ID = TT.TERM_ID

LEFT JOIN AR.AR_RECEIVABLE_APPLICATIONS_ALL RAA
  ON RAA.CUSTOMER_TRX_ID = CTA.CUSTOMER_TRX_ID
 AND RAA.APPLICATION_TYPE = 'CM'
 AND RAA.DISPLAY = 'Y'

LEFT JOIN AR.RA_CUSTOMER_TRX_ALL INV
  ON INV.CUSTOMER_TRX_ID = RAA.APPLIED_CUSTOMER_TRX_ID

WHERE
  GLL.LEDGER_CATEGORY_CODE = 'PRIMARY'
  AND CTTA.NAME NOT LIKE '%EAC%'
  AND CTA.TRX_NUMBER IN ({placeholders})
"""

# ==========================
# HELPER FUNCTIONS
# ==========================

def chunk_list(lst: List[str], n: int) -> Generator[List[str], None, None]:
    """Yield successive n-sized chunks from a list."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def find_latest_coa_file() -> Path:
    """Find the most recent 'RPA-306-001 Cash On Account*' file in Downloads."""
    candidates = []
    for f in DOWNLOADS_DIR.iterdir():
        if f.is_file() and f.name.startswith("RPA-306-001 Cash On Account"):
            candidates.append(f)
    if not candidates:
        raise FileNotFoundError("No 'RPA-306-001 Cash On Account*' file found in Downloads.")
    return max(candidates, key=lambda p: p.stat().st_mtime)


def exclude_group_transactions(df: pd.DataFrame, trx_col: str = "Transaction Type") -> pd.DataFrame:
    """Remove rows where transaction type contains 'GROUP'."""
    if df.empty or trx_col not in df.columns:
        return df

    mask_group = df[trx_col].str.contains("GROUP", case=False, na=False)
    group_trx = df.loc[mask_group, "Transaction Number"].unique()

    if len(group_trx) > 0:
        logger.info(f"Excluding {len(group_trx)} transaction(s) with 'GROUP' type from matching")

    return df[~df["Transaction Number"].isin(group_trx)].copy()


def map_operating_unit(trx_type: str) -> str:
    """Map transaction type prefix to Operating Unit name."""
    if not isinstance(trx_type, str):
        return ""

    prefix_map = {
        "CH": "12305 Expedia Lodging Partner Services Sarl",
        "US": "11105 Expedia, Inc.",
        "TS": "14101 Travelscape, LLC",
        "BR": "11115 Expedia do Brasil Agencia"
    }

    for prefix, ou_name in prefix_map.items():
        if trx_type.startswith(prefix):
            return ou_name

    return ""


def map_activity_name(oper_unit: str) -> str:
    """Map Operating Unit to Activity Name."""
    if not isinstance(oper_unit, str):
        return ""

    activity_map = {
        "12305": "CH WRITE-OFF BAD DEBT",
        "14101": "TS WRITE-OFF BAD DEBT",
        "11105": "US WRITE-OFF BAD DEBT",
        "11115": "BR WRITE-OFF BAD DEBT"
    }

    for code, activity in activity_map.items():
        if oper_unit.startswith(code):
            return activity

    return ""


def normalize_reason(reason: str) -> str:
    """Normalize reason codes like 'UNECONOMICAL_TO_COLLECT' -> 'Uneconomical to Collect'."""
    if pd.isna(reason) or not isinstance(reason, str) or not reason.strip():
        return "Uneconomical to Collect"

    reason_clean = reason.strip().upper()

    overrides = {
        "UNECONOMICAL_TO_COLLECT": "Uneconomical to Collect",
        "SMALL_AMT_REMAINING": "Small Amt Remaining",
    }

    if reason_clean in overrides:
        return overrides[reason_clean]

    return reason_clean.replace("_", " ").title()


def extract_invoice_from_reference(ref: str) -> str:
    """
    Remove 'CREDIT INVOICE' prefix and return only the invoice number
    from the Reference field in RELO CMs query.
    """
    if pd.isna(ref):
        return ""

    s = str(ref).strip()
    s = re.sub(r"(?i)^CREDIT\s+INVOICE\s*", "", s).strip()
    m = re.search(r"\d+", s)

    return m.group(0) if m else s


def write_excel_sheet(file_path: Path, df: pd.DataFrame, sheet_name: str) -> None:
    """Safely write DataFrame to Excel sheet, creating file if needed."""
    try:
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
            df.to_excel(w, sheet_name=sheet_name, index=False)
    except FileNotFoundError:
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as w:
            df.to_excel(w, sheet_name=sheet_name, index=False)

    logger.info(f"Sheet '{sheet_name}' written with {len(df)} rows")


# ==========================
# FILE OPERATIONS
# ==========================

def update_concat_column_b(customers: List[str]) -> None:
    """Populate column B in Concat.xlsx with customer list and copy formulas down."""
    wb = load_workbook(CONCAT_PATH)
    ws = wb.active

    for i, cust in enumerate(customers, start=CONCAT_START_ROW):
        ws.cell(row=i, column=2, value=cust)

    max_row = CONCAT_START_ROW + len(customers) - 1

    # Copy formulas from template row down to all customer rows
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=CONCAT_START_ROW, column=col)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            formula = cell.value
            for r in range(CONCAT_START_ROW + 1, max_row + 1):
                if ws.cell(row=r, column=col).value is None:
                    ws.cell(row=r, column=col).value = formula

    wb.save(CONCAT_PATH)
    logger.info(f"Concat.xlsx updated (column B + formulas down to row {max_row})")


def create_output_folder() -> Path:
    """Create today's output folder in MM.DD.YYYY format."""
    today_str = date.today().strftime("%m.%d.%Y")
    output_folder = BASE_OUTPUT_DIR / today_str
    output_folder.mkdir(parents=True, exist_ok=True)
    return output_folder


# ==========================
# ORACLE DATABASE OPERATIONS
# ==========================

def initialize_oracle_client() -> None:
    """Initialize Oracle client library (only call once per process)."""
    try:
        oracledb.init_oracle_client(lib_dir=ORACLE_CLIENT_DIR)
        logger.info("Oracle client initialized successfully")
    except Exception as e:
        logger.error(f"Failed to initialize Oracle client: {e}")
        raise


def query_oracle_in_batches(customers: List[str], query: str, query_name: str) -> Optional[pd.DataFrame]:
    """Execute Oracle query in batches and return concatenated DataFrame."""
    all_rows, columns = [], None
    total_chunks = (len(customers) + BATCH_SIZE - 1) // BATCH_SIZE

    logger.info(f"Connecting to Oracle for {query_name}...")

    try:
        with oracledb.connect(user=USERNAME, password=PASSWORD, dsn=DSN) as conn:
            with conn.cursor() as cur:
                cur.execute("ALTER SESSION SET CURRENT_SCHEMA = APPS")

            for idx, acc_chunk in enumerate(chunk_list(customers, BATCH_SIZE), start=1):
                logger.info(f"[Batch {idx}/{total_chunks}] Processing {len(acc_chunk)} accounts...")

                placeholders = ", ".join(f":acc{i}" for i in range(len(acc_chunk)))
                params = {f"acc{i}": v for i, v in enumerate(acc_chunk)}

                sql = query.format(placeholders=placeholders)

                with conn.cursor() as cur:
                    cur.execute(sql, params)
                    rows = cur.fetchall()

                    if rows:
                        if columns is None:
                            columns = [d[0] for d in cur.description]
                        all_rows.extend(rows)
                        logger.info(f"  -> Retrieved {len(rows)} rows")

    except oracledb.DatabaseError as e:
        error, = e.args
        logger.error(f"Oracle error in {query_name}: {error.message}")
        raise

    if not all_rows:
        logger.warning(f"No rows returned for {query_name}")
        return None

    return pd.DataFrame(all_rows, columns=columns)


def query_cm_applied_data(cm_list: List[str]) -> Optional[pd.DataFrame]:
    """Query CM applied invoice data for given list of CM transaction numbers."""
    if not cm_list:
        logger.warning("No CMs provided for applied invoice query")
        return None

    logger.info(f"Querying applied invoice data for {len(cm_list)} CM(s)...")

    try:
        with oracledb.connect(user=USERNAME, password=PASSWORD, dsn=DSN) as conn:
            placeholders = ", ".join(f":cm{i}" for i in range(len(cm_list)))
            params = {f"cm{i}": v for i, v in enumerate(cm_list)}
            sql = CM_APPLIED_QUERY.format(placeholders=placeholders)

            with conn.cursor() as cur:
                cur.execute(sql, params)
                rows = cur.fetchall()

                if not rows:
                    logger.info("CM_APPLIED_QUERY returned no rows")
                    return None

                columns = [d[0] for d in cur.description]
                return pd.DataFrame(rows, columns=columns)

    except oracledb.DatabaseError as e:
        error, = e.args
        logger.error(f"Oracle error in CM applied query: {error.message}")
        raise


# ==========================
# DATA EXTRACTION & PROCESSING
# ==========================

def extract_customers(coa_path: Path) -> List[str]:
    """Extract unique customer list from COA Data sheet."""
    logger.info(f"Reading COA file: {coa_path}")
    df_coa = pd.read_excel(coa_path, sheet_name=SHEET_DATA)

    if "CUSTOMER_NBR" not in df_coa.columns:
        raise ValueError(f"Column 'CUSTOMER_NBR' not found in '{SHEET_DATA}' sheet")

    customers = (
        df_coa["CUSTOMER_NBR"]
        .dropna()
        .astype(str)
        .str.strip()
        .loc[lambda s: s != ""]
        .unique()
        .tolist()
    )

    logger.info(f"Extracted {len(customers)} unique customers")

    # Write customer list to COA file
    df_cust = pd.DataFrame({"CUSTOMER_NBR": customers})
    write_excel_sheet(coa_path, df_cust, SHEET_CUST_LIST)

    return customers


def prepare_coa_data(coa_path: Path) -> pd.DataFrame:
    """Load and prepare COA data with necessary transformations."""
    df_coa = pd.read_excel(coa_path, sheet_name=SHEET_DATA)

    required_cols = ["RECEIPT_NUMBER", "LOCAL_RECEIPT_AMOUNT", "CUSTOMER_NBR",
                     "CURRENCY_CODE", "RECEIPT_STATUS"]
    missing = [c for c in required_cols if c not in df_coa.columns]

    if missing:
        raise ValueError(f"Missing required columns in {SHEET_DATA}: {missing}")

    df_coa_match = df_coa.assign(
        CUSTOMER_NBR=lambda x: x["CUSTOMER_NBR"].astype(str).str.strip(),
        LOCAL_RECEIPT_AMOUNT=lambda x: pd.to_numeric(x["LOCAL_RECEIPT_AMOUNT"], errors="coerce"),
        CURRENCY_CODE=lambda x: x["CURRENCY_CODE"].astype(str).str.strip(),
        receipt_amount=lambda x: x["LOCAL_RECEIPT_AMOUNT"].round(2)
    )

    # Extract payment date from column 20 (index 19) if available
    if df_coa.shape[1] > WO_DATA_COLUMN_INDEX:
        payment_raw = df_coa.iloc[:, WO_DATA_COLUMN_INDEX]
        df_coa_match["PAYMENT_DATE"] = pd.to_datetime(payment_raw, errors="coerce")
    else:
        df_coa_match["PAYMENT_DATE"] = pd.NaT

    return df_coa_match


def process_wo_matches(
    df_coa_match: pd.DataFrame,
    df_wo: pd.DataFrame,
    coa_path: Path,
    output_folder: Path
) -> Tuple[Optional[pd.DataFrame], Optional[Path], Optional[Path]]:
    """
    Process WO matches and generate output files.
    Returns: (df_matches, cash_xlsx_path, csv_path)
    """
    required_wo_cols = [
        "Transaction Number", "Transaction Type", "Transaction Date",
        "Account Number", "Entered Amount", "Entered Currency", "Reason Code"
    ]
    missing = [c for c in required_wo_cols if c not in df_wo.columns]

    if missing:
        logger.warning(f"Missing WO columns: {missing}. Skipping WO matches.")
        return None, None, None

    # Prepare WO data
    df_wo_m = df_wo.copy()
    df_wo_m["Account Number"] = df_wo_m["Account Number"].astype(str).str.strip()
    df_wo_m["Entered Amount"] = pd.to_numeric(df_wo_m["Entered Amount"], errors="coerce")
    df_wo_m["Entered Currency"] = df_wo_m["Entered Currency"].astype(str).str.strip()
    df_wo_m["Transaction Type"] = df_wo_m["Transaction Type"].astype(str)

    # Exclude GROUP transactions
    df_wo_m = exclude_group_transactions(df_wo_m)

    # Sum by Transaction Number
    df_wo_sum_trx = (
        df_wo_m
        .groupby(["Account Number", "Transaction Number", "Entered Currency"], as_index=False)["Entered Amount"]
        .sum()
    )
    df_wo_sum_trx["sum_amount"] = df_wo_sum_trx["Entered Amount"].round(2)

    # Match COA receipts with WO transactions
    merged_trx = df_coa_match.merge(
        df_wo_sum_trx[["Account Number", "Transaction Number", "Entered Currency", "sum_amount"]],
        left_on=["CUSTOMER_NBR", "CURRENCY_CODE"],
        right_on=["Account Number", "Entered Currency"],
        how="inner",
    )

    cond_trx = merged_trx["receipt_amount"] == (-merged_trx["sum_amount"])
    keys_ok_trx = merged_trx.loc[cond_trx].drop_duplicates()

    if keys_ok_trx.empty:
        logger.info("No WO matches found by Transaction Number")
        return None, None, None

    # Remove duplicate matches (keep first)
    before_len = len(keys_ok_trx)
    keys_ok_trx = (
        keys_ok_trx
        .sort_values(["CUSTOMER_NBR", "CURRENCY_CODE", "RECEIPT_NUMBER", "Transaction Number"])
        .drop_duplicates(subset=["RECEIPT_NUMBER"], keep="first")
    )
    after_len = len(keys_ok_trx)

    if after_len < before_len:
        logger.info(f"Receipts with multiple WO matches found. Reduced from {before_len} to {after_len} rows.")

    # Join back to get full WO details
    matches_wo = df_wo_m.merge(
        keys_ok_trx,
        on=["Account Number", "Transaction Number", "Entered Currency"],
        how="inner",
    )

    # Aggregate matches
    group_cols = [
        "RECEIPT_NUMBER", "RECEIPT_STATUS", "Transaction Number",
        "CURRENCY_CODE", "Entered Currency", "Transaction Type",
        "Account Number", "Reason Code"
    ]

    df_matches_agg = (
        matches_wo
        .groupby(group_cols, as_index=False)
        .agg({
            "receipt_amount": "first",
            "sum_amount": "first",
            "Entered Amount": "sum",
            "RECEIPT_STATUS": "first",
            "Transaction Date": "first",
            "PAYMENT_DATE": "first",
            "Reason Code": "first",
        })
    )

    # Calculate days between payment and invoice
    if "PAYMENT_DATE" in df_matches_agg.columns and "Transaction Date" in df_matches_agg.columns:
        df_matches_agg["Days Between Payment and Invoice"] = (
            (df_matches_agg["PAYMENT_DATE"] - df_matches_agg["Transaction Date"]).dt.days
        )
    else:
        df_matches_agg["Days Between Payment and Invoice"] = pd.NA

    # Split negative days for separate sheet
    neg_mask = df_matches_agg["Days Between Payment and Invoice"] < 0
    df_negative_days = df_matches_agg[neg_mask].copy()

    # Rename columns
    df_matches_all = df_matches_agg.rename(columns={
        "receipt_amount": "LOCAL_RECEIPT_AMOUNT",
        "sum_amount": "WO_SUMIFS",
        "CURRENCY_CODE": "Receipt Currency Code",
        "Entered Currency": "WO Currency",
    })

    # Split DIR vs non-DIR transactions
    dir_mask = df_matches_all["Transaction Type"].str.contains("DIR", case=False, na=False)
    group_mask = df_matches_all["Transaction Type"].str.contains("GROUP", case=False, na=False)

    df_matches = df_matches_all[dir_mask & ~group_mask].copy()
    df_nl_review = df_matches_all[~dir_mask & ~group_mask].copy()

    # Handle UNKNOWN customer matches
    df_unknown_matches = process_unknown_matches(df_coa_match, df_wo_sum_trx)

    # Write all sheets
    write_excel_sheet(coa_path, df_matches, SHEET_MATCHES)

    if not df_nl_review.empty:
        write_excel_sheet(coa_path, df_nl_review, SHEET_NL_REVIEW)

    if df_unknown_matches is not None and not df_unknown_matches.empty:
        write_excel_sheet(coa_path, df_unknown_matches, SHEET_UNKNOWN_MATCHES)

    if not df_negative_days.empty:
        df_negative_days_renamed = df_negative_days.rename(columns={
            "receipt_amount": "LOCAL_RECEIPT_AMOUNT",
            "sum_amount": "WO_SUMIFS",
            "CURRENCY_CODE": "Receipt Currency Code",
            "Entered Currency": "WO Currency",
        })
        write_excel_sheet(coa_path, df_negative_days_renamed, SHEET_NEGATIVE_DAYS)

    # Generate output files
    cash_xlsx_path = generate_cash_file(df_matches, output_folder)
    csv_path = generate_csv_adjustments(df_matches_agg[dir_mask & ~group_mask], output_folder)

    return df_matches, cash_xlsx_path, csv_path


def process_unknown_matches(df_coa_match: pd.DataFrame, df_wo_sum_trx: pd.DataFrame) -> Optional[pd.DataFrame]:
    """Process matches for customers with UNKNOWN account numbers."""
    unknown_mask = df_coa_match["CUSTOMER_NBR"].astype(str).str.upper().isin(["UNKNOWN", "UNKNWON"])
    df_unknown = df_coa_match[unknown_mask].copy()

    if df_unknown.empty:
        return None

    df_unknown_tmp = df_unknown[["RECEIPT_NUMBER", "receipt_amount", "CURRENCY_CODE"]].copy()
    df_unknown_tmp = df_unknown_tmp.rename(columns={"CURRENCY_CODE": "Entered Currency"})

    merged_unknown = df_unknown_tmp.merge(
        df_wo_sum_trx[["Account Number", "Transaction Number", "Entered Currency", "sum_amount"]],
        on="Entered Currency",
        how="inner",
    )

    cond_unknown = merged_unknown["receipt_amount"] == (-merged_unknown["sum_amount"])
    df_unknown_matches = merged_unknown.loc[cond_unknown].drop_duplicates()

    if not df_unknown_matches.empty:
        logger.info(f"Found {len(df_unknown_matches)} matches for UNKNOWN customers")

    return df_unknown_matches if not df_unknown_matches.empty else None


def generate_cash_file(df_matches: pd.DataFrame, output_folder: Path) -> Path:
    """Generate Cash reconciliation Excel file."""
    today_str = date.today().strftime("%m%d%Y")
    cash_xlsx_name = f"COAvsWO_Cash_{today_str}.xlsx"
    cash_xlsx_path = output_folder / cash_xlsx_name

    df_matches.to_excel(cash_xlsx_path, index=False)
    logger.info(f"Cash file generated: {cash_xlsx_path}")

    return cash_xlsx_path


def generate_csv_adjustments(df_matches_agg_dir: pd.DataFrame, output_folder: Path) -> Optional[Path]:
    """Generate CSV file for Oracle adjustments upload."""
    if df_matches_agg_dir.empty:
        logger.info("No DIR transactions found. CSV will not be generated.")
        return None

    today_dt = date.today()
    today_csv_name = today_dt.strftime("%m%d%Y")
    today_comment = today_dt.strftime("%m/%d/%Y")

    df_csv = pd.DataFrame()
    df_csv["Index"] = range(1, len(df_matches_agg_dir) + 1)

    oper_units = df_matches_agg_dir["Transaction Type"].map(map_operating_unit)
    df_csv["Operating Unit"] = oper_units
    df_csv["Transaction Number"] = df_matches_agg_dir["Transaction Number"]
    df_csv["BFB Number"] = ""
    df_csv["Activity Name"] = oper_units.map(map_activity_name)
    df_csv["Adjustment Type"] = "Line"
    df_csv["Amount to be Adjusted"] = (-df_matches_agg_dir["Entered Amount"]).round(2)
    df_csv["Reason"] = df_matches_agg_dir["Reason Code"].apply(normalize_reason)
    df_csv["Comments"] = "COA vs WO " + today_comment
    df_csv["GL Date"] = ""
    df_csv["Adjust Date"] = ""

    # Remove empty transaction numbers
    df_csv = df_csv[
        df_csv["Transaction Number"].notna()
        & (df_csv["Transaction Number"].astype(str).str.strip() != "")
    ].reset_index(drop=True)

    csv_name = f"COAvsWO{today_csv_name}.csv"
    csv_path = output_folder / csv_name
    df_csv.to_csv(csv_path, index=False, encoding="cp1252")

    logger.info(f"CSV adjustments file generated: {csv_path}")
    return csv_path


def process_relo_matches(
    df_coa_match: pd.DataFrame,
    df_relo: Optional[pd.DataFrame],
    coa_path: Path,
    output_folder: Path,
    cash_xlsx_path: Optional[Path]
) -> Tuple[Optional[pd.DataFrame], Optional[Path], Optional[Path]]:
    """
    Process RELO CM matches and generate DM templates.
    Returns: (df_match_relo, relo_dm_path, hc_dm_path)
    """
    relo_dm_path = None
    hc_dm_path = None

    if df_relo is None or df_relo.empty:
        logger.info("No RELO CMs data to process")
        return None, None, None

    required_relo_cols = [
        "Transaction Number", "Transaction Type", "Account Number",
        "Entered Amount", "Entered Currency", "Reference"
    ]
    missing = [c for c in required_relo_cols if c not in df_relo.columns]

    if missing:
        logger.warning(f"Missing RELO columns: {missing}. Skipping RELO processing.")
        return None, None, None

    # Prepare RELO data
    df_relo_m = df_relo.copy()
    df_relo_m["Account Number"] = df_relo_m["Account Number"].astype(str).str.strip()
    df_relo_m["Entered Amount"] = pd.to_numeric(df_relo_m["Entered Amount"], errors="coerce")
    df_relo_m["Entered Currency"] = df_relo_m["Entered Currency"].astype(str).str.strip()

    # Sum by Transaction Number
    df_relo_sum_trx = (
        df_relo_m
        .groupby(["Account Number", "Transaction Number", "Entered Currency"], as_index=False)["Entered Amount"]
        .sum()
    )
    df_relo_sum_trx["sum_amount"] = df_relo_sum_trx["Entered Amount"].round(2)

    # Match COA receipts with RELO CMs
    merged_relo_trx = df_coa_match.merge(
        df_relo_sum_trx[["Account Number", "Transaction Number", "Entered Currency", "sum_amount"]],
        left_on=["CUSTOMER_NBR", "CURRENCY_CODE"],
        right_on=["Account Number", "Entered Currency"],
        how="inner",
    )

    cond_relo_trx = merged_relo_trx["receipt_amount"] == (-merged_relo_trx["sum_amount"])
    keys_relo_trx = merged_relo_trx.loc[cond_relo_trx].drop_duplicates()

    if keys_relo_trx.empty:
        logger.info("No RELO matches found by Transaction Number")
        return None, None, None

    # Join back to get full RELO details
    matches_relo = df_relo_m.merge(
        keys_relo_trx,
        on=["Account Number", "Transaction Number", "Entered Currency"],
        how="inner",
    )

    df_match_relo = matches_relo[[
        "RECEIPT_NUMBER", "receipt_amount", "Transaction Number",
        "Entered Amount", "CURRENCY_CODE", "Entered Currency",
        "Reference", "Account Number", "Transaction Type"
    ]].rename(columns={
        "receipt_amount": "LOCAL_RECEIPT_AMOUNT",
        "CURRENCY_CODE": "Receipt Currency Code",
        "Entered Currency": "WO Currency",
    })

    # Check for already reversed CMs
    try:
        df_match_relo_filtered = check_relo_cms_already_reversed(df_match_relo, coa_path)
    except Exception as e:
        logger.error(f"Failed to check for already reversed CMs: {e}")
        df_match_relo_filtered = df_match_relo

    # Write Match RELO sheet
    write_excel_sheet(coa_path, df_match_relo_filtered, SHEET_MATCH_RELO)

    # Add to Cash file if available
    if cash_xlsx_path is not None and cash_xlsx_path.exists():
        write_excel_sheet(cash_xlsx_path, df_match_relo_filtered, SHEET_MATCH_RELO)

    # Generate DM templates
    try:
        relo_dm_path = build_ts_relo_dm_file(df_match_relo_filtered, RELO_DM_TEMPLATE, output_folder)
        if relo_dm_path:
            logger.info(f"RELO DM file created: {relo_dm_path}")
    except Exception as e:
        logger.error(f"Failed to generate RELO DM file: {e}")

    try:
        hc_dm_path = build_hc_relo_dm_file(df_match_relo_filtered, HC_DM_TEMPLATE, output_folder)
        if hc_dm_path:
            logger.info(f"HC RELO DM file created: {hc_dm_path}")
    except Exception as e:
        logger.error(f"Failed to generate HC RELO DM file: {e}")

    return df_match_relo_filtered, relo_dm_path, hc_dm_path


def check_relo_cms_already_reversed(df_match_relo: pd.DataFrame, coa_path: Path) -> pd.DataFrame:
    """
    Check if RELO CMs have already been reversed by comparing Reference vs Applied Invoice Number.
    Returns filtered dataframe with reversed CMs removed.
    """
    if df_match_relo is None or df_match_relo.empty:
        logger.info("Match RELO empty. Nothing to check for reversals.")
        return df_match_relo

    cm_list = (
        df_match_relo["Transaction Number"]
        .dropna()
        .astype(str)
        .str.strip()
        .unique()
        .tolist()
    )

    if not cm_list:
        logger.info("No CMs found in Match RELO")
        return df_match_relo

    # Query for applied invoice data
    df_cm = query_cm_applied_data(cm_list)

    if df_cm is None or df_cm.empty:
        return df_match_relo

    # Prepare data for comparison
    df_cm["Transaction Number"] = df_cm["Transaction Number"].astype(str).str.strip()
    df_cm["Applied Invoice Number"] = df_cm["Applied Invoice Number"].astype(str).str.strip()

    df_ref = df_match_relo.copy()
    df_ref["Transaction Number"] = df_ref["Transaction Number"].astype(str).str.strip()
    df_ref["Reference_Invoice"] = (
        df_ref["Reference"]
        .apply(extract_invoice_from_reference)
        .astype(str)
        .str.strip()
    )

    # Get unique applied invoice per CM
    df_applied_small = (
        df_cm[["Transaction Number", "Applied Invoice Number"]]
        .drop_duplicates(subset=["Transaction Number"])
        .copy()
    )

    # Merge and compare
    df_check = df_ref.merge(
        df_applied_small,
        on="Transaction Number",
        how="left",
    )

    mask_has_applied = df_check["Applied Invoice Number"].notna() & (
        df_check["Applied Invoice Number"].astype(str).str.strip() != ""
    )
    mask_diff = (
        df_check["Reference_Invoice"].astype(str).str.strip()
        != df_check["Applied Invoice Number"].astype(str).str.strip()
    )

    df_reversed = df_check[mask_has_applied & mask_diff].copy()

    if df_reversed.empty:
        logger.info("No CMs identified as already reversed")
        return df_match_relo

    # Write reversed CMs to separate sheet
    df_reversed["Status"] = "CMs Already Reversed"
    write_excel_sheet(coa_path, df_reversed, SHEET_CMS_REVERSED)

    logger.info(
        f"Found {len(df_reversed)} CM(s) already reversed "
        "(Reference != Applied Invoice Number)"
    )

    # Filter out reversed CMs
    reversed_trx_numbers = df_reversed["Transaction Number"].unique().tolist()
    df_filtered = df_match_relo[
        ~df_match_relo["Transaction Number"].astype(str).str.strip().isin(reversed_trx_numbers)
    ].copy()

    logger.info(f"Removed {len(df_match_relo) - len(df_filtered)} rows from Match RELO (already reversed)")

    return df_filtered


# ==========================
# TEMPLATE FILE GENERATION
# ==========================

def find_col_letter_by_header(ws, header_text: str, search_rows=range(1, 20)) -> str:
    """Search for header text in first rows and return column letter."""
    header_text = str(header_text).strip().lower()

    for r in search_rows:
        for cell in ws[r]:
            val = cell.value
            if val is None:
                continue
            if str(val).strip().lower() == header_text:
                return cell.column_letter

    raise KeyError(f"Header '{header_text}' not found in template")


def copy_default_from_row(ws, col_letter: str, from_row: int, to_row: int) -> None:
    """Copy cell value from from_row to to_row for given column letter."""
    src = ws[f"{col_letter}{from_row}"]
    dst = ws[f"{col_letter}{to_row}"]
    dst.value = src.value


def build_ts_relo_dm_file(
    df_match_relo: pd.DataFrame,
    template_path: Path,
    out_dir: Path
) -> Optional[Path]:
    """
    Generate TS RELO DM .xlsm file from template.
    Transforms TS_CM_RELO transactions into TS_DM_RELO debit memos.
    """
    if df_match_relo is None or df_match_relo.empty:
        logger.info("Match RELO empty. No TS RELO DM file to generate.")
        return None

    df_src = (
        df_match_relo
        .loc[lambda d: d["Transaction Type"].astype(str).str.upper() == "TS_CM_RELO"]
        .copy()
    )

    if df_src.empty:
        logger.info("No TS_CM_RELO transactions. No TS RELO DM file to generate.")
        return None

    df_src = (
        df_src
        .drop_duplicates(subset=["RECEIPT_NUMBER", "Transaction Number", "Account Number"])
        .loc[:, [
            "RECEIPT_NUMBER",
            "Transaction Number",
            "Account Number",
            "LOCAL_RECEIPT_AMOUNT",
            "WO Currency",
        ]]
        .copy()
    )
    df_src["AMOUNT_POS"] = df_src["LOCAL_RECEIPT_AMOUNT"].abs().round(2)

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(template_path, keep_vba=True, data_only=False)

    if SHEET_EVC_RELO_CM not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_EVC_RELO_CM}' not found in template")

    ws = wb[SHEET_EVC_RELO_CM]

    # Find column letters by header
    col_customer = find_col_letter_by_header(ws, "Customer Number")
    col_shipto = find_col_letter_by_header(ws, "Ship To Cust Numb")
    col_trx_date = find_col_letter_by_header(ws, "Transaction Date")
    col_txn_type = find_col_letter_by_header(ws, "Txn Type")
    col_comments = find_col_letter_by_header(ws, "Comments")
    col_reference = find_col_letter_by_header(ws, "Reference")
    col_context = find_col_letter_by_header(ws, "Context")
    col_line_num = find_col_letter_by_header(ws, "Line Number")

    try:
        col_currency = find_col_letter_by_header(ws, "Currency")
    except KeyError:
        col_currency = None

    col_uom = "AM"
    col_unit_price = "AN"
    col_total_amt = "AP"

    preserve_letters = ["X", "AH", "AI", "AJ", "AK", "AQ", "AR", "AS"]

    start_row = 11
    today_dt = date.today()

    for i, (_, rec) in enumerate(df_src.iterrows(), start=0):
        trx_number = rec["Transaction Number"]
        account_number = rec["Account Number"]
        amount_pos = float(rec["AMOUNT_POS"])
        wo_currency = str(rec["WO Currency"]).strip() if pd.notna(rec["WO Currency"]) else ""

        r = start_row + i

        ws[f"{col_customer}{r}"] = str(account_number)
        ws[f"{col_shipto}{r}"] = str(account_number)
        ws[f"{col_trx_date}{r}"] = today_dt

        if col_currency is not None:
            ws[f"{col_currency}{r}"] = wo_currency

        ws[f"{col_txn_type}{r}"] = "TS_DM_RELO"
        ws[f"P{r}"] = "TS_DM_RELO"

        ws[f"{col_comments}{r}"] = '="COA vs WO " & TEXT(TODAY(),"mm/dd/yyyy")'
        ws[f"{col_reference}{r}"] = f"DM ISSUED TO OFFSET CM {trx_number}"

        for colL in preserve_letters:
            copy_default_from_row(ws, colL, TEMPLATE_ROW_RELO, r)

        copy_default_from_row(ws, col_context, TEMPLATE_ROW_RELO, r)

        ws[f"{col_uom}{r}"] = "EA"
        ws[f"{col_unit_price}{r}"] = amount_pos
        ws[f"{col_total_amt}{r}"] = amount_pos

        if col_line_num != "AH":
            ws[f"{col_line_num}{r}"] = i + 1

    out_name = f"4 Non-Lodging Relo DM_{today_dt.strftime('%m%d%Y')}.xlsm"
    out_path = out_dir / out_name
    wb.save(out_path)

    logger.info(f"TS RELO DM file generated: {out_path}")
    return out_path


def build_hc_relo_dm_file(
    df_match_relo: pd.DataFrame,
    template_path: Path,
    out_dir: Path
) -> Optional[Path]:
    """
    Generate HC RELO DM .xlsm file from template.
    Processes non-TS_CM_RELO transactions (US, CH entities).
    """
    if df_match_relo is None or df_match_relo.empty:
        logger.info("Match RELO empty. No HC RELO DM file to generate.")
        return None

    df_src = (
        df_match_relo
        .loc[lambda d: d["Transaction Type"].astype(str).str.upper() != "TS_CM_RELO"]
        .copy()
    )

    if df_src.empty:
        logger.info("No non-TS_CM_RELO transactions. No HC RELO DM file to generate.")
        return None

    df_src = (
        df_src
        .drop_duplicates(subset=["RECEIPT_NUMBER", "Transaction Number", "Account Number"])
        .loc[:, [
            "RECEIPT_NUMBER",
            "Transaction Number",
            "Account Number",
            "LOCAL_RECEIPT_AMOUNT",
            "WO Currency",
            "Transaction Type",
        ]]
        .copy()
    )

    df_src["AMOUNT_POS"] = df_src["LOCAL_RECEIPT_AMOUNT"].abs().round(2)
    df_src["WO Currency"] = df_src["WO Currency"].astype(str).str.strip()

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(template_path, keep_vba=True, data_only=False)
    today_dt = date.today()
    sheet_cfg = {}

    for _, rec in df_src.iterrows():
        trx_type = str(rec["Transaction Type"] or "")
        trx_type_up = trx_type.upper()

        # Determine entity and new transaction type
        if trx_type_up.startswith("US"):
            entity_key = ENTITY_US
            new_trx_type = "US_DIR_DM_RELO_USD"
        elif trx_type_up.startswith("CH"):
            entity_key = ENTITY_CH
            new_trx_type = "CH_DIR_DM_RELO_USD"
        elif trx_type_up.startswith("TS"):
            entity_key = ENTITY_TS
            new_trx_type = "TS_DIR_DM_RELO_USD"
        else:
            logger.warning(f"Unknown transaction type prefix: {trx_type}")
            continue

        # Initialize sheet configuration for this entity
        if entity_key not in sheet_cfg:
            target_ws = None
            for name in wb.sheetnames:
                if name.upper().startswith(entity_key):
                    target_ws = wb[name]
                    break

            if target_ws is None:
                raise ValueError(f"Sheet for entity '{entity_key}' not found in template")

            ws = target_ws
            next_row = ws.max_row + 1

            def safe_find(header):
                try:
                    return find_col_letter_by_header(ws, header)
                except KeyError:
                    return None

            cfg = {
                "ws": ws,
                "next_row": next_row,
                "col_trx_date": safe_find("Transaction Date"),
                "col_gl_date": safe_find("GL Date"),
                "col_currency": safe_find("Currency Code"),
                "col_comments": safe_find("Comments"),
                "col_line_number": safe_find("Line Number"),
                "col_line_type": safe_find("Line Type"),
                "col_item": safe_find("Item"),
                "col_description": safe_find("Description"),
                "col_quantity": safe_find("Quantity"),
            }
            sheet_cfg[entity_key] = cfg

        cfg = sheet_cfg[entity_key]
        ws = cfg["ws"]
        r = cfg["next_row"]
        cfg["next_row"] += 1

        account_number = str(rec["Account Number"])
        trx_number = str(rec["Transaction Number"])
        amount_pos = float(rec["AMOUNT_POS"])
        currency = str(rec["WO Currency"])

        # Populate row
        ws[f"I{r}"] = account_number
        ws[f"L{r}"] = account_number

        if cfg["col_trx_date"]:
            ws[f"{cfg['col_trx_date']}{r}"] = today_dt
        if cfg["col_gl_date"]:
            ws[f"{cfg['col_gl_date']}{r}"] = today_dt

        ws[f"P{r}"] = new_trx_type

        if cfg["col_comments"]:
            ws[f"{cfg['col_comments']}{r}"] = (
                f'="DM issued to offset Relocation CM - Coa vs WO " & '
                f'TEXT(TODAY(),"mm/dd/yyyy") & " " & "{trx_number}"'
            )

        ws[f"W{r}"] = "Do Not Print"

        if cfg["col_currency"]:
            ws[f"{cfg['col_currency']}{r}"] = currency

        ws[f"AC{r}"] = "DIRECT AGENCY"
        ws[f"AD{r}"] = currency
        ws[f"AG{r}"] = "INDEPENDENTS"
        ws[f"AI{r}"] = "More4Apps"
        ws[f"AJ{r}"] = f"DM Issued to offset CM {trx_number}"

        if cfg["col_line_number"]:
            ws[f"{cfg['col_line_number']}{r}"] = 1
        if cfg["col_line_type"]:
            ws[f"{cfg['col_line_type']}{r}"] = "Line"
        if cfg["col_item"]:
            ws[f"{cfg['col_item']}{r}"] = "EXPWA_COMP_CM"
        if cfg["col_description"]:
            ws[f"{cfg['col_description']}{r}"] = "COMPENSATION CREDIT"
        if cfg["col_quantity"]:
            ws[f"{cfg['col_quantity']}{r}"] = 1

        ws[f"AP{r}"] = "EA"
        ws[f"AQ{r}"] = amount_pos
        ws[f"AS{r}"] = amount_pos

        ws[f"AT{r}"] = "DIRECT AGENCY"
        ws[f"AU{r}"] = "1097"
        ws[f"AV{r}"] = trx_number

        ws[f"AW{r}"] = currency
        ws[f"AX{r}"] = amount_pos

        ws[f"AZ{r}"] = today_dt
        ws[f"BA{r}"] = today_dt
        ws[f"BC{r}"] = today_dt

        ws[f"BD{r}"] = "REQUESTED"
        ws[f"BE{r}"] = "EG_INVOICING"
        ws[f"BF{r}"] = "DIR"

    out_name = f"3 HC Relo Upload DM_{today_dt.strftime('%m%d%Y')}.xlsm"
    out_path = out_dir / out_name
    wb.save(out_path)

    logger.info(f"HC RELO DM file generated: {out_path}")
    return out_path


# ==========================
# EMAIL OPERATIONS
# ==========================

def send_outlook_email_with_attachments(
    to_addr: str,
    subject: str,
    body: str,
    attachments: List[Path]
) -> None:
    """Send email via Outlook with attachments."""
    if not OUTLOOK_AVAILABLE:
        logger.warning("pywin32 not installed. Cannot send Outlook email.")
        return

    try:
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = to_addr
        mail.Subject = subject
        mail.Body = body

        for att in attachments:
            att_path = str(att)
            if os.path.isfile(att_path):
                mail.Attachments.Add(att_path)
            else:
                logger.warning(f"Attachment not found: {att_path}")

        mail.Send()
        logger.info(f"Email sent to {to_addr} with {len(attachments)} attachment(s)")

    except Exception as e:
        logger.error(f"Failed to send Outlook email: {e}")


def send_summary_email(output_files: OutputFiles) -> None:
    """Send summary email with all generated output files."""
    attachments = output_files.get_attachments()

    if not attachments:
        logger.info("No attachments to send. Skipping email.")
        return

    today_comment = date.today().strftime("%m/%d/%Y")
    email_subject = f"COA vs WO {today_comment}"
    email_body = (
        f"Hi team,\n\n"
        f"Please see attached the COA vs WO outputs for {today_comment}.\n\n"
        f"Best regards,\n"
        f"Global Billing\n"
    )

    send_outlook_email_with_attachments(EMAIL_TO, email_subject, email_body, attachments)


# ==========================
# MAIN PIPELINE
# ==========================

def main() -> None:
    """Main execution pipeline for COA vs WO reconciliation."""
    # Initialize variables
    cash_xlsx_path = None
    csv_path = None
    relo_dm_path = None
    hc_dm_path = None

    try:
        logger.info("=" * 60)
        logger.info("COA vs WO Reconciliation - Starting")
        logger.info("=" * 60)

        # Step 1: Find and validate COA file
        coa_path = find_latest_coa_file()
        output_folder = create_output_folder()
        logger.info(f"Output folder: {output_folder}")

        # Step 2: Extract customers and update Concat
        customers = extract_customers(coa_path)
        update_concat_column_b(customers)

        # Step 3: Initialize Oracle and query data
        initialize_oracle_client()

        df_wo = query_oracle_in_batches(customers, BASE_QUERY, "WO Adjustments")
        df_relo = query_oracle_in_batches(customers, RELO_QUERY, "RELO CMs")

        if df_wo is None or df_wo.empty:
            logger.error("No WO data returned. Exiting.")
            return

        # Write raw Oracle data to COA file
        write_excel_sheet(coa_path, df_wo, SHEET_WO)

        if df_relo is not None and not df_relo.empty:
            write_excel_sheet(coa_path, df_relo, SHEET_RELO_CMS)

        # Step 4: Prepare COA data for matching
        df_coa_match = prepare_coa_data(coa_path)

        # Step 5: Process WO matches
        df_matches, cash_xlsx_path, csv_path = process_wo_matches(
            df_coa_match, df_wo, coa_path, output_folder
        )

        # Step 6: Process RELO matches
        df_match_relo, relo_dm_path, hc_dm_path = process_relo_matches(
            df_coa_match, df_relo, coa_path, output_folder, cash_xlsx_path
        )

        # Step 7: Send summary email
        output_files = OutputFiles(
            coa_workbook=coa_path,
            cash_xlsx=cash_xlsx_path,
            csv_adjustments=csv_path,
            relo_dm=relo_dm_path,
            hc_dm=hc_dm_path
        )

        send_summary_email(output_files)

        logger.info("=" * 60)
        logger.info("COA vs WO Reconciliation - Completed Successfully")
        logger.info("=" * 60)

    except FileNotFoundError as e:
        logger.error(f"Required file not found: {e}")
        sys.exit(1)

    except ValueError as e:
        logger.error(f"Data validation error: {e}")
        sys.exit(1)

    except oracledb.DatabaseError as e:
        logger.error(f"Oracle database error: {e}")
        sys.exit(1)

    except Exception as e:
        logger.exception(f"Unexpected error occurred: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
